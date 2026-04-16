#!/usr/bin/env python3



"""



SimplificaÊ Web App - Prototipo MVP



Conversor de Agenda em Lote -> Cotacao + Selecao de URs



"""



import os

import sys



import csv



import json



import shutil



import tempfile



import zipfile



import time
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders as _email_encoders



from io import BytesIO, StringIO



from datetime import datetime



from collections import OrderedDict, defaultdict



from flask import Flask, request, render_template_string, send_file, jsonify, session, redirect



try:



    from dotenv import load_dotenv



    load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env'))



except ImportError:



    pass



try:



    import requests as http_requests



except ImportError:



    http_requests = None



try:



    import openpyxl



    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side



    from openpyxl.utils import get_column_letter



except ImportError:



    print("ERRO: pip install openpyxl")



    exit(1)



APP_VERSION = '22'



app = Flask(__name__)



app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'simplificae-ar-2026')



# Databricks config



DATABRICKS_HOST = os.environ.get('DATABRICKS_HOST', 'https://picpay-principal.cloud.databricks.com')



DATABRICKS_TOKEN = os.environ.get('DATABRICKS_TOKEN', '')



DATABRICKS_WAREHOUSE_ID = os.environ.get('DATABRICKS_WAREHOUSE_ID', '3b94f0935afb32db')



# OAuth U2M config



OAUTH_CLIENT_ID = 'databricks-cli'  # Public client built-in do Databricks



OAUTH_REDIRECT_URI = 'http://localhost:5000/oauth/callback'



OAUTH_AUTHORIZE_URL = f'{DATABRICKS_HOST}/oidc/v1/authorize'



OAUTH_TOKEN_URL = f'{DATABRICKS_HOST}/oidc/v1/token'



OAUTH_SCOPES = 'all-apis offline_access'



OAUTH_TOKEN_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.oauth_token.json')



import hashlib



import base64



import secrets as py_secrets

# Telemetria de uso — importação silenciosa, nunca bloqueia o app
try:
    import usage_log as _usage_log
    _TELEMETRIA_OK = True
except ImportError:
    _TELEMETRIA_OK = False

# Auto-update — importação silenciosa
try:
    import updater as _updater
    _UPDATER_OK = True
except ImportError:
    _UPDATER_OK = False



# Paths



BASE_DIR = os.path.dirname(os.path.abspath(__file__))



RAIZES_PATH = os.path.join(BASE_DIR, 'raizes_conhecidas.json')

# Quando rodando como .exe (PyInstaller frozen), BASE_DIR aponta para
# _internal/ que é temporário e some entre execuções do processo.
# ~/SimplificaE_data é o diretório estável para uploads, outputs e dados.
if getattr(sys, 'frozen', False):
    _DATA_DIR = os.path.join(os.path.expanduser('~'), 'SimplificaE_data')
else:
    _DATA_DIR = BASE_DIR

UPLOAD_DIR = os.path.join(_DATA_DIR, 'uploads')



OUTPUT_DIR = os.path.join(_DATA_DIR, 'output')



os.makedirs(UPLOAD_DIR, exist_ok=True)



os.makedirs(OUTPUT_DIR, exist_ok=True)



# ==============================================================================



# ESTILOS VISUAIS



# ==============================================================================



GREEN_DARK = '1B5E20'



GREEN_MID = '2E7D32'



GREEN_LIGHT = 'C8E6C9'



GREEN_LIGHTER = 'E8F5E9'



GREEN_ACCENT = '4CAF50'



WHITE_HEX = 'FFFFFF'



GRAY_BORDER = 'BDBDBD'



GRAY_TEXT = '616161'



fill_header_dark = PatternFill(start_color=GREEN_DARK, end_color=GREEN_DARK, fill_type='solid')



fill_header_mid = PatternFill(start_color=GREEN_MID, end_color=GREEN_MID, fill_type='solid')



fill_green_light = PatternFill(start_color=GREEN_LIGHT, end_color=GREEN_LIGHT, fill_type='solid')



fill_green_lighter = PatternFill(start_color=GREEN_LIGHTER, end_color=GREEN_LIGHTER, fill_type='solid')



fill_white = PatternFill(start_color=WHITE_HEX, end_color=WHITE_HEX, fill_type='solid')



fill_totals = PatternFill(start_color='A5D6A7', end_color='A5D6A7', fill_type='solid')



fill_inelig = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')  # vermelho claro



font_inelig = Font(color='B71C1C', size=8)  # vermelho escuro
fill_missing = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
font_missing = Font(color='000000', size=8)



thin_border = Border(



    left=Side(style='thin', color=GRAY_BORDER),



    right=Side(style='thin', color=GRAY_BORDER),



    top=Side(style='thin', color=GRAY_BORDER),



    bottom=Side(style='thin', color=GRAY_BORDER)



)



font_title = Font(bold=True, size=16, color=GREEN_DARK)



font_subtitle = Font(bold=True, size=11, color=GREEN_MID)



font_header = Font(bold=True, size=10, color=WHITE_HEX)



font_data = Font(size=10, color='212121')



font_totals = Font(bold=True, size=11, color=GREEN_DARK)



font_premissa_label = Font(size=10, color=GRAY_TEXT)



font_premissa_value = Font(bold=True, size=10, color=GREEN_DARK)



align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)



align_left = Alignment(horizontal='left', vertical='center')



align_right = Alignment(horizontal='right', vertical='center')



# Formatos



fmt_brl = 'R$ #,##0.00'



fmt_pct = '0.0000%'



fmt_pct_long = '0.0000000000%'



fmt_pct_short = '0.00%'



ARRANJO_BANDEIRA = {'ECC': 'Elo', 'VCC': 'Visa', 'ACC': 'Amex', 'MCC': 'Master'}



ADQUIRENTES_ORDER = ['PICPAY', 'Cielo', 'Rede', 'Stone', 'Getnet']



OPERADORES = {



    'Cesar': 'cesar.oda@picpay.com',



    'Deyvis': 'deyvis.balconi@picpay.com',



}



# ==============================================================================



# UTILS



# ==============================================================================



def normalize_cnpj(cnpj):



    if not cnpj:



        return ''



    return cnpj.strip().replace('.', '').replace('/', '').replace('-', '').zfill(14)



def get_root(cnpj):



    return normalize_cnpj(cnpj)[:8]



def parse_number(val):



    if not val or val.strip() == '':



        return 0.0



    val = val.strip()



    # Detectar formato: se tem virgula E ponto, virgula eh decimal (BR)



    if ',' in val and '.' in val:



        val = val.replace('.', '').replace(',', '.')



    elif ',' in val and '.' not in val:



        val = val.replace(',', '.')



    try:



        return float(val)



    except ValueError:



        return 0.0



def parse_date(date_str):



    if not date_str:



        return None



    date_str = date_str.strip()



    for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d/%m/%Y %H:%M', '%Y-%m-%d %H:%M:%S']:



        try:



            return datetime.strptime(date_str, fmt)



        except ValueError:



            continue



    return None



def load_raizes():



    if os.path.exists(RAIZES_PATH):



        with open(RAIZES_PATH, 'r', encoding='utf-8-sig') as f:



            return json.load(f)



    return {}



def apply_zebra(ws, row, col_start, col_end, is_even):



    # Otimizado: so aplica fill em linhas pares, sem border (performance)



    if is_even:



        fill = fill_green_lighter



        for col in range(col_start, col_end + 1):



            ws.cell(row=row, column=col).fill = fill



def set_header_row(ws, row, headers, col_start=1, fill=None, font_h=None):



    if fill is None:



        fill = fill_header_dark



    if font_h is None:



        font_h = font_header



    for i, h in enumerate(headers):



        cell = ws.cell(row=row, column=col_start + i, value=h)



        cell.font = font_h



        cell.fill = fill



        cell.alignment = align_center



        cell.border = thin_border



# ==============================================================================



# PARSE AGENDA



# ==============================================================================



def parse_agenda(filepath):



    records = []



    for encoding in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:



        try:



            with open(filepath, 'r', encoding=encoding) as f:



                first_line = f.readline()



                f.seek(0)



                sep = ';' if ';' in first_line else ','



                reader = csv.DictReader(f, delimiter=sep)



                for row in reader:



                    n = {}



                    for k, v in row.items():



                        key = k.strip().lower()



                        val = v.strip() if v else ''



                        if 'receivable_id' in key:



                            n['receivable_id'] = val



                        elif 'adquirente' in key and 'cnpj' in key:



                            n['cnpj_adquirente'] = normalize_cnpj(val)



                        elif 'documento' in key or ('cnpj' in key and 'adquirente' not in key):



                            n['cnpj'] = val



                        elif 'nome adquirente' in key:



                            n['adquirente'] = val



                        elif 'arranjo' in key:



                            n['arranjo'] = val.upper()



                        elif ('liquidac' in key or 'liquida' in key) and 'data' in key:



                            n['data_liquidacao'] = val



                        elif 'atualizac' in key:



                            n['data_atualizacao'] = val



                        elif 'constitu' in key:



                            n['agenda_total'] = parse_number(val)



                        elif 'comprometida' in key:



                            n['agenda_comprometida'] = parse_number(val)



                        elif 'livre' in key:



                            n['agenda_livre'] = parse_number(val)



                        elif 'antecipar' in key or 'dispon' in key:



                            n['disponivel'] = parse_number(val)



                        elif 'buffer' in key:



                            n['buffer'] = parse_number(val)



                    if 'cnpj' in n and 'receivable_id' in n:



                        n['cnpj_original'] = n['cnpj']



                        n['cnpj'] = normalize_cnpj(n['cnpj'])



                        n['raiz'] = get_root(n['cnpj'])



                        records.append(n)



                if records:



                    break



        except (UnicodeDecodeError, KeyError):



            records = []



    return records



# ==============================================================================



# ANALYZE



# ==============================================================================



def analyze_records(records, raiz_map):



    empresas = {}



    for r in records:



        raiz = r.get('raiz', '')



        emp_name = raiz_map.get(raiz, f'DESCONHECIDA ({raiz})')



        if emp_name not in empresas:



            empresas[emp_name] = {



                'raizes': set(), 'cnpjs': set(), 'urs': 0, 'valor': 0.0,



                'adquirentes': set(), 'arranjos': set(), 'datas': set()



            }



        e = empresas[emp_name]



        e['raizes'].add(raiz)



        e['cnpjs'].add(r.get('cnpj', ''))



        e['urs'] += 1



        e['valor'] += r.get('disponivel', 0)



        e['adquirentes'].add(r.get('adquirente', ''))



        e['arranjos'].add(r.get('arranjo', ''))



        ds = r.get('data_liquidacao', '')



        if ds:



            e['datas'].add(ds)



    return empresas



# ==============================================================================



# GENERATE COTACAO



# ==============================================================================



# -*- coding: utf-8 -*-

"""

_generate_cotacao_xlsxwriter.py

Reimplementacao de generate_cotacao usando xlsxwriter.

Mesmas 4 abas, mesma estrutura, mesma formatacao.

Testado isoladamente antes de inserir no webapp.

"""

import xlsxwriter

from datetime import datetime

from collections import OrderedDict



# ─────────────────────────────────────────────────────────────────────────────
# CALCULADORA AR — leitura de curvas/feriados e cálculo de indicadores
# ─────────────────────────────────────────────────────────────────────────────
_ar_curvas_cache = None
_ar_curvas_cache_ts = None
_ar_cerc_cache      = None  # {receivable_id: custo_carrego_r$}
_ar_cerc_cache_ts   = None
_ar_cerc_tarifa_faixas = []  # [(qtd_min, qtd_max, tarifa_r$)]
# Planilha de curvas: busca em Desktop/SimplificaE primeiro, depois junto ao script
_AR_PLANILHA = next(
    (p for p in [
        os.path.join(os.path.expanduser('~'), 'Desktop', 'SimplificaE', 'calculadora_ar.xlsx'),
        os.path.join(os.path.dirname(os.path.abspath(__file__)), 'calculadora_ar.xlsx'),
        os.path.join(os.path.expanduser('~'), 'SimplificaE_data', 'calculadora_ar.xlsx'),
    ] if os.path.exists(p)),
    os.path.join(os.path.expanduser('~'), 'Desktop', 'SimplificaE', 'calculadora_ar.xlsx')
)
_AR_CDI_DEFAULT = 0.1465  # 14.65% a.a. — fallback se planilha indisponível


def carregar_curvas_ar(forcar=False):
    """
    Lê as abas 'Curva Mensal' e 'Feriados' da calculadora_ar.xlsx.
    Retorna (curvas_dict, feriados_set, aviso) onde:
      curvas_dict  : {prazo_du (int): {'cdi_du': float, 'cof_am': float}}
      feriados_set : set de datetime.date
      aviso        : str ou None
    Cacheia por 12 horas.
    """
    global _ar_curvas_cache, _ar_curvas_cache_ts
    agora = datetime.now()
    if (not forcar and _ar_curvas_cache is not None and _ar_curvas_cache_ts is not None
            and (agora - _ar_curvas_cache_ts).total_seconds() < 43200):
        return _ar_curvas_cache

    curvas_dict = {}
    feriados_set = set()
    aviso = None

    try:
        if not os.path.exists(_AR_PLANILHA):
            raise FileNotFoundError('calculadora_ar.xlsx não encontrada')
        import openpyxl
        wb = openpyxl.load_workbook(_AR_PLANILHA, read_only=True, data_only=True)
        # Curva Mensal: A=data, B=dias corridos, C=dias úteis, D=rate aa, E=COF%CDI, F=cdim, G=CDI base
        ws_curva = wb['Curva Mensal']
        for row in ws_curva.iter_rows(min_row=2, values_only=True):
            if row[2] is None:
                continue
            try:
                prazo_du    = int(row[2])
                cof_am      = float(row[4]) if row[4] is not None else 0.012   # col E — COF a.m. (já decimal)
                cdi_base_aa = float(row[6]) if row[6] is not None else _AR_CDI_DEFAULT  # col G — CDI a.a. (já decimal)
                if prazo_du > 0:
                    curvas_dict[prazo_du] = {'cof_am': cof_am, 'cdi_base_aa': cdi_base_aa}
            except (TypeError, ValueError):
                continue
        # Aba Curvas: date, calendarDays, businessDays, rate(% aa), pcdi(%), cdim, cdi(% aa)
        # rate_aa: taxa de spread em % a.a. — usada para calcular Receita Bruta via (1+rate_aa)^(dc/360)-1
        try:
            ws_curvas = wb['Curvas']
            for row_c in ws_curvas.iter_rows(min_row=2, values_only=True):
                if row_c[2] is None: continue
                try:
                    du_c = int(float(str(row_c[2])))
                    rate_aa_pct = float(str(row_c[3])) / 100.0  # converte % para decimal
                    if du_c > 0 and du_c in curvas_dict:
                        curvas_dict[du_c]['rate_aa'] = rate_aa_pct
                except (TypeError, ValueError): continue
        except Exception: pass  # aba Curvas ausente — rate_aa fica sem valor
        # Feriados: A=data
        ws_fer = wb['Feriados']
        for row in ws_fer.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            try:
                feriados_set.add(row[0].date() if hasattr(row[0], 'date') else row[0])
            except Exception:
                continue
        wb.close()
    except Exception as e:
        aviso = f'Curvas AR indisponíveis ({e}). Usando CDI default {_AR_CDI_DEFAULT*100:.2f}% a.a.'
        cdi_diario = (1 + _AR_CDI_DEFAULT) ** (1 / 252) - 1
        for du in range(1, 365):
            cdi_acum = (1 + cdi_diario) ** du - 1
            curvas_dict[du] = {'cdi_du': cdi_acum, 'cof_am': cdi_acum * 0.85}

    _ar_curvas_cache = (curvas_dict, feriados_set, aviso)
    _ar_curvas_cache_ts = agora
    return _ar_curvas_cache



def carregar_cerc_ar(forcar=False):
    """
    Lê a aba 'Custos CERC' da calculadora_ar.xlsx.
    Retorna (cerc_map, tarifa_faixas, taxa_registro_bps) onde:
      cerc_map        : {receivable_id: custo_carrego_r$} — custo de carrego por UR
      tarifa_faixas   : [(qtd_min, qtd_max, tarifa_r$)] — tarifa por faixa de qtd URs
      taxa_registro_bps: float — taxa de registro em bps (col F linha 43 / volume total)
    Cacheia por 12 horas.
    """
    global _ar_cerc_cache, _ar_cerc_cache_ts, _ar_cerc_tarifa_faixas
    agora = datetime.now()
    if (not forcar and _ar_cerc_cache is not None and _ar_cerc_cache_ts is not None
            and (agora - _ar_cerc_cache_ts).total_seconds() < 43200):
        return _ar_cerc_cache

    cerc_map = {}
    tarifa_faixas = []
    taxa_registro_bps = 0.0

    try:
        if not os.path.exists(_AR_PLANILHA):
            raise FileNotFoundError('calculadora_ar.xlsx não encontrada')
        import openpyxl
        wb = openpyxl.load_workbook(_AR_PLANILHA, read_only=True, data_only=True)
        ws = wb['Custos CERC']

        rows = list(ws.iter_rows(min_row=1, values_only=True))

        # Tarifa por contrato (faixas): linhas 9-18 (índice 8-17), cols A=0, B=1, C=2
        for row in rows[8:18]:
            try:
                qmin = int(float(row[0])) if row[0] is not None else None
                qmax = int(float(row[1])) if row[1] is not None else None
                tar  = float(row[2])      if row[2] is not None else None
                if qmin is not None and qmax is not None and tar is not None:
                    tarifa_faixas.append((qmin, qmax, tar))
            except (TypeError, ValueError):
                continue

        # Linha 43 (índice 42): col C=qtd, D=tarifa, E=carrego, F=registro
        # col F / volume total = taxa de registro em bps
        if len(rows) > 42:
            row43 = rows[42]
            vol_total_linha22 = None
            if len(rows) > 21:
                for comp_row in rows[21:42]:
                    try:
                        v = float(comp_row[1]) if comp_row[1] is not None else 0.0
                        if v > 0:
                            vol_total_linha22 = (vol_total_linha22 or 0.0) + v
                    except (TypeError, ValueError):
                        continue
            try:
                reg_total = float(row43[5]) if row43[5] is not None else 0.0
                if vol_total_linha22 and vol_total_linha22 > 0:
                    taxa_registro_bps = reg_total / vol_total_linha22  # negativo, ex: -0.0000638
                else:
                    taxa_registro_bps = 0.0
            except (TypeError, ValueError):
                taxa_registro_bps = 0.0

        # Mapa receivable_id -> custo_carrego (col H=7, K=10), a partir da linha 4 (índice 3)
        for row in rows[3:]:
            if row[7] is None:
                continue
            rid = str(row[7]).strip()
            custo = row[10]
            if rid and custo is not None:
                try:
                    cerc_map[rid] = float(custo)
                except (TypeError, ValueError):
                    continue

        wb.close()
    except Exception as e:
        pass  # sem dados CERC — custos ficam zerados

    _ar_cerc_tarifa_faixas = tarifa_faixas
    result = (cerc_map, tarifa_faixas, taxa_registro_bps)
    _ar_cerc_cache = result
    _ar_cerc_cache_ts = agora
    return result

def calcular_indicadores_ar(records, taxa_nominal_am, custos_cerc_fixos=0.0, data_operacao=None):
    """
    Replica os cálculos da Calculadora de Antecipação de Recebíveis.

    Parâmetros:
      records           lista de dicts do SimplificaÊ
      taxa_nominal_am   taxa % CDI a.m. (ex: 0.0137 = 1.37%)
      custos_cerc_fixos custos CERC fixos da operação em R$ (default 0)
      data_operacao     datetime.date da operação (default: hoje)

    Retorna dict com indicadores ou {'erro': mensagem}.
    """
    import datetime as _dt

    if not records:
        return {'erro': 'Sem URs para calcular.'}

    if data_operacao is None:
        data_operacao = datetime.now().date()
    elif hasattr(data_operacao, 'date'):
        data_operacao = data_operacao.date()

    curvas_dict, feriados_set, aviso_curvas = carregar_curvas_ar()

    def _networkdays(d1, d2):
        """Conta dias úteis de d1 (exclusive) até d2 (inclusive), excluindo feriados."""
        if d2 <= d1:
            return 0
        count = 0
        cur = d1
        one = _dt.timedelta(days=1)
        while cur < d2:
            cur += one
            if cur.weekday() < 5 and cur not in feriados_set:
                count += 1
        return count

    def _vlookup(prazo_du):
        if prazo_du <= 0:
            return {'cdi_du': 0.0, 'cof_am': 0.0}
        if prazo_du in curvas_dict:
            return curvas_dict[prazo_du]
        menores = [k for k in curvas_dict if k <= prazo_du]
        if menores:
            return curvas_dict[max(menores)]
        return curvas_dict[min(curvas_dict)] if curvas_dict else {'cdi_du': 0.0, 'cof_am': 0.0}

    # Equivalencia taxa nominal DU->DC: planilha usa base DC (dias corridos/30)
    # taxa_nominal_am = % CDI (ex: 0.0110 = 1.10% CDI a.m.)
    # A taxa efetiva por UR = taxa_nominal_am * CDI_am_DC(prazo_du) [base DC/30]
    total_vb = sum_pdc_w = sum_pdu_w = 0.0
    sum_rb_du_jc = sum_cof_du = 0.0
    qtd_urs = 0
    _last_cdi_aa = _AR_CDI_DEFAULT  # fallback

    for r in records:
        disponivel = r.get('disponivel') or r.get('valor_bruto') or 0
        if disponivel <= 0:
            continue
        dl_raw = r.get('data_liquidacao', '')
        try:
            if hasattr(dl_raw, 'date'):
                dl = dl_raw.date()
            elif isinstance(dl_raw, str) and dl_raw:
                for _fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d/%m/%Y %H:%M'):
                    try:
                        dl = _dt.datetime.strptime(dl_raw.strip()[:10], _fmt[:10]).date()
                        break
                    except Exception:
                        dl = None
                if not dl:
                    continue
            else:
                continue
        except Exception:
            continue
        if dl <= data_operacao:
            continue

        prazo_dc = (dl - data_operacao).days
        prazo_du = _networkdays(data_operacao, dl)
        if prazo_du <= 0 or prazo_dc <= 0:
            continue

        curva      = _vlookup(prazo_du)
        cof_am     = curva['cof_am']
        cdi_base_aa = curva.get('cdi_base_aa', _AR_CDI_DEFAULT)
        _last_cdi_aa = cdi_base_aa  # guardar para usar no prazo médio

        # Receita Bruta: formula exata da planilha col W (Juros Compostos base DC)
        # RB = K * ((1 + taxa_am_dc)^(DC/30) - 1)
        # taxa_am_dc = taxa_nominal_am  (input do operador: taxa a.m. base DC)
        # Nota: taxa_nominal_am aqui e a taxa mensal DC que o operador digita
        rb_dc        = disponivel * ((1 + taxa_nominal_am) ** (prazo_dc / 30.0) - 1)

        cof_du       = -disponivel * ((1 + cof_am) ** (prazo_du / 21.0) - 1)   # COF: base DU
        total_vb     += disponivel
        sum_pdc_w    += prazo_dc * disponivel
        sum_pdu_w    += prazo_du * disponivel
        sum_rb_du_jc += rb_dc
        sum_cof_du   += cof_du
        qtd_urs      += 1

    if total_vb <= 0:
        return {'erro': 'Nenhuma UR válida com valor disponível.'}

    prazo_medio_dc = sum_pdc_w / total_vb
    prazo_medio_du = sum_pdu_w / total_vb
    # CDI a.m. a partir do CDI base a.a. (21 DU por mês)
    curva_med  = _vlookup(max(1, round(prazo_medio_du)))
    cdi_base_aa = curva_med.get('cdi_base_aa', _last_cdi_aa)
    cdi_am_du   = (1 + cdi_base_aa) ** (21.0 / 252.0) - 1   # CDI a.m. base DU
    # CDI equivalencia DC para o prazo medio: (1+CDI_aa)^(PM_DU/252 * 360/PM_DC) - 1
    _exp_dc_equiv = (prazo_medio_du / 252.0) * (360.0 / prazo_medio_dc) if prazo_medio_dc > 0 else 1.0
    cdi_equiv_dc  = (1 + cdi_base_aa) ** _exp_dc_equiv - 1             # CDI equiv DC (F5)
    cdi_dc_am     = (1 + cdi_equiv_dc) ** (30.0 / 360.0) - 1           # Taxa CDI a.m. DC (F6)
    pct_cdi       = taxa_nominal_am / cdi_dc_am if cdi_dc_am > 0 else 0.0  # %CDI = F12/F6

    receita_bruta = sum_rb_du_jc
    cof_total     = sum_cof_du
    imposto       = max(0.0, receita_bruta + cof_total) * -0.0465
    receita_liq   = receita_bruta + cof_total + imposto

    # Custo CERC -- 3 componentes calculados a partir da agenda, por mes de competencia
    # Conforme planilha calculadora_ar.xlsx, aba Custos CERC
    _cerc_batch_carrego = 0.0   # Pendente: logica B6 a confirmar
    _cerc_registro_val  = 0.0
    _cerc_tarifa_val    = 0.0
    if custos_cerc_fixos == 0.0:
        try:
            # Taxa registro:  = taxa ponderada da faixa de volume
            # Faixas (col F = taxa ponderada acumulada):
            _TAXA_REG_FAIXAS = [
                (0,       25e6,    -0.0003),
                (25e6,    75e6,    -0.00017),
                (75e6,    175e6,   -0.000107142857),
                (175e6,   400e6,   -0.0000637499),
                (400e6,   float('inf'), -0.000015),
            ]
            # Tarifa por contrato por faixa de qtd URs (por mes)
            _FAIXAS_CONTRATO = [
                (0,30,-0.33),(31,100,-1.5),(101,500,-2.28),(501,1000,-4.57),
                (1001,2000,-9.14),(2001,4000,-18.3),(4001,8000,-36.59),
                (8001,16000,-54.89),(16001,32000,-109.78),(32001,999999,-105.6033),
            ]
            import datetime as _dt2
            from collections import defaultdict as _dd
            _mes_vol = _dd(float)
            _mes_qtd = _dd(int)
            for _r in records:
                _v = _r.get('disponivel', 0)
                if _v <= 0: continue
                _dl_raw = _r.get('data_liquidacao', '')
                _dl2 = None
                for _fmt in ('%d/%m/%Y', '%Y-%m-%d'):
                    try: _dl2 = _dt2.datetime.strptime(_dl_raw.strip()[:10], _fmt).date(); break
                    except: pass
                if not _dl2 or _dl2 <= (data_operacao or datetime.now().date()): continue
                _mes = _dt2.date(_dl2.year, _dl2.month, 1)
                _mes_vol[_mes] += _v
                _mes_qtd[_mes] += 1
            # Soma por mes
            # Taxa registro = F53: fixo = -6.375e-5
            # (SUMPRODUCT das faixas ate 400MM, ponderado por volume de faixa)
            _taxa_reg_total = -6.374999999999998e-05
            # for removido - taxa e fixa independente do volume
            for _mes, _vol_m in _mes_vol.items():
                _qtd_m = _mes_qtd[_mes]
                # Tarifa por contrato: VLOOKUP(qtd_mes, faixas, 3, 1)
                _tar_m = 0.0
                for _qlo, _qhi, _tv in _FAIXAS_CONTRATO:
                    if _qlo <= _qtd_m <= _qhi: _tar_m = abs(_tv); break
                _cerc_tarifa_val += _tar_m
                # Registro: volume_mes *  (taxa baseada no VOLUME TOTAL da operacao)
                # A planilha usa  = taxa ponderada do volume total (nao por mes)
                _cerc_registro_val += abs(_taxa_reg_total) * _vol_m
        except Exception:
            pass
    else:
        _cerc_batch_carrego = abs(custos_cerc_fixos)
    custos_cerc_calc = _cerc_batch_carrego + _cerc_registro_val + _cerc_tarifa_val

    margem        = receita_liq - custos_cerc_calc
    yield_margem  = margem / total_vb if total_vb > 0 else 0.0

    capital_alocado = total_vb - receita_bruta
    giro_ano        = 252.0 / prazo_medio_du if prazo_medio_du > 0 else 0.0

    if capital_alocado > 0 and prazo_medio_du > 0:
        roic_aa       = (1 + (margem - cof_total) / capital_alocado) ** (252 / prazo_medio_du) - 1
        custo_capital = (1 + abs(cof_total) / capital_alocado)       ** (252 / prazo_medio_du) - 1
    else:
        roic_aa = custo_capital = 0.0

    criacao_valor = roic_aa - custo_capital

    if criacao_valor < 0:
        semaforo, semaforo_cor = 'Recusar',   '#D32F2F'
    elif criacao_valor < 0.0075:
        semaforo, semaforo_cor = 'Avaliar',   '#F57C00'
    elif criacao_valor < 0.02:
        semaforo, semaforo_cor = 'Aceitar',   '#388E3C'
    else:
        semaforo, semaforo_cor = 'Priorizar', '#1B5E20'

    return {
        'volume_total':    round(total_vb, 2),
        'qtd_urs':         qtd_urs,
        'prazo_medio_dc':  round(prazo_medio_dc, 1),
        'prazo_medio_du':  round(prazo_medio_du, 1),
        'cdi_base_aa':     round(cdi_base_aa, 6),
        'cdi_am_du':       round(cdi_am_du, 6),
        'cdi_equiv_dc':    round(cdi_equiv_dc, 6),    # CDI equivalencia DC (F5)
        'cdi_dc_am':       round(cdi_dc_am, 6),       # Taxa CDI a.m. DC (F6)
        'pct_cdi':         round(pct_cdi, 4),
        'taxa_nominal_am': round(taxa_nominal_am, 6),
        'receita_bruta':   round(receita_bruta, 2),
        'cof_total':       round(cof_total, 2),
        'imposto':         round(imposto, 2),
        'receita_liquida': round(receita_liq, 2),
        'custos_cerc':          round(-custos_cerc_calc, 2),
        'cerc_batch_carrego':   round(-_cerc_batch_carrego, 2),
        'cerc_registro':        round(-_cerc_registro_val, 2),
        'cerc_tarifa':          round(-_cerc_tarifa_val, 2),
        'margem':          round(margem, 2),
        'yield_margem':    round(yield_margem, 6),
        'capital_alocado': round(capital_alocado, 2),
        'giro_ano':        round(giro_ano, 2),
        'roic_aa':         round(roic_aa, 6),
        'custo_capital':   round(custo_capital, 6),
        'criacao_valor':   round(criacao_valor, 6),
        'semaforo':        semaforo,
        'semaforo_cor':    semaforo_cor,
        'aviso_curvas':    aviso_curvas,
    }


# ─────────────────────────────────────────────────────────────────────────────
def generate_cotacao_xw(records, empresa_nome, taxa_nominal, di_periodo,

                        seller_map, output_path, ineligible_cnpjs=None, missing_cnpjs=None):

    """

    Substituto de generate_cotacao para arquivos grandes (>= 50k URs).

    Usa xlsxwriter constant_memory=True: escreve sequencialmente sem

    acumular em RAM. Todos os totais sao pre-calculados antes de abrir

    o workbook para permitir escrita 100% linear.

    """



    today = datetime.now()

    taxa_diaria = (1 + taxa_nominal) ** (1 / 30) - 1

    _inelig = set(ineligible_cnpjs) if ineligible_cnpjs else set()
    _missing = set(missing_cnpjs)    if missing_cnpjs    else set()



    # ------------------------------------------------------------------

    # PRE-CALCULO: varrer records UMA vez acumulando tudo que precisamos

    # ------------------------------------------------------------------

    from collections import defaultdict

    import sys, os

    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))



    # Importar helpers do webapp

    from webapp import parse_date, ARRANJO_BANDEIRA, ADQUIRENTES_ORDER, \
        GREEN_DARK, GREEN_MID, GREEN_LIGHT, GREEN_LIGHTER, WHITE_HEX



    def _pd(s, _cache={}):

        if s not in _cache:

            _cache[s] = parse_date(s)

        return _cache[s]



    # Totais para premissas

    td = tl = total_vb = agenda_total_sum = avg_prazo_num = 0.0

    cdi_last = 0.0



    # Dados pre-calculados por linha (lista de listas para write_row)

    rows_data = []



    # Grupos por CNPJ para a aba Consolidado

    cnpj_groups = OrderedDict()



    for idx, r in enumerate(records):

        dl = _pd(r.get('data_liquidacao', ''))

        da = _pd(r.get('data_atualizacao', ''))

        vb = r.get('disponivel', 0)

        pdc = max(0, (dl - today).days) if dl else 30

        dp = (1 + taxa_diaria) ** pdc - 1

        dr = vb * dp

        liq = vb - dr

        cdi = ((1 + dp) ** (360 / pdc) - 1) / di_periodo if (di_periodo > 0 and pdc > 0) else 0

        buf = r.get('buffer', 0)

        buf = buf / 100 if buf > 1 else buf

        cnpj = r.get('cnpj', '')

        sid_val = seller_map.get(cnpj, seller_map.get(r.get('cnpj_original', ''), ''))

        band = ARRANJO_BANDEIRA.get(r.get('arranjo', ''), r.get('arranjo', ''))



        td += dr

        tl += liq

        total_vb += vb

        cdi_last = cdi

        agenda_total_sum += r.get('agenda_total', 0)

        if dl and vb > 0:

            avg_prazo_num += pdc * vb



        rows_data.append([

            idx + 1,

            r.get('receivable_id', ''),

            cnpj,

            r.get('cnpj_adquirente', ''),

            r.get('adquirente', ''),

            r.get('arranjo', ''),

            dl.strftime('%d/%m/%Y') if dl else '',

            da.strftime('%d/%m/%Y %H:%M') if da else '',

            r.get('agenda_total', 0),

            r.get('agenda_comprometida', 0),

            r.get('agenda_livre', 0),

            vb, buf, sid_val, band,

            pdc, 0, dp, dr, liq, cdi,

            cnpj in _inelig   # col 22: flag inelegivel (nao exibida)

        ])



        if cnpj not in cnpj_groups:

            cnpj_groups[cnpj] = []

        cnpj_groups[cnpj].append(r)



    dtp = td / total_vb if total_vb > 0 else 0

    avg_prazo = avg_prazo_num / total_vb if total_vb > 0 else 30



    # ------------------------------------------------------------------

    # WORKBOOK  (constant_memory: escrita sequencial, sem RAM extra)

    # ------------------------------------------------------------------

    wb = xlsxwriter.Workbook(output_path, {'constant_memory': True})



    # --- Formatos ---

    brl     = wb.add_format({'num_format': 'R$ #,##0.00'})

    pct     = wb.add_format({'num_format': '0.0000%'})

    pct_s   = wb.add_format({'num_format': '0.00%'})

    pct_l   = wb.add_format({'num_format': '0.0000000000%'})

    intfmt  = wb.add_format({'num_format': '0'})

    txt     = wb.add_format()

    hdr     = wb.add_format({'bold': True, 'bg_color': '#' + GREEN_MID,

                              'font_color': '#' + WHITE_HEX, 'border': 1})

    tot     = wb.add_format({'bold': True, 'bg_color': '#A5D6A7',

                              'font_color': '#' + GREEN_DARK})

    tot_brl = wb.add_format({'bold': True, 'bg_color': '#A5D6A7',

                              'font_color': '#' + GREEN_DARK, 'num_format': 'R$ #,##0.00'})

    tot_pct = wb.add_format({'bold': True, 'bg_color': '#A5D6A7',

                              'font_color': '#' + GREEN_DARK, 'num_format': '0.0000%'})

    pre_lbl = wb.add_format({'font_color': '#9E9E9E', 'bg_color': '#' + GREEN_LIGHTER,

                              'border': 1, 'align': 'center', 'font_size': 10})

    pre_brl = wb.add_format({'bold': True, 'font_color': '#' + GREEN_DARK,

                              'bg_color': '#' + GREEN_LIGHTER, 'border': 1,

                              'num_format': 'R$ #,##0.00'})

    pre_pct = wb.add_format({'bold': True, 'font_color': '#' + GREEN_DARK,

                              'bg_color': '#' + GREEN_LIGHTER, 'border': 1,

                              'num_format': '0.0000%'})

    pre_ps  = wb.add_format({'bold': True, 'font_color': '#' + GREEN_DARK,

                              'bg_color': '#' + GREEN_LIGHTER, 'border': 1,

                              'num_format': '0.00%'})

    pre_pl  = wb.add_format({'bold': True, 'font_color': '#' + GREEN_DARK,

                              'bg_color': '#' + GREEN_LIGHTER, 'border': 1,

                              'num_format': '0.0000000000%'})

    pre_val = wb.add_format({'bold': True, 'font_color': '#' + GREEN_DARK,

                              'bg_color': '#' + GREEN_LIGHTER, 'border': 1})

    inelig_fmt = wb.add_format({'bg_color': '#FFCDD2', 'font_color': '#B71C1C',

                                 'font_size': 8})

    title_fmt  = wb.add_format({'bold': True, 'font_color': '#' + GREEN_DARK,

                                 'font_size': 16})

    sub_fmt    = wb.add_format({'bold': True, 'font_color': '#' + GREEN_MID,

                                 'font_size': 11})

    dark_hdr   = wb.add_format({'bold': True, 'bg_color': '#' + GREEN_DARK,

                                 'font_color': '#' + WHITE_HEX, 'border': 1})

    mid_hdr    = wb.add_format({'bold': True, 'bg_color': '#' + GREEN_MID,

                                 'font_color': '#' + WHITE_HEX, 'border': 1})

    c_tot_brl  = wb.add_format({'bold': True, 'bg_color': '#A5D6A7',

                                 'font_color': '#' + GREEN_DARK,

                                 'num_format': 'R$ #,##0.00', 'border': 1})

    c_tot_flt  = wb.add_format({'bold': True, 'bg_color': '#A5D6A7',

                                 'font_color': '#' + GREEN_DARK,

                                 'num_format': '0.00', 'border': 1})

    c_tot_txt  = wb.add_format({'bold': True, 'bg_color': '#A5D6A7',

                                 'font_color': '#' + GREEN_DARK, 'border': 1})

    c_pre_lbl  = wb.add_format({'font_color': '#9E9E9E', 'bg_color': '#' + GREEN_LIGHTER,

                                 'border': 1})

    c_pre_pct  = wb.add_format({'bold': True, 'font_color': '#' + GREEN_DARK,

                                 'bg_color': '#' + GREEN_LIGHTER, 'border': 1,

                                 'num_format': '0.0000%'})

    c_pre_ps   = wb.add_format({'bold': True, 'font_color': '#' + GREEN_DARK,

                                 'bg_color': '#' + GREEN_LIGHTER, 'border': 1,

                                 'num_format': '0.00%'})

    c_pre_flt  = wb.add_format({'bold': True, 'font_color': '#' + GREEN_DARK,

                                 'bg_color': '#' + GREEN_LIGHTER, 'border': 1,

                                 'num_format': '0.00'})

    data_fmt   = wb.add_format({'font_size': 10})

    even_fmt   = wb.add_format({'bg_color': '#' + GREEN_LIGHTER})

    even_brl   = wb.add_format({'bg_color': '#' + GREEN_LIGHTER,

                                 'num_format': 'R$ #,##0.00'})

    even_flt   = wb.add_format({'bg_color': '#' + GREEN_LIGHTER,

                                 'num_format': '0.00'})

    inelig_brl = wb.add_format({'bg_color': '#FFCDD2', 'font_color': '#B71C1C',

                                 'num_format': 'R$ #,##0.00'})

    inelig_flt = wb.add_format({'bg_color': '#FFCDD2', 'font_color': '#B71C1C',

                                 'num_format': '0.00'})



    # ------------------------------------------------------------------

    # ABA ANALITICO  (escrita sequencial: premissas → headers → dados → total)

    # ------------------------------------------------------------------

    ws_a = wb.add_worksheet('Analitico')

    ws_a.set_tab_color('#' + GREEN_MID)



    # Larguras de coluna (definidas antes de qualquer write)

    cw = {0:5, 1:38, 2:18, 3:18, 4:14, 5:10, 6:14, 7:18,

          8:16, 9:18, 10:16, 11:16, 12:8, 13:12, 14:10,

          15:10, 16:10, 17:14, 18:14, 19:14, 20:14}

    for col, w in cw.items():

        ws_a.set_column(col, col, w)



    # Linha 1: titulo

    ws_a.merge_range('B1:F1', 'ANALITICO - ' + empresa_nome.upper(), title_fmt)

    ws_a.write('B2', today.strftime('%d/%m/%Y'), sub_fmt)



    # Linha 2: headers das premissas (col G=6 em diante)

    p_headers = ['Taxa diaria', 'Agenda Total', '', 'Valor Bruto', '', '',

                 'Taxa Nominal', 'Prazo DC', 'Prazo DU',

                 'Desconto %', 'Desconto R$', 'Liquido', '%CDI']

    pcs = 6

    for i, h in enumerate(p_headers):

        ws_a.write(1, pcs + i, h, pre_lbl)



    # Linha 3: valores das premissas (pre-calculados — escrita sequencial OK)

    ws_a.write_formula(2, pcs,     '=(1+M3)^(1/30)-1',   pre_pl)

    ws_a.write        (2, pcs + 1,  agenda_total_sum,      pre_brl)

    ws_a.write        (2, pcs + 2,  '',                    pre_val)

    ws_a.write        (2, pcs + 3,  total_vb,              pre_brl)

    ws_a.write        (2, pcs + 4,  '',                    pre_val)

    ws_a.write        (2, pcs + 5,  '',                    pre_val)

    ws_a.write_formula(2, pcs + 6, '=Consolidado!D5',      pre_pct)   # M3 taxa

    ws_a.write        (2, pcs + 7,  avg_prazo,             pre_val)

    ws_a.write        (2, pcs + 8,  0,                     pre_val)

    ws_a.write        (2, pcs + 9,  dtp,                   pre_pct)

    ws_a.write        (2, pcs + 10, td,                    pre_brl)

    ws_a.write        (2, pcs + 11, tl,                    pre_brl)

    ws_a.write_formula(2, pcs + 12, '=Consolidado!D8',     pre_ps)    # S3 DI



    # Linha 5: headers dos dados

    a_headers = ['#', 'ID', 'CNPJ', 'CNPJ Adquirente', 'Adquirente', 'Arranjo',

                 'Data Liquidacao', 'Data Atualizacao', 'Agenda Total',

                 'Agenda Comprometida', 'Agenda Livre', 'Valor Bruto',

                 'Buffer', 'Seller ID', 'Bandeira', 'Prazo DC', 'Prazo DU',

                 'Desconto %', 'Desconto R$', 'Liquido', '%CDI']

    for col, h in enumerate(a_headers):

        ws_a.write(4, col, h, hdr)



    # Formatos por coluna para write_row (sem flag inelig)

    col_fmts = [intfmt, txt, txt, txt, txt, txt,

                txt, txt, brl, brl, brl, brl,

                pct_s, txt, txt, intfmt, intfmt,

                pct, brl, brl, pct]



    # Dados — write_row para linhas normais, write individual para inelegiveis

    for i, row in enumerate(rows_data):

        rn = 5 + i

        if row[21]:  # inelegivel

            for col, (val, fmt) in enumerate(zip(row[:21], col_fmts)):

                ws_a.write(rn, col, val, inelig_fmt)

        else:

            ws_a.write_row(rn, 0, row[:21])



    # Total

    tr = 5 + len(records)

    ws_a.write(tr, 0,  'TOTAL',           tot)

    ws_a.write(tr, 8,  agenda_total_sum,  tot_brl)

    ws_a.write(tr, 11, total_vb,          tot_brl)

    ws_a.write(tr, 17, dtp,               tot_pct)

    ws_a.write(tr, 18, td,                tot_brl)

    ws_a.write(tr, 19, tl,                tot_brl)

    for col in [1,2,3,4,5,6,7,9,10,12,13,14,15,16,20]:

        ws_a.write(tr, col, '', tot)



    ws_a.freeze_panes(5, 0)



    # ------------------------------------------------------------------

    # ABA CONSOLIDADO

    # ------------------------------------------------------------------

    ws_c = wb.add_worksheet('Consolidado')

    ws_c.set_tab_color('#' + GREEN_DARK)



    ws_c.set_column(1, 1, 6)

    ws_c.set_column(2, 2, 28)

    ws_c.set_column(3, 3, 18)

    ws_c.set_column(4, 7, 16)

    ws_c.set_column(8, 8, 3)

    ws_c.set_column(9, 13, 16)



    # Linha 1: titulo

    title_c = wb.add_format({'bold': True, 'font_size': 18,

                              'font_color': '#' + WHITE_HEX,

                              'bg_color': '#' + GREEN_DARK, 'align': 'left'})

    ws_c.merge_range('B1:I1', '  ' + empresa_nome.upper(), title_c)

    for col in range(1, 15):

        ws_c.write(0, col, '', wb.add_format({'bg_color': '#' + GREEN_DARK}))

    ws_c.write('B2', 'Cotacao de Antecipacao de Recebiveis - ' +

               today.strftime('%d/%m/%Y'), sub_fmt)



    # Premissas

    ws_c.write(3, 2, 'PREMISSAS',

               wb.add_format({'bold': True, 'font_size': 11,

                               'font_color': '#' + WHITE_HEX,

                               'bg_color': '#' + GREEN_MID}))

    ws_c.write(3, 3, '', wb.add_format({'bg_color': '#' + GREEN_MID}))



    prem_rows = [

        ('Taxa Nominal %', taxa_nominal,  c_pre_pct),

        ('Taxa efetiva',   dtp,           c_pre_pct),

        ('% do DI',        cdi_last,      c_pre_pct),

        ('DI Periodo',     di_periodo,    c_pre_ps),

        ('Prazo DC',       avg_prazo,     c_pre_flt),

        ('Prazo DU',       0,             c_pre_flt),

    ]

    for i, (lbl, val, fmt) in enumerate(prem_rows):

        ws_c.write(4 + i, 2, lbl, c_pre_lbl)

        ws_c.write(4 + i, 3, val, fmt)



    # Headers consolidado (linha 12 = row 11)

    mh = ['#', 'Nome do cliente', 'CPF/CNPJ', 'Valor bruto',

          'Desconto', 'Valor liquido', 'Prazo DC']

    for col, h in enumerate(mh, 1):

        ws_c.write(11, col, h, dark_hdr)

    ws_c.write(11, 8, '', dark_hdr)

    for i, adq in enumerate(ADQUIRENTES_ORDER):

        ws_c.write(11, 9 + i, adq, mid_hdr)



    # Linha de totais (linha 13 = row 12) — pre-calculada

    sums = []

    for cnpj, crecs in cnpj_groups.items():

        vb = sum(r.get('disponivel', 0) for r in crecs)

        dr2 = wp2 = 0.0

        for r in crecs:

            dl = _pd(r.get('data_liquidacao', ''))

            v = r.get('disponivel', 0)

            if dl and v > 0:

                p = max(0, (dl - today).days)

                wp2 += p * v

                dr2 += v * ((1 + taxa_diaria) ** p - 1)

        pdc_c = wp2 / vb if vb > 0 else 30

        liq2 = vb - dr2

        adqv = {a: 0.0 for a in ADQUIRENTES_ORDER}

        for r in crecs:

            an = r.get('adquirente', '').strip().upper()

            if 'PICPAY' in an:   adqv['PICPAY'] += r.get('disponivel', 0)

            elif 'CIELO' in an:  adqv['Cielo']  += r.get('disponivel', 0)

            elif 'REDE' in an:   adqv['Rede']   += r.get('disponivel', 0)

            elif 'STONE' in an:  adqv['Stone']  += r.get('disponivel', 0)

            elif 'GETNET' in an: adqv['Getnet'] += r.get('disponivel', 0)

        sums.append({'vb': vb, 'dr': dr2, 'liq': liq2,

                     'pdc': pdc_c, 'adq': adqv, 'cnpj': cnpj})



    # Total geral (row 12)

    tot_vb2  = sum(s['vb']  for s in sums)

    tot_dr2  = sum(s['dr']  for s in sums)

    tot_liq2 = sum(s['liq'] for s in sums)

    tot_pdc2 = (sum(s['pdc'] * s['vb'] for s in sums) / tot_vb2

                if tot_vb2 > 0 else 0)

    ws_c.write(12, 1, 'TOTAL',           c_tot_txt)

    ws_c.write(12, 4, tot_vb2,           c_tot_brl)

    ws_c.write(12, 5, tot_dr2,           c_tot_brl)

    ws_c.write(12, 6, tot_liq2,          c_tot_brl)

    ws_c.write(12, 7, round(tot_pdc2,2), c_tot_flt)

    for i, adq in enumerate(ADQUIRENTES_ORDER):

        ws_c.write(12, 9+i, sum(s['adq'][adq] for s in sums), c_tot_brl)

    for col in [2, 3, 8]:

        ws_c.write(12, col, '', c_tot_txt)



    # Linhas por CNPJ (rows 13+)

    for idx, s in enumerate(sums, 1):

        row = 12 + idx

        is_even = idx % 2 == 0

        is_inelig  = s['cnpj'] in _inelig

        is_missing = s['cnpj'] in _missing and not is_inelig

        rf     = inelig_fmt  if is_inelig  else (missing_fmt if is_missing else (even_fmt if is_even else txt))

        rf_brl = inelig_brl  if is_inelig  else (missing_brl if is_missing else (even_brl if is_even else brl))

        rf_flt = inelig_flt  if is_inelig  else (missing_flt if is_missing else (even_flt if is_even else c_pre_flt))

        ws_c.write(row, 1, idx,                      rf)

        ws_c.write(row, 2, empresa_nome.upper(),      rf)

        ws_c.write(row, 3, s['cnpj'],                 rf)

        ws_c.write(row, 4, s['vb'],                   rf_brl)

        ws_c.write(row, 5, s['dr'],                   rf_brl)

        ws_c.write(row, 6, s['liq'],                  rf_brl)

        ws_c.write(row, 7, round(s['pdc'], 2),        rf_flt)

        ws_c.write(row, 8, '',                        rf)

        for i, adq in enumerate(ADQUIRENTES_ORDER):

            ws_c.write(row, 9+i, s['adq'][adq], rf_brl)



    ws_c.freeze_panes(13, 0)



    # ------------------------------------------------------------------

    # ABA ARRANJOS

    # ------------------------------------------------------------------

    ws_arr = wb.add_worksheet('Arranjos')

    ws_arr.set_tab_color('#' + GREEN_DARK)

    ws_arr.write(0, 0, 'Codigo',   wb.add_format({'bold': True, 'bg_color': '#' + GREEN_MID, 'font_color': 'white', 'border': 1}))

    ws_arr.write(0, 1, 'Bandeira', wb.add_format({'bold': True, 'bg_color': '#' + GREEN_MID, 'font_color': 'white', 'border': 1}))

    for i, (code, name) in enumerate(ARRANJO_BANDEIRA.items(), 1):

        ws_arr.write(i, 0, code, data_fmt)

        ws_arr.write(i, 1, name, data_fmt)



    # ------------------------------------------------------------------

    # ABA SELLERS ID

    # ------------------------------------------------------------------

    ws_sel = wb.add_worksheet('Sellers ID')

    ws_sel.set_tab_color('#' + GREEN_DARK)

    ws_sel.write(0, 0, 'CNPJ',      wb.add_format({'bold': True, 'bg_color': '#' + GREEN_MID, 'font_color': 'white', 'border': 1}))

    ws_sel.write(0, 1, 'Seller ID', wb.add_format({'bold': True, 'bg_color': '#' + GREEN_MID, 'font_color': 'white', 'border': 1}))

    for i, cnpj in enumerate(sorted(cnpj_groups.keys()), 1):

        sid_val = seller_map.get(cnpj, '')

        ws_sel.write(i, 0, cnpj,    data_fmt)

        ws_sel.write(i, 1, sid_val, data_fmt)



    wb.close()

    return len(records), len(cnpj_groups)



def generate_cotacao(records, empresa_nome, taxa_nominal, di_periodo, seller_map, output_path, ineligible_cnpjs=None, missing_cnpjs=None):



    # Para arquivos grandes: usar xlsxwriter (3x mais rapido, sem OOMKill)



    if len(records) >= 3000:  # xlsxwriter: ~3x mais rapido que openpyxl



        return generate_cotacao_xw(records, empresa_nome, taxa_nominal, di_periodo,



                                   seller_map, output_path, ineligible_cnpjs,



                                   missing_cnpjs=missing_cnpjs)



    wb = openpyxl.Workbook()



    today = datetime.now()



    taxa_diaria = (1 + taxa_nominal) ** (1/30) - 1



    _ineligible_cnpjs = set(ineligible_cnpjs) if ineligible_cnpjs else set()

    _missing_cnpjs    = set(missing_cnpjs)    if missing_cnpjs    else set()



    # ANALITICO



    ws_a = wb.active



    ws_a.title = 'Analitico'



    ws_a.sheet_properties.tabColor = GREEN_MID



    ws_a.merge_cells('B1:F1')



    c = ws_a['B1']



    c.value = f'ANALITICO - {empresa_nome.upper()}'



    c.font = font_title



    c.alignment = align_left



    ws_a['B2'] = today.strftime('%d/%m/%Y')



    ws_a['B2'].font = font_subtitle



    # Premissas header



    p_headers = ['Taxa diaria', 'Agenda Total', '', 'Valor Bruto', '', '',



                 'Taxa Nominal', 'Prazo DC', 'Prazo DU', 'Desconto %',



                 'Desconto R$', 'Liquido', '%CDI']



    pcs = 7



    for i, h in enumerate(p_headers):



        cell = ws_a.cell(row=2, column=pcs + i, value=h)



        cell.font = font_premissa_label



        cell.fill = fill_green_light



        cell.border = thin_border



        cell.alignment = align_center



    total_vb = sum(r.get('disponivel', 0) for r in records)



    agenda_total = sum(r.get('agenda_total', 0) for r in records)



    wp = sum(max(0, (parse_date(r.get('data_liquidacao', '')) - today).days) * r.get('disponivel', 0)



             for r in records if parse_date(r.get('data_liquidacao', '')) and r.get('disponivel', 0) > 0)



    avg_prazo = wp / total_vb if total_vb > 0 else 30



    pvals = [(None, fmt_pct_long), (agenda_total, fmt_brl), None, (total_vb, fmt_brl),



             None, None, (taxa_nominal, fmt_pct), (avg_prazo, '0.00'), (0, '0'),



             None, None, None, None]



    for i, pv in enumerate(pvals):



        cell = ws_a.cell(row=3, column=pcs + i)



        cell.fill = fill_green_lighter



        cell.border = thin_border



        cell.alignment = align_center



        if pv:



            cell.value = pv[0]



            cell.number_format = pv[1]



            cell.font = font_premissa_value



    # G3 = taxa diaria como formula derivada de M3 (permite alterar taxa no Consolidado e recalcular tudo)



    ws_a['G3'] = '=(1+M3)^(1/30)-1'



    ws_a['G3'].number_format = fmt_pct_long



    ws_a['G3'].font = font_premissa_value



    ws_a['G3'].fill = fill_green_lighter



    ws_a['G3'].border = thin_border



    ws_a['G3'].alignment = align_center



    # M3 = taxa nominal referenciando Consolidado!D5 (fonte unica da taxa — altere la, reflete aqui)



    ws_a['M3'] = '=Consolidado!D5'



    ws_a['M3'].number_format = fmt_pct



    ws_a['M3'].font = font_premissa_value



    ws_a['M3'].fill = fill_green_lighter



    ws_a['M3'].border = thin_border



    ws_a['M3'].alignment = align_center



    # S3 = DI Periodo referenciando Consolidado!D8 (altere la, %CDI reflete automaticamente)



    ws_a['S3'] = '=Consolidado!D8'



    ws_a['S3'].number_format = fmt_pct_short



    ws_a['S3'].font = font_premissa_value



    ws_a['S3'].fill = fill_green_lighter



    ws_a['S3'].border = thin_border



    ws_a['S3'].alignment = align_center



    # Threshold: se muitas URs, usar valores em vez de formulas (performance)



    USE_FORMULAS = len(records) < 70000  # acima de 70k URs: valores pre-calculados (mais rapido)



    # Headers



    a_headers = ['#', 'ID', 'CNPJ', 'CNPJ Adquirente', 'Adquirente', 'Arranjo',



                 'Data Liquidacao', 'Data Atualizacao', 'Agenda Total',



                 'Agenda Comprometida', 'Agenda Livre', 'Valor Bruto',



                 'Buffer', 'Seller ID', 'Bandeira', 'Prazo DC', 'Prazo DU',



                 'Desconto %', 'Desconto R$', 'Liquido', '%CDI']



    set_header_row(ws_a, 5, a_headers)



    td = 0



    tl = 0



    cdi_last = 0



    _date_cache = {}



    def _pd(s):



        if s not in _date_cache: _date_cache[s] = parse_date(s)



        return _date_cache[s]



    # column_dimensions: define fmt de coluna UMA vez - Excel aplica a todas as celulas



    _col_fmts = {7: 'DD/MM/YYYY', 8: 'DD/MM/YYYY HH:MM',



                 9: fmt_brl, 10: fmt_brl, 11: fmt_brl, 12: fmt_brl,



                 13: fmt_pct_short,



                 16: '0', 18: fmt_pct, 19: fmt_brl, 20: fmt_brl, 21: fmt_pct}



    from openpyxl.utils import get_column_letter as _gcl



    for _ci, _cf in _col_fmts.items():



        ws_a.column_dimensions[_gcl(_ci)].number_format = _cf



    for idx, r in enumerate(records):



        rn = 6 + idx



        dl = _pd(r.get('data_liquidacao', ''))



        da = _pd(r.get('data_atualizacao', ''))



        vb = r.get('disponivel', 0)



        pdc = max(0, (dl - today).days) if dl else 30



        dp = (1 + taxa_diaria) ** pdc - 1



        dr = vb * dp



        liq = vb - dr



        td += dr



        tl += liq



        if di_periodo > 0 and pdc > 0:



            cdi = ((1 + dp) ** (360 / pdc) - 1) / di_periodo



        else:



            cdi = 0



        cdi_last = cdi



        band = ARRANJO_BANDEIRA.get(r.get('arranjo', ''), r.get('arranjo', ''))



        cnpj = r.get('cnpj', '')



        sid = seller_map.get(cnpj, seller_map.get(r.get('cnpj_original', ''), ''))



        buf = r.get('buffer', 0)



        buf_val = buf / 100 if buf > 1 else buf



        # Usar append() para a linha inteira - muito mais rapido que ws_a.cell() por coluna



        if USE_FORMULAS:



            ws_a.append([idx+1, r.get('receivable_id',''), cnpj, r.get('cnpj_adquirente',''),



                         r.get('adquirente',''), r.get('arranjo',''), dl, da,



                         r.get('agenda_total',0), r.get('agenda_comprometida',0),



                         r.get('agenda_livre',0), vb, buf_val, sid, band,



                         f'=MAX(0,G{rn}-TODAY())', 0,



                         f'=(1+$G$3)^P{rn}-1', f'=L{rn}*R{rn}', f'=L{rn}-S{rn}',



                         f'=IF(P{rn}>0,((1+R{rn})^(360/P{rn})-1)/$S$3,0)'])



        else:



            ws_a.append([idx+1, r.get('receivable_id',''), cnpj, r.get('cnpj_adquirente',''),



                         r.get('adquirente',''), r.get('arranjo',''), dl, da,



                         r.get('agenda_total',0), r.get('agenda_comprometida',0),



                         r.get('agenda_livre',0), vb, buf_val, sid, band,



                         pdc, 0, dp, dr, liq, cdi])



        # Colorir linha inteira se CNPJ inelegivel (fmt via column_dimensions)



        if cnpj in _ineligible_cnpjs:



            for _c in ws_a[rn]:



                _c.fill = fill_inelig



                _c.font = font_inelig



        elif cnpj in _missing_cnpjs:



            for _c in ws_a[rn]:



                _c.fill = fill_missing



                _c.font = font_missing



    # Totals



    tr = 6 + len(records)



    dtp = td / total_vb if total_vb > 0 else 0



    for col in range(1, 22):



        ws_a.cell(row=tr, column=col).fill = fill_totals



        ws_a.cell(row=tr, column=col).border = thin_border



        ws_a.cell(row=tr, column=col).font = font_totals



    ws_a.cell(row=tr, column=1, value='TOTAL')



    if USE_FORMULAS:



        ws_a.cell(row=tr, column=9, value=f'=SUM(I6:I{tr-1})').number_format = fmt_brl



        ws_a.cell(row=tr, column=12, value=f'=SUM(L6:L{tr-1})').number_format = fmt_brl



        ws_a.cell(row=tr, column=18, value=f'=IF(L{tr}>0,S{tr}/L{tr},0)').number_format = fmt_pct



        ws_a.cell(row=tr, column=19, value=f'=SUM(S6:S{tr-1})').number_format = fmt_brl



        ws_a.cell(row=tr, column=20, value=f'=SUM(T6:T{tr-1})').number_format = fmt_brl



    else:



        ws_a.cell(row=tr, column=9, value=agenda_total).number_format = fmt_brl



        ws_a.cell(row=tr, column=12, value=total_vb).number_format = fmt_brl



        ws_a.cell(row=tr, column=18, value=dtp).number_format = fmt_pct



        ws_a.cell(row=tr, column=19, value=td).number_format = fmt_brl



        ws_a.cell(row=tr, column=20, value=tl).number_format = fmt_brl



    # Update premissas



    if USE_FORMULAS:



        ws_a.cell(row=3, column=pcs+9).value = f'=IF(L{tr}>0,S{tr}/L{tr},0)'



        ws_a.cell(row=3, column=pcs+10).value = f'=S{tr}'



        ws_a.cell(row=3, column=pcs+11).value = f'=T{tr}'



    else:



        ws_a.cell(row=3, column=pcs+9).value = dtp



        ws_a.cell(row=3, column=pcs+10).value = td



        ws_a.cell(row=3, column=pcs+11).value = tl



    ws_a.cell(row=3, column=pcs+9).number_format = fmt_pct



    ws_a.cell(row=3, column=pcs+9).font = font_premissa_value



    ws_a.cell(row=3, column=pcs+10).number_format = fmt_brl



    ws_a.cell(row=3, column=pcs+10).font = font_premissa_value



    ws_a.cell(row=3, column=pcs+11).number_format = fmt_brl



    ws_a.cell(row=3, column=pcs+11).font = font_premissa_value



    ws_a.cell(row=3, column=pcs+12).value = di_periodo



    ws_a.cell(row=3, column=pcs+12).number_format = fmt_pct_short



    ws_a.cell(row=3, column=pcs+12).font = font_premissa_value



    cw = {1:5,2:38,3:18,4:18,5:14,6:10,7:14,8:18,9:16,10:18,11:16,12:16,13:8,14:12,15:10,16:10,17:10,18:14,19:14,20:14,21:14}



    for col, w in cw.items():



        ws_a.column_dimensions[get_column_letter(col)].width = w



    ws_a.freeze_panes = 'A6'



    # CONSOLIDADO



    ws_c = wb.create_sheet('Consolidado', 0)



    ws_c.sheet_properties.tabColor = GREEN_DARK



    for col in range(2, 16):



        ws_c.cell(row=1, column=col).fill = fill_header_dark



    ws_c.merge_cells('B1:I1')



    c = ws_c['B1']



    c.value = f'  {empresa_nome.upper()}'



    c.font = Font(bold=True, size=18, color=WHITE_HEX)



    c.alignment = align_left



    c.fill = fill_header_dark



    ws_c['B2'] = f'Cotação de Antecipação de Recebíveis - {today.strftime("%d/%m/%Y")}'



    ws_c['B2'].font = font_subtitle



    ws_c['C4'] = 'PREMISSAS'



    ws_c['C4'].font = Font(bold=True, size=11, color=WHITE_HEX)



    ws_c['C4'].fill = fill_header_mid



    ws_c['D4'].fill = fill_header_mid



    ws_c['E4'].fill = fill_header_mid



    prem = [('Taxa Nominal %', taxa_nominal, fmt_pct), ('Taxa efetiva', None, fmt_pct),



            ('% do DI', None, fmt_pct), ('DI Periodo', di_periodo, fmt_pct_short),



            ('Prazo DC', None, '0.00'), ('Prazo DU', 0, '0')]



    for i, (label, value, fmt) in enumerate(prem):



        r = 5 + i



        lc = ws_c.cell(row=r, column=3, value=label)



        lc.font = font_premissa_label



        lc.fill = fill_green_lighter



        lc.border = thin_border



        vc = ws_c.cell(row=r, column=4, value=value)



        vc.font = font_premissa_value



        vc.fill = fill_green_lighter



        vc.border = thin_border



        vc.number_format = fmt



    if USE_FORMULAS:



        ws_c.cell(row=6, column=4, value='=IF(E13>0,F13/E13,0)')



        ws_c.cell(row=7, column=4, value='=IF(AND(D9>0,D8>0),((1+D6)^(360/D9)-1)/D8,0)')



        ws_c.cell(row=9, column=4, value='=H13')



    else:



        ws_c.cell(row=6, column=4, value=dtp)



        ws_c.cell(row=7, column=4, value=cdi_last)



        ws_c.cell(row=9, column=4, value=avg_prazo)



    ws_c.cell(row=6, column=4).number_format = fmt_pct



    ws_c.cell(row=6, column=4).font = font_premissa_value



    ws_c.cell(row=7, column=4).number_format = fmt_pct



    ws_c.cell(row=7, column=4).font = font_premissa_value



    ws_c.cell(row=9, column=4).number_format = '0.00'



    ws_c.cell(row=9, column=4).font = font_premissa_value



    mh = ['#', 'Nome do cliente', 'CPF/CNPJ', 'Valor bruto', 'Desconto', 'Valor liquido', 'Prazo DC']



    set_header_row(ws_c, 12, mh, col_start=2)



    ws_c.cell(row=12, column=9).fill = fill_white



    set_header_row(ws_c, 12, ADQUIRENTES_ORDER, col_start=10, fill=fill_header_mid)



    cnpj_groups = OrderedDict()



    for r in records:



        cnpj = r.get('cnpj', '')



        if cnpj not in cnpj_groups:



            cnpj_groups[cnpj] = []



        cnpj_groups[cnpj].append(r)



    sums = []



    for idx, (cnpj, crecs) in enumerate(cnpj_groups.items(), 1):



        vb = sum(r.get('disponivel', 0) for r in crecs)



        # Calcular desconto UR por UR (soma individual, nao prazo medio)



        dr2 = 0



        wp2 = 0



        for r in crecs:



            dl = parse_date(r.get('data_liquidacao', ''))



            v = r.get('disponivel', 0)



            if dl and v > 0:



                p = max(0, (dl - today).days)



                wp2 += p * v



                dp_ur = (1 + taxa_diaria) ** p - 1



                dr2 += v * dp_ur



        pdc = wp2 / vb if vb > 0 else 30



        liq2 = vb - dr2



        adqv = {a: 0 for a in ADQUIRENTES_ORDER}



        for r in crecs:



            an = r.get('adquirente', '').strip().upper()



            if 'PICPAY' in an: adqv['PICPAY'] += r.get('disponivel', 0)



            elif 'CIELO' in an: adqv['Cielo'] += r.get('disponivel', 0)



            elif 'REDE' in an: adqv['Rede'] += r.get('disponivel', 0)



            elif 'STONE' in an: adqv['Stone'] += r.get('disponivel', 0)



            elif 'GETNET' in an: adqv['Getnet'] += r.get('disponivel', 0)



        sums.append({'vb': vb, 'dr': dr2, 'liq': liq2, 'adq': adqv})



        row = 14 + idx - 1



        ie = idx % 2 == 0



        apply_zebra(ws_c, row, 2, 8, ie)



        apply_zebra(ws_c, row, 10, 14, ie)



        # Colorir linha consolidado se CNPJ inelegivel



        if cnpj in _ineligible_cnpjs:



            for _col in range(2, 15):



                _c = ws_c.cell(row=row, column=_col)



                _c.fill = fill_inelig



                _c.font = font_inelig



        elif cnpj in _missing_cnpjs:



            for _col in range(2, 15):



                _c = ws_c.cell(row=row, column=_col)



                _c.fill = fill_missing



                _c.font = font_missing



        last_a_row = 5 + len(records)



        ws_c.cell(row=row, column=2, value=idx)



        ws_c.cell(row=row, column=3, value=empresa_nome.upper())



        ws_c.cell(row=row, column=4, value=cnpj)



        if USE_FORMULAS:



            ws_c.cell(row=row, column=5, value=f"=SUMIF(Analitico!C6:C{last_a_row},D{row},Analitico!L6:L{last_a_row})")



            ws_c.cell(row=row, column=6, value=f"=SUMIF(Analitico!C6:C{last_a_row},D{row},Analitico!S6:S{last_a_row})")



            ws_c.cell(row=row, column=7, value=f"=E{row}-F{row}")



            ws_c.cell(row=row, column=8, value=f"=IF(E{row}>0,SUMPRODUCT((Analitico!C6:C{last_a_row}=D{row})*Analitico!P6:P{last_a_row}*Analitico!L6:L{last_a_row})/E{row},0)")



        else:



            ws_c.cell(row=row, column=5, value=vb)



            ws_c.cell(row=row, column=6, value=dr2)



            ws_c.cell(row=row, column=7, value=liq2)



            ws_c.cell(row=row, column=8, value=round(pdc, 2))



        ws_c.cell(row=row, column=5).number_format = fmt_brl



        ws_c.cell(row=row, column=6).number_format = fmt_brl



        ws_c.cell(row=row, column=7).number_format = fmt_brl



        ws_c.cell(row=row, column=8).number_format = '0.00'



        # Adquirentes (valores fixos - nao mudam com TODAY)



        for i, adq in enumerate(ADQUIRENTES_ORDER):



            ws_c.cell(row=row, column=10+i, value=adqv[adq]).number_format = fmt_brl



    # Totals row 13



    for col in range(2, 15):



        ws_c.cell(row=13, column=col).fill = fill_totals



        ws_c.cell(row=13, column=col).border = thin_border



    last_c_row = 13 + len(cnpj_groups)



    ws_c.cell(row=13, column=2, value='TOTAL').font = font_totals



    if USE_FORMULAS:



        ws_c.cell(row=13, column=5, value=f'=SUM(E14:E{last_c_row})')



        ws_c.cell(row=13, column=6, value=f'=SUM(F14:F{last_c_row})')



        ws_c.cell(row=13, column=7, value=f'=SUM(G14:G{last_c_row})')



    else:



        ws_c.cell(row=13, column=5, value=sum(s['vb'] for s in sums))



        ws_c.cell(row=13, column=6, value=sum(s['dr'] for s in sums))



        ws_c.cell(row=13, column=7, value=sum(s['liq'] for s in sums))



    for c in [5, 6, 7]:



        ws_c.cell(row=13, column=c).font = font_totals



        ws_c.cell(row=13, column=c).number_format = fmt_brl



    # Prazo DC medio geral ponderado



    ws_c.cell(row=13, column=8, value=f'=IF(E13>0,SUMPRODUCT(H14:H{last_c_row}*E14:E{last_c_row})/E13,0)')



    ws_c.cell(row=13, column=8).number_format = '0.00'



    ws_c.cell(row=13, column=8).font = font_totals



    for col, key in [(5,'vb'),(6,'dr'),(7,'liq')]:



        pass  # ja setado acima com formulas



    for i, adq in enumerate(ADQUIRENTES_ORDER):



        col_letter = get_column_letter(10 + i)



        cell = ws_c.cell(row=13, column=10+i, value=f'=SUM({col_letter}14:{col_letter}{last_c_row})')



        cell.number_format = fmt_brl



        cell.font = font_totals



    ws_c.column_dimensions['B'].width = 6



    ws_c.column_dimensions['C'].width = 28



    ws_c.column_dimensions['D'].width = 18



    for cl in ['E','F','G','H']: ws_c.column_dimensions[cl].width = 16



    ws_c.column_dimensions['I'].width = 3



    for cl in ['J','K','L','M','N']: ws_c.column_dimensions[cl].width = 16



    ws_c.freeze_panes = 'B14'



    # Arranjos



    ws_arr = wb.create_sheet('Arranjos')



    ws_arr.sheet_properties.tabColor = GREEN_ACCENT



    set_header_row(ws_arr, 1, ['Codigo', 'Bandeira'])



    for i, (code, name) in enumerate(ARRANJO_BANDEIRA.items(), 2):



        ws_arr.cell(row=i, column=1, value=code).font = font_data



        ws_arr.cell(row=i, column=1).border = thin_border



        ws_arr.cell(row=i, column=2, value=name).font = font_data



        ws_arr.cell(row=i, column=2).border = thin_border



    # Sellers ID



    ws_sel = wb.create_sheet('Sellers ID')



    ws_sel.sheet_properties.tabColor = GREEN_ACCENT



    set_header_row(ws_sel, 1, ['CNPJ', 'Seller ID'])



    for i, cnpj in enumerate(sorted(cnpj_groups.keys()), 2):



        sid = seller_map.get(cnpj, '')



        ws_sel.cell(row=i, column=1, value=cnpj).font = font_data



        ws_sel.cell(row=i, column=1).border = thin_border



        ws_sel.cell(row=i, column=2, value=sid).font = font_data



        ws_sel.cell(row=i, column=2).border = thin_border



    wb.save(output_path)



    return len(records), len(cnpj_groups)



def generate_selecao(records, taxa_mensal, operator_email, seller_map, output_path):



    # 4 casas decimais para preservar taxas como 1.055% sem arredondamento
    taxa_str = f"{taxa_mensal:.4f}%"



    with open(output_path, 'w', newline='', encoding='utf-8') as f:



        writer = csv.writer(f)



        writer.writerow(['cnpj','seller_id','cnpj_acquirer','acquirer','arrangement',



                         'settlement_date','available_amount','requested_amount',



                         'monthly_rate','operator_email','platform_id'])



        for r in records:



            cnpj = r.get('cnpj', '')



            sid = seller_map.get(cnpj, seller_map.get(r.get('cnpj_original', ''), ''))



            if not sid:  # Pular URs sem seller_id (nao foram encontrados no Databricks)



                continue



            dl = parse_date(r.get('data_liquidacao', ''))



            data_liq = dl.strftime('%d/%m/%Y') if dl else r.get('data_liquidacao', '')



            disp = r.get('disponivel', 0)

            # Se seleção parcial foi aplicada (valor_alvo), usar o valor cedido parcial
            requested = r.get('_valor_cedido', disp)

            writer.writerow([

                r.get('cnpj_original', cnpj), sid, r.get('cnpj_adquirente', ''),

                r.get('adquirente', ''), r.get('arranjo', ''), data_liq,

                f"{disp:.2f}", f"{requested:.2f}", taxa_str, operator_email,

                r.get('receivable_id', '')])



def generate_inelegiveis_csv(inelig_cnpjs, seller_map, sid_to_cnpjs, records, raiz_to_emp, output_path):



    """Gera CSV com lista de CNPJs/sellers inelegiveis excluidos da cotacao."""



    with open(output_path, 'w', newline='', encoding='utf-8') as f:



        writer = csv.writer(f)



        writer.writerow(['cnpj', 'seller_id', 'empresa', 'urs', 'valor_bruto'])



        sid_cnpjs = {}



        for cnpj in inelig_cnpjs:



            s_id = seller_map.get(cnpj, '')



            if s_id not in sid_cnpjs:



                sid_cnpjs[s_id] = []



            sid_cnpjs[s_id].append(cnpj)



        for s_id, cnpjs in sorted(sid_cnpjs.items()):



            for cnpj in sorted(cnpjs):



                raiz = cnpj[:8] if len(cnpj) >= 8 else ''



                empresa = raiz_to_emp.get(raiz, '')



                urs = sum(1 for r in records if r.get('cnpj', '') == cnpj)



                valor = sum(r.get('disponivel', 0) for r in records if r.get('cnpj', '') == cnpj)



                writer.writerow([cnpj, s_id, empresa, urs, f'{valor:.2f}'])



# ==============================================================================



# OAUTH U2M - Token Management



# ==============================================================================



def save_oauth_token(token_data):



    token_data['saved_at'] = datetime.now().isoformat()



    with open(OAUTH_TOKEN_FILE, 'w') as f:



        json.dump(token_data, f, indent=2)



def load_oauth_token():



    if os.path.exists(OAUTH_TOKEN_FILE):



        with open(OAUTH_TOKEN_FILE, 'r') as f:



            return json.load(f)



    return None



def refresh_oauth_token():



    """Usa refresh_token pra obter novo access_token."""



    token_data = load_oauth_token()



    if not token_data or 'refresh_token' not in token_data:



        return None



    try:



        resp = http_requests.post(OAUTH_TOKEN_URL, data={



            'grant_type': 'refresh_token',



            'client_id': OAUTH_CLIENT_ID,



            'refresh_token': token_data['refresh_token'],



        }, timeout=15)



        if resp.status_code == 200:



            new_data = resp.json()



            # Manter refresh_token se nao veio novo



            if 'refresh_token' not in new_data:



                new_data['refresh_token'] = token_data['refresh_token']



            save_oauth_token(new_data)



            return new_data.get('access_token')



    except Exception:



        pass



    return None



def get_databricks_token():



    """Retorna o melhor token disponivel: OAuth > .env PAT."""



    # 1. Tentar OAuth token salvo



    token_data = load_oauth_token()



    if token_data and 'access_token' in token_data:



        # Testar se ainda funciona



        try:



            resp = http_requests.get(



                f'{DATABRICKS_HOST}/api/2.0/clusters/list',



                headers={'Authorization': f'Bearer {token_data["access_token"]}'},



                timeout=10



            )



            if resp.status_code == 200:



                return token_data['access_token']



        except Exception:



            pass



        # Tentar refresh



        new_token = refresh_oauth_token()



        if new_token:



            return new_token



    # 2. Tentar via Databricks CLI



    cli_token = get_token_via_cli()



    if cli_token:



        return cli_token



    # 3. Fallback pro PAT do .env ? validar antes de usar
    #    PAT expirado (dapi*) retorna 403 e nao deve ser retornado
    if DATABRICKS_TOKEN:
        _pat = DATABRICKS_TOKEN.strip()
        # Se comecar com 'dapi', eh PAT ? testar se ainda e valido
        if _pat.startswith('dapi'):
            try:
                _r = http_requests.get(
                    f'{DATABRICKS_HOST}/api/2.0/clusters/list',
                    headers={'Authorization': f'Bearer {_pat}'},
                    timeout=8
                )
                if _r.status_code == 200:
                    return _pat
                # 403/401 = PAT expirado, ignorar
            except Exception:
                pass
            return None  # PAT invalido ? nao usar
        return _pat  # token nao-dapi (OAuth salvo no .env), retornar sem validar
    return None



# ==============================================================================



# DATABRICKS - SELLER IDs



# ==============================================================================



def get_token_via_cli():



    """Tenta obter token OAuth via Databricks CLI.
    Busca o CLI em múltiplos locais e tenta todos os profiles configurados.
    """

    import subprocess
    import shutil as _shutil
    import configparser as _cp

    # --- Encontrar o CLI ---
    cli_candidates = []
    _local = os.environ.get('LOCALAPPDATA', '')
    _winget = os.path.join(_local, 'Microsoft', 'WinGet', 'Packages')
    if os.path.exists(_winget):
        for _folder in os.listdir(_winget):
            if 'databricks' in _folder.lower():
                _c = os.path.join(_winget, _folder, 'databricks.exe')
                if os.path.exists(_c):
                    cli_candidates.append(_c)
    _in_path = _shutil.which('databricks')
    if _in_path:
        cli_candidates.append(_in_path)
    cli_candidates += [
        os.path.join(_local, 'Programs', 'databricks', 'databricks.exe'),
        'databricks',
        'databricks.exe',
    ]

    # --- Ler profiles do .databrickscfg ---
    _cfg_path = os.path.join(os.environ.get('USERPROFILE', ''), '.databrickscfg')
    _profiles = [None]  # None = sem --profile (usa DEFAULT)
    if os.path.exists(_cfg_path):
        try:
            _cfg = _cp.ConfigParser()
            _cfg.read(_cfg_path, encoding='utf-8')
            for _s in _cfg.sections():
                _host = _cfg.get(_s, 'host', fallback='')
                if 'picpay' in _host or 'databricks' in _host:
                    _profiles.insert(0, _s)  # profiles PicPay têm prioridade
        except Exception:
            pass

    # --- Tentar cada CLI × cada profile ---
    for cli in cli_candidates:
        for profile in _profiles:
            try:
                cmd = [cli, 'auth', 'token', '--host', DATABRICKS_HOST]
                if profile:
                    cmd += ['--profile', profile]
                result = subprocess.run(cmd, capture_output=True, text=True, timeout=15)
                if result.returncode == 0:
                    data = json.loads(result.stdout)
                    token = data.get('access_token')
                    if token:
                        return token
            except Exception:
                continue

    return None



def fetch_empresa_names(raizes, cnpjs_por_raiz=None):



    """Busca nome da empresa via economic_group no sf_accounts pra raizes desconhecidas.



    Retorna dict: {raiz: nome_empresa}



    """



    token = get_databricks_token()



    if not token or not http_requests:



        # Tentar via CLI



        token = get_token_via_cli()



    if not token:



        return {}



    raizes_str = ",".join(f"\'{r}\'" for r in raizes)



    sql = f"""SELECT root_cnpj, 



  COALESCE(



    NULLIF(economic_group, ''),



    NULLIF(account_name, ''),



    NULLIF(corporate_name, '')



  ) as empresa_name



FROM picpay.sales.sf_accounts



WHERE root_cnpj IN ({raizes_str})



  AND COALESCE(NULLIF(economic_group, ''), NULLIF(account_name, ''), NULLIF(corporate_name, '')) IS NOT NULL



GROUP BY root_cnpj, empresa_name



ORDER BY root_cnpj"""



    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}



    try:



        # Timeout agressivo: 8s sync + 4 polls de 3s = max ~20s no upload



        resp = http_requests.post(



            f"{DATABRICKS_HOST}/api/2.0/sql/statements/",



            headers=headers,



            json={"warehouse_id": DATABRICKS_WAREHOUSE_ID, "statement": sql, "wait_timeout": "8s"},



            timeout=12



        )



        data = resp.json()



        state = data.get("status", {}).get("state", "")



        if state == "SUCCEEDED":



            result = {}



            for row in data.get("result", {}).get("data_array", []):



                raiz = str(row[0]).zfill(8)



                nome = str(row[1]).strip()



                if nome.upper().startswith("GRUPO "):



                    nome = nome[6:].strip()



                result[raiz] = nome



            return result



        # Polling curto (warehouse frio): 4x3s = 12s



        stmt_id = data.get("statement_id")



        if stmt_id:



            for _ in range(4):



                time.sleep(3)



                resp = http_requests.get(



                    f"{DATABRICKS_HOST}/api/2.0/sql/statements/{stmt_id}",



                    headers=headers, timeout=8



                )



                r = resp.json()



                if r.get("status", {}).get("state") == "SUCCEEDED":



                    result = {}



                    for row in r.get("result", {}).get("data_array", []):



                        raiz = str(row[0]).zfill(8)



                        nome = str(row[1]).strip()



                        if nome.upper().startswith("GRUPO "):



                            nome = nome[6:].strip()



                        result[raiz] = nome



                    return result



                if r.get("status", {}).get("state") == "FAILED":



                    return {}



        # Warehouse muito frio ou timeout &#8212; retorna vazio, operador nomeia manualmente



        return {}



    except Exception:



        pass



    # Fallback: buscar nome via ReceitaWS por CNPJ (um por raiz)



    return _fetch_empresa_names_web(raizes, cnpjs_por_raiz)



def _fetch_empresa_names_web(raizes, cnpjs_por_raiz=None):



    """Fallback: consulta ReceitaWS para obter nome da empresa por CNPJ.



    Usa um CNPJ de exemplo por raiz (a propria raiz + sufixo fixo nao funciona,



    entao tenta raiz + 000100 e raiz + 000101 como candidatos).



    Retorna dict: {raiz: nome_empresa}



    """



    if not http_requests:



        return {}



    result = {}



    for raiz in raizes:



        # Tentar CNPJ matriz padrao: raiz + 0001 + 01 (filial 0001, digitos 01 genericos)



        # Na pratica buscamos a raiz com sufixo comum



        # Preferir CNPJs reais, depois construidos
        _cmap = cnpjs_por_raiz or {}
        candidates = list(_cmap.get(raiz, []))[:3]
        if not candidates:
            candidates = [raiz + "0001" + "00", raiz + "000100", raiz + "000101"]



        found = False



        for cnpj_try in candidates:



            cnpj_clean = cnpj_try.zfill(14)[:14]



            try:



                resp = http_requests.get(



                    f"https://receitaws.com.br/v1/cnpj/{cnpj_clean}",



                    timeout=8, verify=False



                )



                if resp.status_code == 200:



                    data = resp.json()



                    if data.get("status") != "ERROR":



                        nome = data.get("fantasia") or data.get("nome") or ""



                        nome = nome.strip()



                        if nome:



                            result[raiz] = nome



                            found = True



                            break



            except Exception:



                pass



        if not found:



            # Tentar casa dos dados como segundo fallback



            try:



                resp = http_requests.get(



                    f"https://publica.cnpj.ws/cnpj/{raiz}000100",



                    timeout=8,



                    headers={"User-Agent": "Mozilla/5.0"}



                )



                if resp.status_code == 200:



                    data = resp.json()



                    razao = data.get("razao_social", "")



                    fantasia = (data.get("estabelecimento") or {}).get("nome_fantasia", "")



                    nome = fantasia or razao



                    nome = nome.strip()



                    if nome:



                        result[raiz] = nome



            except Exception:



                pass



    return result



def fetch_seller_ids(raizes, max_retries=2):



    """Busca seller IDs no Databricks via sf_accounts + sf_capture_solutions.



    Otimizado: wait_timeout sync primeiro, async com poll rapido depois.



    Retorna tuple: (seller_map dict, error_message ou None)



    """



    token = get_databricks_token()



    if not token or not DATABRICKS_WAREHOUSE_ID or not http_requests:



        return {}, "Databricks nao conectado. Clique em Conectar ao Databricks no topo da pagina para autenticar via SSO PicPay."



    raizes_str = ",".join(f"\'{r}\'" for r in raizes)



    sql = f"""WITH ranked AS (



    SELECT 



        a.cnpj,



        cs.merchant_seller_id,



        ROW_NUMBER() OVER (



            PARTITION BY a.cnpj 



            ORDER BY 



                CASE WHEN cs.product_name = 'BIZ' THEN 0 ELSE 1 END,



                CAST(cs.merchant_seller_id AS BIGINT)



        ) as rn



    FROM picpay.sales.sf_accounts a



    JOIN picpay.sales.sf_capture_solutions cs ON a.account_id = cs.account_id



    WHERE a.root_cnpj IN ({raizes_str})



      AND cs.capture_solution_status = 'Ativado'



      AND cs.merchant_seller_id IS NOT NULL



      AND cs.merchant_seller_id RLIKE '^[0-9]+$'



)



SELECT cnpj, merchant_seller_id as seller_id



FROM ranked



WHERE rn = 1



ORDER BY cnpj"""



    headers = {



        "Authorization": f"Bearer {token}",



        "Content-Type": "application/json"



    }



    last_error = None



    for attempt in range(1, max_retries + 1):



        try:



            # Tentar sync primeiro (wait_timeout=50s - maximo permitido)



            resp = http_requests.post(



                f"{DATABRICKS_HOST}/api/2.0/sql/statements/",



                headers=headers,



                json={"warehouse_id": DATABRICKS_WAREHOUSE_ID, "statement": sql, "wait_timeout": "50s"},



                timeout=60



            )



            if resp.status_code == 403:



                return {}, "Token Databricks expirado ou invalido. Clique em Conectar ao Databricks no topo da pagina para autenticar via SSO."



            data = resp.json()



            state = data.get("status", {}).get("state", "")



            # Se ja retornou com sucesso (sync)



            if state == "SUCCEEDED":



                seller_map = {}



                for row in data.get("result", {}).get("data_array", []):



                    cnpj = str(row[0]).zfill(14)



                    sid = str(row[1])



                    seller_map[cnpj] = sid



                return seller_map, None



            if state == "FAILED":



                err_msg = data.get("status", {}).get("error", {}).get("message", "Erro na query")



                if "PERMISSION_DENIED" in err_msg:



                    return {}, ("Sem permissao nas tabelas do Databricks. Solicite acesso em: "



                        "https://picpedia.picpay.com/glossario-de-negocios/tables/sales/sf_accounts e "



                        "https://picpedia.picpay.com/glossario-de-negocios/tables/sales/sf_capture_solutions")



                last_error = f"Erro na query: {err_msg}"



                if attempt < max_retries:



                    time.sleep(3)



                    continue



                return {}, f"{last_error} (apos {max_retries} tentativas)"



            # PENDING - precisa poll (warehouse estava frio)



            stmt_id = data.get("statement_id")



            if not stmt_id:



                last_error = "Erro ao submeter query"



                if attempt < max_retries:



                    time.sleep(3)



                    continue



                return {}, last_error



            # Poll rapido: 2s intervalo, max 45 tentativas (~90s)



            for poll in range(45):



                time.sleep(2)



                resp = http_requests.get(



                    f"{DATABRICKS_HOST}/api/2.0/sql/statements/{stmt_id}",



                    headers=headers,



                    timeout=15



                )



                result = resp.json()



                state = result.get("status", {}).get("state", "")



                if state == "SUCCEEDED":



                    seller_map = {}



                    for row in result.get("result", {}).get("data_array", []):



                        cnpj = str(row[0]).zfill(14)



                        sid = str(row[1])



                        seller_map[cnpj] = sid



                    return seller_map, None



                elif state == "FAILED":



                    err_msg = result.get("status", {}).get("error", {}).get("message", "Erro na query")



                    if "PERMISSION_DENIED" in err_msg:



                        return {}, ("Sem permissao nas tabelas do Databricks. Solicite acesso em: "



                            "https://picpedia.picpay.com/glossario-de-negocios/tables/sales/sf_accounts e "



                            "https://picpedia.picpay.com/glossario-de-negocios/tables/sales/sf_capture_solutions")



                    last_error = f"Erro na query: {err_msg}"



                    break



            else:



                last_error = "Query demorou demais no Databricks."



            if attempt < max_retries:



                time.sleep(3)



                continue



        except Exception as e:



            err_str = str(e)



            if "ConnectionError" in str(type(e)) or "connection" in err_str.lower():



                last_error = "Sem conexao com o Databricks. Verifique sua rede/VPN."



            else:



                last_error = f"Erro inesperado: {err_str}"



            if attempt < max_retries:



                time.sleep(5)



                continue



    return {}, f"{last_error} (apos {max_retries} tentativas)"



# ==============================================================================



# ELEGIBILIDADE DE SELLERS



# ==============================================================================



def fetch_eligibility(seller_ids, token=None):



    """Verifica elegibilidade dos seller_ids na tabela sfpj_advance_receivable_elegibilities.



    



    Regra: seller e elegivel se existir na tabela com



    type IN ('advance_receivable', 'external_advance_receivable') E total_amount > 0



    em pelo menos um dos tipos.



    



    Retorna:



        dict: {



            seller_id: {



                'eligible': bool,



                'advance_total': float,



                'external_total': float



            }



        }



    """



    if not seller_ids:



        return {}



    if not token:



        token = get_databricks_token()



    if not token or not http_requests:



        return {}



    ids_str = ",".join(f"'{sid}'" for sid in seller_ids)



    sql = f"""SELECT



    seller_id,



    type,



    MAX(total_amount) as total_amount



FROM picpay.self_service_analytics.sfpj_advance_receivable_elegibilities



WHERE seller_id IN ({ids_str})



  AND type IN ('advance_receivable', 'external_advance_receivable')



GROUP BY seller_id, type"""



    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}



    try:



        resp = http_requests.post(



            f"{DATABRICKS_HOST}/api/2.0/sql/statements/",



            headers=headers,



            json={"warehouse_id": DATABRICKS_WAREHOUSE_ID, "statement": sql, "wait_timeout": "50s"},



            timeout=65



        )



        data = resp.json()



        state = data.get("status", {}).get("state", "")



        rows = []



        if state == "SUCCEEDED":



            rows = data.get("result", {}).get("data_array", [])



        elif state in ("PENDING", "RUNNING"):



            stmt_id = data.get("statement_id")



            if stmt_id:



                for _ in range(30):



                    time.sleep(2)



                    r2 = http_requests.get(



                        f"{DATABRICKS_HOST}/api/2.0/sql/statements/{stmt_id}",



                        headers=headers, timeout=15



                    ).json()



                    s2 = r2.get("status", {}).get("state", "")



                    if s2 == "SUCCEEDED":



                        rows = r2.get("result", {}).get("data_array", [])



                        break



                    elif s2 == "FAILED":



                        return {}



        else:



            return {}



        # Montar resultado por seller_id



        result = {}



        for row in rows:



            sid = str(row[0])



            tipo = str(row[1])



            amount = float(row[2]) if row[2] is not None else 0.0



            if sid not in result:



                result[sid] = {'advance_total': 0.0, 'external_total': 0.0}



            if tipo == 'advance_receivable':



                result[sid]['advance_total'] = amount



            elif tipo == 'external_advance_receivable':



                result[sid]['external_total'] = amount



        # Calcular elegibilidade



        for sid in result:



            adv = result[sid]['advance_total']



            ext = result[sid]['external_total']



            result[sid]['eligible'] = (adv > 0 or ext > 0)



        # Sellers ausentes da tabela = inelegiveis



        for sid in seller_ids:



            if sid not in result:



                result[sid] = {'eligible': False, 'advance_total': 0.0, 'external_total': 0.0, 'absent': True}



        return result



    except Exception as e:



        return {}



# ==============================================================================



# HTML TEMPLATE



# ==============================================================================



HTML_TEMPLATE = '''



<!DOCTYPE html>



<html lang="pt-BR">



<head>



    <meta charset="UTF-8">



    <meta name="viewport" content="width=device-width, initial-scale=1.0">



    <title>SimplificaÊ 🏭 - Antecipação de Recebíveis</title>



    <style>



        * { margin: 0; padding: 0; box-sizing: border-box; }



        body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; background: #f5f5f5; color: #212121; }



        .header { background: linear-gradient(135deg, #1B5E20, #2E7D32); color: white; padding: 20px 40px; display: flex; align-items: center; gap: 16px; }



        .header h1 { font-size: 28px; font-weight: 700; }



        .header .subtitle { opacity: 0.8; font-size: 14px; }



        .tabs { background: white; border-bottom: 2px solid #E0E0E0; padding: 0 40px; display: flex; gap: 0; max-width: 100%; }



        .tab { padding: 14px 24px; font-size: 14px; font-weight: 600; color: #757575; cursor: pointer; border-bottom: 3px solid transparent; transition: all 0.2s; }



        .tab:hover { color: #1B5E20; background: #E8F5E9; }



        .tab.active { color: #1B5E20; border-bottom-color: #1B5E20; }



        .tab-content { display: none; }



        .tab-content.active { display: block; }



        .container { max-width: 1200px; margin: 24px auto; padding: 0 24px; }



        .card { background: white; border-radius: 12px; padding: 24px; margin-bottom: 20px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); }



        .card h2 { color: #1B5E20; margin-bottom: 16px; font-size: 20px; }



        .card h3 { color: #2E7D32; margin-bottom: 12px; }



        .form-group { margin-bottom: 16px; }



        .form-group label { display: block; font-weight: 600; margin-bottom: 6px; color: #424242; }



        .form-group input, .form-group select { width: 100%; padding: 10px 14px; border: 1px solid #E0E0E0; border-radius: 8px; font-size: 14px; }



        .form-group input:focus, .form-group select:focus { outline: none; border-color: #4CAF50; box-shadow: 0 0 0 3px rgba(76,175,80,0.15); }



        .form-row { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 16px; }



        .btn { padding: 12px 24px; border: none; border-radius: 8px; font-size: 14px; font-weight: 600; cursor: pointer; transition: all 0.2s; }



        .btn-primary { background: #1B5E20; color: white; }



        .btn-primary:hover { background: #2E7D32; transform: translateY(-1px); }



        .btn-secondary { background: #E8F5E9; color: #1B5E20; border: 1px solid #C8E6C9; }



        .btn-secondary:hover { background: #C8E6C9; }



        .btn-download { background: #4CAF50; color: white; }



        .btn-download:hover { background: #388E3C; }



        .btn:disabled { opacity: 0.5; cursor: not-allowed; }



        .eligibility-panel { background: #FFF8E1; border: 2px solid #F9A825; border-radius: 12px; padding: 20px; margin-bottom: 20px; }



        .eligibility-panel h3 { color: #E65100; margin-bottom: 12px; font-size: 16px; }



        .eligibility-panel .warning-icon { font-size: 20px; margin-right: 8px; }



        .eligibility-table td { font-size: 12px; padding: 8px 12px; }



        .eligibility-table th { background: #E65100; font-size: 12px; padding: 8px 12px; }



        .eligibility-table tr:nth-child(even) { background: #FFF3E0; }



        .seller-checkbox { width: 16px; height: 16px; cursor: pointer; accent-color: #E65100; }



        .eligibility-actions { margin-top: 16px; display: flex; gap: 12px; align-items: center; flex-wrap: wrap; }



        .eligibility-note { font-size: 12px; color: #6D4C41; margin-top: 8px; }



        .badge-inelegivel { background: #FFCCBC; color: #BF360C; padding: 2px 8px; border-radius: 10px; font-size: 11px; font-weight: 600; }



        .badge-ausente { background: #EF9A9A; color: #B71C1C; padding: 2px 8px; border-radius: 10px; font-size: 11px; font-weight: 600; }



        table { width: 100%; border-collapse: collapse; margin-top: 12px; }



        th { background: #1B5E20; color: white; padding: 10px 14px; text-align: left; font-size: 13px; }



        td { padding: 10px 14px; border-bottom: 1px solid #E0E0E0; font-size: 13px; }



        tr:nth-child(even) { background: #E8F5E9; }



        tr:hover { background: #C8E6C9; }



        .total-row { background: #A5D6A7 !important; font-weight: 700; }



        .badge { display: inline-block; padding: 3px 10px; border-radius: 12px; font-size: 11px; font-weight: 600; }



        .badge-green { background: #E8F5E9; color: #1B5E20; }



        .badge-blue { background: #E3F2FD; color: #1565C0; }



        .upload-area { border: 2px dashed #C8E6C9; border-radius: 12px; padding: 40px; text-align: center; cursor: pointer; transition: all 0.2s; }



        .upload-area:hover { border-color: #4CAF50; background: #E8F5E9; }



        .upload-area.dragover { border-color: #1B5E20; background: #C8E6C9; }



        .upload-area p { color: #616161; margin-top: 8px; }



        .upload-area .icon { font-size: 48px; }



        .status { padding: 12px 16px; border-radius: 8px; margin-top: 12px; }



        .status-success { background: #E8F5E9; color: #1B5E20; border: 1px solid #C8E6C9; }



        .status-error { background: #FFEBEE; color: #C62828; border: 1px solid #FFCDD2; }



        .status-info { background: #E3F2FD; color: #1565C0; border: 1px solid #BBDEFB; }



        .loader { display: inline-block; width: 20px; height: 20px; border: 3px solid #C8E6C9; border-top-color: #1B5E20; border-radius: 50%; animation: spin 0.8s linear infinite; }



        @keyframes spin { to { transform: rotate(360deg); } }



        .hidden { display: none; }



        .text-right { text-align: right; }



        .text-center { text-align: center; }



        .mt-2 { margin-top: 8px; }



        .mt-4 { margin-top: 16px; }



        .gap-2 { display: flex; gap: 8px; align-items: center; }



        .chip { display: inline-block; padding: 4px 12px; background: #E8F5E9; color: #1B5E20; border-radius: 16px; font-size: 12px; margin: 2px; }



    </style>



</head>



<body>



    <div class="header">



        <div>



            <h1>🏭 SimplificaÊ</h1>



            <div class="subtitle">Cotação de Antecipação de Recebíveis</div>



        </div>



        <div style="font-size:11px;color:rgba(255,255,255,0.55);align-self:flex-end;padding-bottom:6px;">v{{ app_version }}</div>



    </div>



    <div class="tabs">



        <div class="tab active" onclick="switchTab('cotacao')">🏭 Cotação</div>



        <div class="tab" onclick="switchTab('historico')">📜 Histórico</div>



    </div>



    <div class="container">



        <!-- TAB: Cotação -->



        <div id="tab-cotacao" class="tab-content active">



        <!-- Step 1: Operador -->



        <div class="card" id="step-operador">



            <h2>1. Operador</h2>



            <div class="form-row">



                <div class="form-group">



                    <label>Quem é o operador?</label>



                    <select id="operador">



                        {% for name, email in operadores.items() %}



                        <option value="{{ email }}">{{ name }} ({{ email }})</option>



                        {% endfor %}



                        <option value="outro">Outro...</option>



                    </select>



                </div>



                <div class="form-group hidden" id="operador-custom">



                    <label>Email do operador</label>



                    <input type="email" id="operador-email" placeholder="email@picpay.com">



                </div>



                <div class="form-group">



                    <label>Servidor Central</label>



                    <div id="server-status" style="padding:4px 0;font-size:12px;color:#9E9E9E;">Verificando...</div>



                </div>



                <div class="form-group">



                    <label>Databricks</label>



                    <div id="db-status" style="padding:8px 0;">



                        <span class="loader"></span> Verificando...



                    </div>



                </div>



            </div>



            <!-- Config E-mail -->



            <div style="margin-top:16px;padding-top:16px;border-top:1px solid #E0E0E0;">



                <label style="font-size:13px;font-weight:600;display:block;margin-bottom:8px;">



                    ✉️ E-mail do Operador (para envio de cotações)



                </label>



                <div style="display:flex;gap:8px;flex-wrap:wrap;align-items:flex-end;">



                    <div>



                        <label style="font-size:11px;color:#9E9E9E;">Gmail corporativo</label>



                        <input type="email" id="cfg-email-user" placeholder="operador@picpay.com" style="padding:6px 10px;border:1px solid #E0E0E0;border-radius:6px;font-size:13px;width:220px;">



                    </div>



                    <div>



                        <label style="font-size:11px;color:#9E9E9E;">App Password</label>



                        <input type="password" id="cfg-email-pass" placeholder="xxxx xxxx xxxx xxxx" style="padding:6px 10px;border:1px solid #E0E0E0;border-radius:6px;font-size:13px;width:180px;">



                    </div>



                    <div>



                        <label style="font-size:11px;color:#9E9E9E;">Nome exibido</label>



                        <input type="text" id="cfg-email-name" value="PicPay AR" style="padding:6px 10px;border:1px solid #E0E0E0;border-radius:6px;font-size:13px;width:140px;">



                    </div>



                    <button onclick="saveEmailConfig()" class="btn btn-secondary" style="padding:6px 14px;font-size:12px;">Salvar e Testar</button>



                </div>



                <div id="cfg-email-status" style="margin-top:8px;font-size:13px;min-height:28px;"></div>
                



                <p style="font-size:11px;color:#9E9E9E;margin-top:4px;">



                    Gere o App Password em: myaccount.google.com > Segurança > Senhas de app



                </p>



            </div>



        </div>



        <!-- HeroDash: Cotacao Rapida -->
        <div class="card" id="card-hd-flow" style="border-left: 4px solid #1a73e8;">
            <h2 style="color:#1a73e8;">&#x26A1; Cotacao Rapida - HeroDash</h2>
            <div id="hd-token-badge" style="display:inline-block; font-size:11px; padding:2px 10px; border-radius:10px; margin-bottom:8px; background:#e0e0e0; color:#666;">Verificando login HeroDash...</div>
            <p style="color:#555; font-size:13px; margin-bottom:12px;">
                Informe as raizes ou CNPJs, configure os e-mails por empresa e gere cotacoes indicativas (taxa 0%).<br>
                <span style="color:#888; font-size:11px;">Maximo 5 raizes por requisicao ao HeroDash. Mais de 5 serao processadas em lotes automaticamente.</span>
            </p>
            <div style="margin-bottom:10px; overflow-x:auto;">
                <table style="width:100%;border-collapse:collapse;font-size:13px;" id="hd-empresas-table">
                    <thead>
                        <tr style="background:#e3f2fd;">
                            <th style="padding:7px 10px;text-align:left;border:1px solid #ccc;width:190px;">Raiz / CNPJ</th>
                            <th style="padding:7px 10px;text-align:left;border:1px solid #ccc;">E-mails do cliente <span style="color:#999;font-weight:normal;">(opcional)</span></th>
                            <th style="padding:7px 10px;border:1px solid #ccc;width:36px;"></th>
                        </tr>
                    </thead>
                    <tbody id="hd-empresas-body">
                        <tr id="hd-row-0">
                            <td style="padding:6px 8px;border:1px solid #e0e0e0;vertical-align:top;">
                                <input type="text" class="hd-cnpj-input" placeholder="17678232" maxlength="18"
                                    style="width:100%;padding:5px 8px;border:1px solid #ccc;border-radius:4px;font-size:13px;box-sizing:border-box;">
                            </td>
                            <td style="padding:6px 8px;border:1px solid #e0e0e0;vertical-align:top;">
                                <div class="hd-emails-list">
                                    <div style="display:flex;gap:6px;margin-bottom:4px;">
                                        <input type="email" class="hd-email-input" placeholder="cliente@empresa.com"
                                            style="flex:1;padding:5px 8px;border:1px solid #ccc;border-radius:4px;font-size:13px;">
                                        <button type="button" onclick="hdRemoveEmail(this)" style="background:none;border:none;color:#999;cursor:pointer;font-size:16px;padding:0 4px;">&#x2715;</button>
                                    </div>
                                </div>
                                <button type="button" onclick="hdAddEmail(this)"
                                    style="font-size:11px;color:#1a73e8;background:none;border:none;cursor:pointer;padding:0;">+ adicionar e-mail</button>
                            </td>
                            <td style="padding:6px;border:1px solid #e0e0e0;text-align:center;vertical-align:middle;">
                                <button type="button" onclick="hdRemoveRow(this)"
                                    style="background:none;border:none;color:#e53935;cursor:pointer;font-size:18px;line-height:1;">&#x2715;</button>
                            </td>
                        </tr>
                    </tbody>
                </table>
                <button type="button" onclick="hdAddRow()"
                    style="margin-top:6px;font-size:12px;color:#1a73e8;background:none;border:1px dashed #90caf9;border-radius:4px;padding:5px 14px;cursor:pointer;">
                    + Adicionar empresa
                </button>
            </div>
            <div style="display:flex;gap:16px;align-items:center;flex-wrap:wrap;margin-top:8px;">
                <label style="display:flex;align-items:center;gap:6px;font-size:13px;cursor:pointer;">
                    <input type="checkbox" id="hd-use-raiz" checked style="width:14px;height:14px;"> Buscar por raiz
                </label>
                <label style="display:flex;align-items:center;gap:6px;font-size:13px;cursor:pointer;">
                    <input type="checkbox" id="hd-enviar-email" style="width:14px;height:14px;"> Enviar e-mail para cada empresa
                </label>
                <button id="btn-hd-cotacao" onclick="hdCotacaoRapida()"
                        style="background:#1a73e8;color:#fff;border:none;padding:10px 24px;border-radius:6px;font-size:15px;cursor:pointer;font-weight:600;">
                    &#x26A1; Gerar Cotacoes
                </button>
                <span id="hd-status-msg" style="font-size:13px;color:#555;"></span>
            </div>
            <div id="hd-result" style="margin-top:14px;display:none;">
                <div style="padding:12px;background:#e8f5e9;border-radius:6px;font-size:13px;" id="hd-result-body"></div>
            </div>
        </div>
        <script>
        (function(){
            fetch('/herodash/token_status').then(function(r){return r.json();}).then(function(d){
                var b=document.getElementById('hd-token-badge'); if(!b) return;
                if(d.ok){var m=Math.round((d.expires_in_seconds||0)/60);b.style.background='#c8e6c9';b.style.color='#2e7d32';b.textContent='HeroDash conectado (expira em '+m+' min)';}
                else{
  b.style.background='#ffccbc';b.style.color='#bf360c';
  b.innerHTML='HeroDash desconectado &nbsp;'
    +'<button onclick="hdRenovarToken()" style="background:#bf360c;color:white;border:none;border-radius:4px;padding:2px 10px;font-size:11px;cursor:pointer;font-weight:600;">Renovar token</button>';
}
            }).catch(function(){var b=document.getElementById('hd-token-badge');if(b)b.textContent='Status HeroDash indisponivel';});
        })();
        function hdRenovarToken() {
    var w = window.open('https://herodash.picpay.com/login', '_blank', 'width=900,height=650');
    if (!w) { alert('Popup bloqueado. Abra https://herodash.picpay.com/login manualmente e faça login.'); return; }
    var b = document.getElementById('hd-token-badge');
    if (b) b.innerHTML = 'Aguardando login no HeroDash... <button onclick="hdVerificarTokenAposLogin()" style="background:#1B5E20;color:white;border:none;border-radius:4px;padding:2px 10px;font-size:11px;cursor:pointer;font-weight:600;">Já fiz login</button>';
}
function hdVerificarTokenAposLogin() {
    fetch('/herodash/token_status').then(function(r){return r.json();}).then(function(d){
        var b = document.getElementById('hd-token-badge');
        if (!b) return;
        if (d.ok) {
            var m = Math.round((d.expires_in_seconds||0)/60);
            b.style.background='#c8e6c9'; b.style.color='#2e7d32';
            b.textContent='HeroDash conectado (expira em '+m+' min)';
        } else {
            b.innerHTML='Token ainda inválido. <button onclick="hdRenovarToken()" style="background:#bf360c;color:white;border:none;border-radius:4px;padding:2px 10px;font-size:11px;cursor:pointer;">Tentar novamente</button>';
        }
    });
}
var _hdRowCount=1;
        function hdAddRow(){
            var idx=_hdRowCount++; var tbody=document.getElementById('hd-empresas-body');
            var tr=document.createElement('tr'); tr.id='hd-row-'+idx;
            tr.innerHTML='<td style="padding:6px 8px;border:1px solid #e0e0e0;vertical-align:top;">'
                +'<input type="text" class="hd-cnpj-input" placeholder="17678232" maxlength="18"'
                +'<input type="text" class="hd-cnpj-input" placeholder="17678232 (8 digitos)" maxlength="18"'
                +' style="width:100%;padding:5px 8px;border:1px solid #ccc;border-radius:4px;font-size:13px;box-sizing:border-box;"></td>'
                +'<td style="padding:6px 8px;border:1px solid #e0e0e0;vertical-align:top;">'
                +'<div class="hd-emails-list"><div style="display:flex;gap:6px;margin-bottom:4px;">'
                +'<input type="email" class="hd-email-input" placeholder="cliente@empresa.com"'
                +' style="flex:1;padding:5px 8px;border:1px solid #ccc;border-radius:4px;font-size:13px;">'
                +'<button type="button" onclick="hdRemoveEmail(this)" style="background:none;border:none;color:#999;cursor:pointer;font-size:16px;padding:0 4px;">&#x2715;</button>'
                +'</div></div>'
                +'<button type="button" onclick="hdAddEmail(this)" style="font-size:11px;color:#1a73e8;background:none;border:none;cursor:pointer;padding:0;">+ adicionar e-mail</button>'
                +'</td><td style="padding:6px;border:1px solid #e0e0e0;text-align:center;vertical-align:middle;">'
                +'<button type="button" onclick="hdRemoveRow(this)" style="background:none;border:none;color:#e53935;cursor:pointer;font-size:18px;line-height:1;">&#x2715;</button></td>';
            tbody.appendChild(tr);
        }
        function hdRemoveRow(btn){
            var tr=btn.closest('tr'); var tbody=document.getElementById('hd-empresas-body');
            if(tbody.rows.length>1){tr.remove();}else{tr.querySelectorAll('input').forEach(function(i){i.value='';});}
        }
        function hdAddEmail(btn){
            var list=btn.previousElementSibling;
            var div=document.createElement('div'); div.style.cssText='display:flex;gap:6px;margin-bottom:4px;';
            div.innerHTML='<input type="email" class="hd-email-input" placeholder="cliente@empresa.com"'
                +' style="flex:1;padding:5px 8px;border:1px solid #ccc;border-radius:4px;font-size:13px;">'
                +'<button type="button" onclick="hdRemoveEmail(this)" style="background:none;border:none;color:#999;cursor:pointer;font-size:16px;padding:0 4px;">&#x2715;</button>';
            list.appendChild(div);
        }
        function hdRemoveEmail(btn){
            var div=btn.parentElement; var list=div.parentElement;
            if(list.children.length>1){div.remove();}else{div.querySelector('input').value='';}
        }
        function hdColetarEmpresas(){
            var rows=document.querySelectorAll('#hd-empresas-body tr'); var empresas=[];
            rows.forEach(function(tr){
                var inp=tr.querySelector('.hd-cnpj-input');
                var cnpj=inp?(inp.value||'').trim():''; if(!cnpj) return;
                var emails=[];
                tr.querySelectorAll('.hd-email-input').forEach(function(i){var v=i.value.trim();if(v)emails.push(v);});
                empresas.push({cnpj:cnpj,emails:emails});
            });
            return empresas;
        }
        function hdCotacaoRapida(){
            var empresas=hdColetarEmpresas();
            if(!empresas.length){alert('Informe pelo menos um CNPJ ou raiz.');return;}
            var useRaiz=document.getElementById('hd-use-raiz').checked;
            var enviarEmail=document.getElementById('hd-enviar-email').checked;
            var opEl=document.getElementById('operador'); var opEmail=opEl?opEl.value:'';
            if(opEmail==='outro'){var oc=document.getElementById('operador-email');opEmail=oc?oc.value:'';}
            var btn=document.getElementById('btn-hd-cotacao');
            var msg=document.getElementById('hd-status-msg');
            var res=document.getElementById('hd-result');
            var resBody=document.getElementById('hd-result-body');
            btn.disabled=true;
            msg.textContent='Processando '+empresas.length+' empresa(s)... aguarde';
            res.style.display='none';
            fetch('/herodash/cotacao_rapida',{
                method:'POST',headers:{'Content-Type':'application/json'},
                body:JSON.stringify({empresas:empresas,use_raiz:useRaiz,operator_email:opEmail,enviar_email:enviarEmail})
            })
            .then(function(r){return r.json().then(function(d){d._status=r.status;return d;});})
            .then(function(d){
                btn.disabled=false;
                if(d.error){
                    var isTokenErr = d._status===401 || d._status===403
                        || !!(d.error && (d.error.indexOf('Token')>=0||d.error.indexOf('token')>=0||d.error.indexOf('expirado')>=0||d.error.indexOf('login')>=0));
                    if(isTokenErr){
                        var b=document.getElementById('hd-token-badge');
                        if(b){b.style.background='#ffccbc';b.style.color='#bf360c';
                            b.innerHTML='Token HeroDash expirado \u00a0<button onclick="hdRenovarToken()" style="background:#bf360c;color:white;border:none;border-radius:4px;padding:2px 10px;font-size:11px;cursor:pointer;font-weight:600;">Renovar token</button>';}
                        msg.innerHTML='<span style="color:#bf360c;font-weight:600;">&#9888; Token HeroDash expirado.</span> Renove acima e tente novamente.';
                    } else { msg.textContent='Erro: '+d.error; }
                    return;
                }
                msg.textContent='Cotacoes geradas com sucesso!';
                var html='<b>Cotacao Rapida HD</b><br><br>';
                html+='Registros totais: <b>'+d.records+'</b> URs | Taxa: <b>'+d.taxa_pct+'%</b><br><br>';
                if(d.empresas&&d.empresas.length>0){
                    html+='<b>Empresas:</b><ul style="margin:6px 0 8px 16px;">';
                    d.empresas.forEach(function(e){
                        if(e.error){html+='<li style="color:#c62828;">ERRO '+e.empresa+': '+e.error+'</li>';}
                        else{
                            html+='<li>'+e.empresa+' - '+e.urs+' URs';
                            html+=' <a href="/download/'+d.session_id+'/'+encodeURIComponent(e.empresa)+'" target="_blank" style="color:#1a73e8;">[baixar]</a>';
                            if(e.email_enviado) html+=' <span style="color:#2e7d32;">&#10003; e-mail enviado</span>';
                            if(e.email_erro)    html+=' <span style="color:#c62828;">&#x26A0; '+e.email_erro+'</span>';
                            html+='</li>';
                        }
                    });
                    html+='</ul>';
                }
                html+='<a href="/download_all/'+d.session_id+'" style="color:#1a73e8;font-weight:600;">Baixar todos (.zip)</a>';
                resBody.innerHTML=html; res.style.display='block';
            })
            .catch(function(err){btn.disabled=false;msg.textContent='Erro: '+err.message;});
        }
        </script>

        <!-- Step 2: Upload -->



        <div class="card">



            <h2>2. Upload da Agenda</h2>



            <div class="upload-area" id="upload-area" onclick="document.getElementById('file-input').click()">



                <div class="icon">📁</div>



                <p><strong>Clique ou arraste</strong> um ou mais arquivos CSV da agenda em lote</p>



                <p style="font-size: 12px; color: #9E9E9E;">Aceita múltiplos arquivos | Separador ; ou , | Encoding UTF-8 ou Latin-1</p>



            </div>



            <input type="file" id="file-input" accept=".csv" multiple class="hidden">



            <div id="upload-status" class="hidden"></div>



        </div>



        <!-- Step 3: Resumo -->



        <div class="card hidden" id="step-resumo">



            <h2>3. Resumo da Agenda</h2>



            <div id="resumo-content"></div>



        </div>



        <!-- Step: Raizes desconhecidas -->



        <div class="card hidden" id="step-unknown"></div>



        <!-- Step 4: Parametros -->



        <div class="card hidden" id="step-params">



            <h2>4. Parâmetros</h2>



            <div class="form-row">



                <div class="form-group">



                    <label>Taxa Mensal (%)</label>



                    <input type="number" id="taxa" step="0.01" min="0" max="10" placeholder="1.33"

                    oninput="syncTaxaEmpresas(this.value)">



                </div>



                <div class="form-group">



                    <label>DI Período (% a.a.)</label>



                    <input type="number" id="di-periodo" step="0.001" value="14.65" placeholder="14.65">



                </div>



            </div>



            <div style="margin:4px 0 8px;">



                <label style="font-size:12px;color:#616161;cursor:pointer;">



                    <input type="checkbox" id="taxa-individual" onchange="toggleTaxasIndividuais(this.checked)">



                    Definir taxa diferente por empresa <span style="color:#9E9E9E;">(marque aqui, ajuste as taxas e depois clique em Gerar)</span>



                </label>



            </div>



            <div id="taxas-por-empresa-wrap" style="display:none;border:1px solid #E0E0E0;border-radius:6px;padding:12px;margin-bottom:8px;">



                <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px;">



                    <strong style="font-size:13px;color:#424242;">Taxa por empresa</strong>



                    <button type="button" class="btn btn-secondary" style="padding:3px 10px;font-size:11px;"

                        onclick="resetTaxasEmpresas()">Igualar todas</button>



                </div>



                <div id="taxas-empresas-list"><p style="font-size:12px;color:#E65100;background:#FFF3E0;padding:8px;border-radius:4px;margin:0;">⚠️ Defina as taxas aqui <strong>antes</strong> de clicar em “Gerar Cotações”. Após a primeira geração, as empresas aparecerão aqui para ajuste.</p></div>



            </div>



            <div class="form-group">



                <label>Arquivos a gerar</label>



                <div style="display:flex;gap:16px;flex-wrap:wrap;">



                    <label style="cursor:pointer;"><input type="checkbox" id="gen-cotacao" checked> Cotação Elegíveis (XLSX)</label>



                    <label style="cursor:pointer;"><input type="checkbox" id="gen-selecao" checked> Seleção de URs (CSV)</label>



                    <label style="cursor:pointer;" title="Gera também a cotação com todos os sellers, marcando inlegíveis em vermelho"><input type="checkbox" id="gen-completo" checked> Cotação Completa c/ Ineligíveis (XLSX)</label>



                </div>



            </div>



            <div class="mt-4">



                <button class="btn btn-primary" onclick="generateAll()">



                    🏭 Gerar Cotacoes (Todas as Empresas)



                </button>



            </div>



        </div>



        <!-- Step 5: Resultados -->



        <div class="card hidden" id="step-results">



            <h2>5. Cotações Geradas</h2>



            <div id="results-content"></div>



            <div class="mt-4 gap-2">



                <button class="btn btn-download" onclick="downloadAll()">⬇️ Download Tudo (ZIP)</button>

                <button id="btn-send-all" onclick="openSendAllModal()" style="background:#E3F2FD;border:1px solid #90CAF9;color:#1565C0;padding:8px 18px;font-size:13px;border-radius:6px;cursor:pointer;display:none;">&#9993; Enviar todos por e-mail</button>



            </div>



        </div>



        <!-- Step 6: Calculadora AR -->
        <div class="card hidden" id="step-calc-ar" style="border-left:4px solid #1B5E20;">
            <h2>6. Calculadora de Antecipa\u00e7\u00e3o</h2>
            <p style="color:#555;font-size:13px;margin-bottom:10px;">Indicadores de rentabilidade por empresa. Ajuste a taxa individualmente ou aplique a todas de uma vez.</p>
            <div style="display:flex;flex-wrap:wrap;gap:10px;align-items:center;margin-bottom:14px;padding-bottom:12px;border-bottom:1px solid #E8F5E9;">
                <span style="font-size:12px;color:#555;font-weight:600;">Aplicar taxa a todas:</span>
                <button class="btn btn-secondary" style="padding:5px 11px;font-size:12px;" onclick="setArTaxaAll(1.10)">1,10%</button>
                <button class="btn btn-secondary" style="padding:5px 11px;font-size:12px;" onclick="setArTaxaAll(1.22)">1,22%</button>
                <button class="btn btn-secondary" style="padding:5px 11px;font-size:12px;" onclick="setArTaxaAll(1.37)">1,37%</button>
                <button class="btn btn-secondary" style="padding:5px 11px;font-size:12px;" onclick="setArTaxaAll(1.50)">1,50%</button>
                
                <input type="number" id="ar-taxa-default" value="1.37" min="0.01" max="50" step="0.01" style="width:75px;padding:4px 8px;border:1px solid #ccc;border-radius:5px;font-size:12px;display:none;">
            </div>
            <div id="ar-inputs" style="display:none;"></div>
            <div id="ar-results" style="min-height:40px;overflow-x:auto;"></div>
        </div>

        <!-- Step 7: Personalizar -->
        <div class="card hidden" id="step-custom">



            <h2>6. Cotação Personalizada</h2>



            <div class="form-row">



                <div class="form-group">



                    <label>Empresa</label>



                    <select id="custom-empresa" onchange="loadEmpresaFilters()"></select>



                </div>



                <div class="form-group">



                    <label>Valor Alvo (R$) <span style="color:#9E9E9E;font-size:11px;font-weight:normal;">ex: 5M, 500K, 1.5M</span></label>



                    <input type="text" id="custom-valor" placeholder="ex: 5M ou 50000000" oninput="_previewValor()">



                    <div id="custom-valor-preview" style="font-size:12px;color:#1B5E20;min-height:16px;margin-top:2px;"></div>



                    <div style="display:flex;flex-wrap:wrap;gap:5px;margin-top:5px;">



                        <button type="button" onclick="_setValorCustom(500000)" class="btn btn-secondary" style="padding:2px 8px;font-size:11px;">500K</button>



                        <button type="button" onclick="_setValorCustom(1000000)" class="btn btn-secondary" style="padding:2px 8px;font-size:11px;">1M</button>



                        <button type="button" onclick="_setValorCustom(5000000)" class="btn btn-secondary" style="padding:2px 8px;font-size:11px;">5M</button>



                        <button type="button" onclick="_setValorCustom(10000000)" class="btn btn-secondary" style="padding:2px 8px;font-size:11px;">10M</button>



                        <button type="button" onclick="_setValorCustom(50000000)" class="btn btn-secondary" style="padding:2px 8px;font-size:11px;">50M</button>



                        <button type="button" onclick="_setValorCustom(100000000)" class="btn btn-secondary" style="padding:2px 8px;font-size:11px;">100M</button>



                        <button type="button" onclick="_setValorCustom(0)" class="btn btn-secondary" style="padding:2px 8px;font-size:11px;background:#FFF8E1;border-color:#F9A825;">Tudo</button>



                    </div>



                </div>



                <div class="form-group">



                    <label>Adquirente <span style="color:#9E9E9E;font-size:11px;font-weight:normal;">(múltipla seleção)</span></label>



                    <div id="custom-adquirente-wrap" style="display:flex;flex-wrap:wrap;gap:6px;min-height:28px;align-items:center;">



                        <span style="color:#9E9E9E;font-size:12px;">Selecione a empresa primeiro</span>



                    </div>



                </div>



                <div class="form-group">



                    <label>Arranjo <span style="color:#9E9E9E;font-size:11px;font-weight:normal;">(múltipla seleção)</span></label>



                    <div id="custom-arranjo-wrap" style="display:flex;flex-wrap:wrap;gap:6px;min-height:28px;align-items:center;">



                        <span style="color:#9E9E9E;font-size:12px;">Selecione a empresa primeiro</span>



                    </div>



                </div>



                <div class="form-group">



                    <label>Taxa Mensal (%) <span style="color:#9E9E9E;font-weight:normal;">vazio = usa a taxa geral</span></label>



                    <input type="number" id="custom-taxa" step="0.01" min="0" max="10" placeholder="Mesma taxa geral">



                </div>



            </div>



            <div id="empresa-info" class="hidden" style="margin-bottom:12px;">



                <span class="badge badge-green" id="empresa-info-urs"></span>



                <span class="badge badge-blue" id="empresa-info-valor"></span>



            

                    <span class="badge" id="empresa-info-operavel" style="display:none;background:#E8F5E9;color:#1B5E20;border:1.5px solid #4CAF50;"></span></div>



            <div class="form-group" id="custom-cnpj-group" style="display:none;">

                    <label style="display:flex;justify-content:space-between;align-items:center;">

                        <span>CNPJ</span>

                        <span id="cnpj-count" style="font-size:11px;color:#1B5E20;font-weight:600;"></span>

                    </label>

                    <input type="text" id="cnpj-search"

                        placeholder="Buscar CNPJ..."

                        oninput="filterCnpjs(this.value)"

                        style="width:100%;box-sizing:border-box;padding:6px 10px;border:1px solid #E0E0E0;border-radius:6px;font-size:12px;margin-bottom:4px;">

                    <select id="custom-cnpj-select" multiple size="5"

                        onchange="updateCnpjCount(); updateFiltersForCnpj();"

                        style="width:100%;border:1px solid #E0E0E0;border-radius:6px;font-family:monospace;font-size:12px;padding:4px;color:#212121;">

                    </select>

                    <div style="display:flex;gap:6px;margin-top:4px;">

                        <button type="button" onclick="selectAllCnpjs(true)" style="flex:1;font-size:11px;padding:3px;border:1px solid #E0E0E0;border-radius:4px;background:white;cursor:pointer;">✓ Todos</button>

                        <button type="button" onclick="selectAllCnpjs(false)" style="flex:1;font-size:11px;padding:3px;border:1px solid #E0E0E0;border-radius:4px;background:white;cursor:pointer;">□ Nenhum</button>

                    </div>

                </div>

                <div class="form-group">



                <label>Datas de Liquidação</label>



                <div class="form-row" style="align-items:center;">



                    <div>



                        <select id="custom-datas-mode" onchange="toggleDatasCustom()">



                            <option value="todas">Todas as datas</option>



                            <option value="range">Intervalo (de/ate)</option>



                            <option value="select">Selecionar datas</option>



                        </select>



                    </div>



                    <div id="datas-range" class="hidden" style="display:flex;gap:8px;align-items:center;">



                        <input type="text" id="custom-data-de" placeholder="dd/mm/aaaa" style="width:140px;">



                        <span>ate</span>



                        <input type="text" id="custom-data-ate" placeholder="dd/mm/aaaa" style="width:140px;">



                    </div>



                </div>



                <div id="datas-select" class="hidden mt-2" style="max-height:200px;overflow-y:auto;border:1px solid #E0E0E0;border-radius:8px;padding:8px;">



                    <div class="gap-2 mt-2" style="margin-bottom:8px;">



                        <button class="btn btn-secondary" style="padding:4px 12px;font-size:11px;" onclick="toggleAllDatas(true)">Selecionar Todas</button>



                        <button class="btn btn-secondary" style="padding:4px 12px;font-size:11px;" onclick="toggleAllDatas(false)">Limpar</button>



                    </div>



                    <div id="datas-checkboxes"></div>



                </div>



                <div id="datas-summary" class="mt-2" style="font-size:12px;color:#616161;"></div>



            </div>



            <div class="mt-4">



                <button class="btn btn-secondary" onclick="generateCustom()">📋 Gerar Personalizada</button>



            </div>



            <div id="custom-result" class="hidden mt-4"></div>



        </div>



        </div><!-- fim tab-cotacao -->



        <!-- TAB: Histórico -->



        <div id="tab-historico" class="tab-content">



            <div class="card">



                <h2>📜 Histórico de Cotações</h2>



                <p style="color:#757575;margin-bottom:16px;">Registro de todas as cotações geradas por todos os operadores.</p>



                <div style="margin-bottom:12px;">



                    <button class="btn btn-secondary" onclick="loadHistory()" style="margin-right:8px;">🔄 Atualizar</button>



                    <span id="history-sync-status" style="font-size:12px;color:#9E9E9E;"></span>



                </div>



                <div id="history-content"><p style="color:#9E9E9E;">Carregando...</p></div>



            </div>



        </div><!-- fim tab-historico -->



    </div>



    <script>



        // Auto-update check



        function checkForUpdates() {
            fetch('/check_update')
            .then(r => r.json())
            .then(function(data) {
                var temUpdate = data.has_update || data.update_available;
                var remote = data.remote_version || '';
                var local  = data.local_version  || '';
                if (!temUpdate) {
                    var el = document.getElementById('update-banner');
                    if (el) { el.remove(); document.body.style.paddingTop = ''; }
                    return;
                }
                if (document.getElementById('update-banner')) return;
                var banner = document.createElement('div');
                banner.id = 'update-banner';
                banner.style.cssText = 'position:fixed;top:0;left:0;right:0;z-index:9999;'
                    + 'background:linear-gradient(90deg,#1B5E20,#2E7D32);color:white;'
                    + 'padding:10px 24px;display:flex;align-items:center;'
                    + 'justify-content:space-between;box-shadow:0 2px 8px rgba(0,0,0,0.3);font-size:14px;';
                var localTxt = local ? ' &mdash; voc&#234; est&#225; na <strong>' + local + '</strong>' : '';
                banner.innerHTML =
                    '<span>'
                    + '<strong style="font-size:15px;">&#128260; Nova vers&#227;o dispon&#237;vel!</strong>'
                    + '<span style="margin-left:12px;opacity:0.9;">Vers&#227;o <strong>' + remote + '</strong> pronta' + localTxt + '</span>'
                    + '</span>'
                    + '<div style="display:flex;gap:8px;align-items:center;">'
                    + '<button id="btn-update" onclick="applyUpdate()" style="background:white;color:#1B5E20;border:none;border-radius:6px;padding:6px 18px;font-size:13px;font-weight:bold;cursor:pointer;">&#11015; Atualizar agora</button>'
                    + '<button onclick="document.getElementById(\"update-banner\").remove();document.body.style.paddingTop=\"\";" style="background:transparent;color:rgba(255,255,255,0.7);border:1px solid rgba(255,255,255,0.3);border-radius:6px;padding:6px 12px;font-size:12px;cursor:pointer;">Agora n&#227;o</button>'
                    + '</div>';
                document.body.prepend(banner);
                document.body.style.paddingTop = '52px';
            })
            .catch(function() {});
        }

                function _updateBannerProgress(pct, status) {
            const btn = document.getElementById('btn-update');
            if (!btn) return;
            if (status === 'downloading') {
                btn.textContent = 'Baixando... ' + pct + '%';
                btn.disabled = true;
            } else if (status === 'ready') {
                btn.textContent = '✅ Reiniciando...';
                btn.disabled = true;
            } else if (status === 'error') {
                btn.textContent = '❌ Erro — tente novamente';
                btn.disabled = false;
            }
        }



        function restartApp() {
            if (!confirm('Reiniciar o Simplifica? para carregar a vers?o mais recente?')) return;
            const btn = document.getElementById('btn-restart');
            if (btn) { btn.disabled=true; btn.textContent='Reiniciando...'; }
            fetch('/restart', {method: 'POST'})
            .then(() => {
                setTimeout(() => location.reload(), 3000);
            })
            .catch(() => setTimeout(() => location.reload(), 3000));
        }

        function applyUpdate() {
            fetch('/apply_update', {method: 'POST'})
            .then(r => r.json())
            .then(data => {
                if (data.status === 'download iniciado') {
                    // Polling do progresso a cada 1s
                    const poll = setInterval(() => {
                        fetch('/check_update').then(r => r.json()).then(s => {
                            _updateBannerProgress(s.download_progress || 0, s.download_status || '');
                            if (s.download_status === 'ready') {
                                clearInterval(poll);
                                // App vai reiniciar sozinho via update.bat — avisa operador
                                setTimeout(() => {
                                    alert('Atualização concluída! O app vai reiniciar automaticamente. Aguarde alguns segundos e reabra o SimplificaÊ.');
                                }, 1500);
                            } else if (s.download_status === 'error') {
                                clearInterval(poll);
                            }
                        });
                    }, 1000);
                } else if (data.status === 'ok') {
                    alert('Atualizado! A página vai recarregar.');
                    location.reload();
                } else {
                    alert('Erro: ' + (data.error || data.status || 'desconhecido'));
                }
            })
            .catch(e => alert('Erro ao iniciar atualização: ' + e));
        }



        // Helpers multi-select checkboxes



        function buildCheckboxes(prefix, items, labelMap) {



            if (!items || items.length === 0) return '<span style="color:#9E9E9E;font-size:12px;">Nenhum</span>';



            return items.map(function(v) {



                const label = labelMap && labelMap[v] ? v + ' (' + labelMap[v] + ')' : v;



                const id = prefix + '_' + v.replace(/[^a-zA-Z0-9]/g,'_');



                return '<label style="display:flex;align-items:center;gap:5px;font-size:12px;cursor:pointer;padding:2px 0;">'



                    + '<input type="checkbox" id="'+id+'" value="'+v+'" checked style="cursor:pointer;"> '+label+'</label>';



            }).join('');



        }



        function getCheckedValues(wrapId) {



            const boxes = document.querySelectorAll('#'+wrapId+' input[type=checkbox]');



            if (!boxes.length) return '';



            const checked = Array.from(boxes).filter(b => b.checked).map(b => b.value);



            const all = Array.from(boxes).map(b => b.value);



            if (checked.length === 0 || checked.length === all.length) return '';



            return checked.join(',');



        }



        // Preview do valor alvo com multiplicador



        function updateValorPreview() {



            const raw = parseFloat(document.getElementById('custom-valor').value) || 0;



            const mult = parseInt(document.getElementById('custom-valor-mult').value) || 1;



            const total = raw * mult;



            const el = document.getElementById('custom-valor-preview');



            if (el) el.textContent = raw > 0 ? '= R$ ' + total.toLocaleString('pt-BR', {minimumFractionDigits:2, maximumFractionDigits:2}) : '';



        }



        document.addEventListener('DOMContentLoaded', function() {



            var vi = document.getElementById('custom-valor');



            var vm = document.getElementById('custom-valor-mult');



            if (vi) vi.addEventListener('input', updateValorPreview);



            if (vm) vm.addEventListener('change', updateValorPreview);



        });



        checkForUpdates();
        setInterval(checkForUpdates, 2 * 60 * 1000);



        // Server status



        function checkServerStatus() {



            fetch('/server_status')



            .then(r => r.json())



            .then(data => {



                const el = document.getElementById('server-status');



                if (data.connected) {



                    el.innerHTML = '<span class="badge badge-green">Conectado (' + data.url + ')</span>';



                } else if (data.is_server) {



                    el.innerHTML = '<span class="badge badge-blue">Modo Servidor (este e o central)</span>';



                } else {



                    el.innerHTML = '<span style="color:#9E9E9E;">Local (sem servidor central)</span>';



                }



            })



            .catch(() => {});



        }



        // Tab switching



        function switchTab(tab) {



            document.querySelectorAll('.tab').forEach(t => t.classList.remove('active'));



            document.querySelectorAll('.tab-content').forEach(t => t.classList.remove('active'));



            document.getElementById('tab-' + tab).classList.add('active');



            document.querySelectorAll('.tab')[tab === 'cotacao' ? 0 : 1].classList.add('active');



            if (tab === 'historico') loadHistory();



        }



        // Operador toggle



        document.getElementById('operador').addEventListener('change', function() {



            document.getElementById('operador-custom').classList.toggle('hidden', this.value !== 'outro');



        });



        // Databricks status check



        // full=false: check leve (token local), instantaneo



        // full=true:  SELECT 1 real no Databricks, ~5-10s



        function checkDbStatus(full) {



            const el = document.getElementById('db-status');



            if (full) {



                el.innerHTML = '<span class="loader"></span> Testando conexao real...';



            }



            fetch('/oauth/status' + (full ? '?full=1' : '?full=0'))



            .then(r => r.json())



            .then(data => {



                if (data.connected) {



                    const method = data.method === 'oauth' ? 'OAuth' : 'PAT';



                    const testedLabel = data.full_check ? ' (testado)' : '';



                    el.innerHTML = '<span class="badge badge-green">Conectado (' + method + testedLabel + ')</span> '



                        + '<button onclick="checkDbStatus(true)" class="btn btn-secondary" style="padding:3px 10px;font-size:11px;margin-left:6px;">Testar</button>'



                        + ' <a href="/oauth/login" class="btn btn-secondary" style="padding:3px 10px;font-size:11px;margin-left:4px;">Reconectar</a>'




                } else {



                    let msg = '<span style="color:#C62828;">Desconectado</span>';



                    if (data.error) {



                        msg += ' <span style="color:#9E9E9E;font-size:11px;">(' + data.error + ')</span>';



                    }



                    msg += '<br><a href="/oauth/login" class="btn btn-primary" style="padding:6px 14px;font-size:12px;margin-top:6px;">Conectar ao Databricks</a>';



                    msg += '<p style="font-size:11px;color:#757575;margin-top:4px;">Clique para autenticar via SSO PicPay (nao precisa de token manual)</p>';



                    el.innerHTML = msg;



                }



            })



            .catch(() => {



                el.innerHTML = '<span style="color:#9E9E9E;">Sem verificacao</span>'



                    + '<br><a href="/oauth/login" class="btn btn-primary" style="padding:6px 14px;font-size:12px;margin-top:6px;">Conectar ao Databricks</a>';



            });



        }



        function disconnectDb() {



            fetch('/oauth/disconnect').then(() => checkDbStatus(false));



        }



        // ================================================================



        // E-MAIL: modal + config + historico de envios



        // ================================================================



        





        // Delegacao de clique para botoes de email (usa data-attributes, sem escapes)

        document.addEventListener('click', function(ev) {

            // Botao email cotacao personalizada

            var btnCustom = ev.target.closest('.btn-email-action-custom');

            if (btnCustom) {

                var sid      = btnCustom.getAttribute('data-sid');

                var safe     = btnCustom.getAttribute('data-safe');

                var nome     = btnCustom.getAttribute('data-nome');

                var urs      = parseInt(btnCustom.getAttribute('data-urs')) || 0;

                var valor    = parseFloat(btnCustom.getAttribute('data-valor')) || 0;

                var operavel = parseFloat(btnCustom.getAttribute('data-operavel')) || 0;

                var taxa     = parseFloat(btnCustom.getAttribute('data-taxa')) || 0;

                openEmailModalCustom(sid, safe, nome, urs, valor, operavel, taxa);

                return;

            }

            var btn = ev.target.closest('.btn-email-action');

            if (!btn) return;

            var sid      = btn.getAttribute('data-sid');

            var safe     = btn.getAttribute('data-safe');

            var nome     = btn.getAttribute('data-nome');

            var urs      = parseInt(btn.getAttribute('data-urs')) || 0;

            var valor    = parseFloat(btn.getAttribute('data-valor')) || 0;

            var operavel = parseFloat(btn.getAttribute('data-operavel')) || 0;

            var taxa     = parseFloat(btn.getAttribute('data-taxa')) || 0;

            openEmailModal(sid, safe, nome, urs, valor, operavel, taxa);

        });







        // ================================================================

        // TAXAS INDIVIDUAIS POR EMPRESA

        // ================================================================

        function toggleTaxasIndividuais(enabled) {

            var wrap = document.getElementById('taxas-por-empresa-wrap');

            if (!wrap) return;

            wrap.style.display = enabled ? 'block' : 'none';

            if (enabled) buildTaxasEmpresas();

        }



        function buildTaxasEmpresas() {

            var list = document.getElementById('taxas-empresas-list');

            if (!list) return;

            var taxaGlobal = parseFloat(document.getElementById('taxa').value) || 0;

            if (_allEmpresasData.length === 0) {

                list.innerHTML = '<p style="font-size:12px;color:#E65100;background:#FFF3E0;padding:8px;border-radius:4px;margin:0;">\u26a0\ufe0f Defina as taxas aqui <strong>antes</strong> de clicar em Gerar Cota\u00e7\u00f5es. Ap\u00f3s a primeira gera\u00e7\u00e3o, as empresas aparecer\u00e3o aqui.</p>';

                return;

            }

            var html = '<table style="width:100%;border-collapse:collapse;">';

            _allEmpresasData.forEach(function(e) {

                html += '<tr style="border-bottom:1px solid #F0F0F0;">'

                    + '<td style="padding:6px 8px;font-size:13px;width:55%;">' + e.nome + '</td>'

                    + '<td style="padding:6px 8px;">'

                    + '<input type="number" class="taxa-empresa-input" data-empresa="' + e.nome + '" '

                    + 'step="0.01" min="0" max="10" value="' + taxaGlobal + '" '

                    + 'style="width:90px;padding:4px 8px;border:1px solid #E0E0E0;border-radius:4px;font-size:13px;">'

                    + ' %</td></tr>';

            });

            html += '</table>';

            list.innerHTML = html;

        }



        function syncTaxaEmpresas(val) {

            var cb = document.getElementById('taxa-individual');

            if (!cb || !cb.checked) return;

            document.querySelectorAll('.taxa-empresa-input').forEach(function(inp) {

                inp.value = val;

            });

        }



        function resetTaxasEmpresas() {

            var taxaGlobal = parseFloat(document.getElementById('taxa').value) || 0;

            document.querySelectorAll('.taxa-empresa-input').forEach(function(inp) {

                inp.value = taxaGlobal;

            });

        }



        function getTaxaMap() {

            var cb = document.getElementById('taxa-individual');

            if (!cb || !cb.checked) return null;

            var map = {};

            document.querySelectorAll('.taxa-empresa-input').forEach(function(inp) {

                var nome = inp.getAttribute('data-empresa');

                var val  = parseFloat(inp.value);

                if (nome && !isNaN(val)) map[nome] = val;

            });

            return Object.keys(map).length > 0 ? map : null;

        }



        // ================================================================

        // ENVIAR TODOS POR E-MAIL

        // ================================================================

        var _allEmpresasData = [];



        function openSendAllModal() {

            if (_allEmpresasData.length === 0) { alert('Nenhuma cota\u00e7\u00e3o gerada ainda.'); return; }

            var existing = document.getElementById('sendAllModal');

            if (existing) existing.remove();



            // Carregar emails salvos de forma sincrona (usa cache se disponivel)
            window._sendAllSavedEmails = window._sendAllSavedEmails || {};
            // Atualizar em background para proxima vez
            fetch('/email_destinatarios').then(function(r){return r.json();}).then(function(d){window._sendAllSavedEmails=d;}).catch(function(){});

            var overlay = document.createElement('div');

            overlay.id = 'sendAllModal';

            overlay.style.cssText = 'position:fixed;top:0;left:0;width:100%;height:100%;background:rgba(0,0,0,0.5);z-index:9999;display:flex;align-items:center;justify-content:center;';



            var card = document.createElement('div');

            card.style.cssText = 'background:white;border-radius:12px;padding:28px;max-width:640px;width:96%;max-height:85vh;overflow-y:auto;box-shadow:0 8px 32px rgba(0,0,0,0.2);';



            // Header via createElement (evita aspas simples em onclick inline)

            var hdrEl = document.createElement('div');

            hdrEl.style.cssText = 'display:flex;justify-content:space-between;align-items:center;margin-bottom:16px;';

            var hdrTitle = document.createElement('h3');

            hdrTitle.style.cssText = 'margin:0;color:#1B5E20;font-size:16px;';

            hdrTitle.textContent = '\u2709 Enviar cota\u00e7\u00f5es por e-mail';

            var hdrClose = document.createElement('button');

            hdrClose.innerHTML = '&times;';

            hdrClose.style.cssText = 'background:none;border:none;font-size:22px;cursor:pointer;color:#757575;';

            hdrClose.onclick = function(){ overlay.remove(); };

            hdrEl.appendChild(hdrTitle); hdrEl.appendChild(hdrClose);



            var _se = window._sendAllSavedEmails || {};
            var rows = '';

            _allEmpresasData.forEach(function(e, i) {

                var _sv = _se[(e.nome||'').toUpperCase()] || [];
                if (!Array.isArray(_sv)) _sv = _sv ? [_sv] : [];
                var _em = _sv.length > 0 ? _sv : [''];
                var _inRows = _em.map(function(v) {
                    return '<div style="display:flex;gap:3px;align-items:center;">'
                        + '<input type="email" class="sendall-email-input" data-idx="' + i + '"'
                        + ' value="' + (v||'').replace(/"/g,'&quot;') + '"'
                        + ' placeholder="destinatario@empresa.com.br"'
                        + ' style="flex:1;padding:5px 8px;border:1px solid #E0E0E0;border-radius:4px;font-size:12px;">'
                        + '<button type="button"'
                        + ' onclick="var r=this.parentNode,p=r.parentNode;if(p.children.length>1)p.removeChild(r);"'
                        + ' style="border:1px solid #E0E0E0;border-radius:4px;padding:3px 7px;cursor:pointer;background:none;color:#9E9E9E;flex-shrink:0;">&times;</button>'
                        + '</div>';
                }).join('');

                rows += '<tr style="border-bottom:1px solid #F0F0F0;">'

                    + '<td style="padding:8px;font-size:13px;font-weight:600;width:35%;vertical-align:top;">' + e.nome + '</td>'

                    + '<td style="padding:8px;">'
                    + '<div id="sendall-list-' + i + '">' + _inRows + '</div>'
                    + '<button type="button" class="sendall-add-btn" data-idx="' + i + '"'
                    + ' style="border:1px dashed #90CAF9;color:#1565C0;background:none;border-radius:4px;padding:3px 10px;font-size:11px;cursor:pointer;margin-top:3px;width:100%;">+ Adicionar destinat\u00e1rio</button>'
                    + '</td>'

                    + '<td style="padding:8px;text-align:center;white-space:nowrap;vertical-align:top;">'

                    + '<span id="sendall-status-' + i + '" style="font-size:11px;"></span>'

                    + '</td></tr>';

            });



            // Tabela de empresas

            var tblDiv = document.createElement('div');

            tblDiv.innerHTML = '<table style="width:100%;border-collapse:collapse;margin-bottom:16px;">'

                + '<thead><tr style="background:#E8F5E9;">'

                + '<th style="padding:8px;text-align:left;font-size:12px;">Empresa</th>'

                + '<th style="padding:8px;text-align:left;font-size:12px;">E-mail do destinat\u00e1rio</th>'

                + '<th style="padding:8px;font-size:12px;">Status</th>'

                + '</tr></thead><tbody>' + rows + '</tbody></table>';

            // Botoes

            var btnsDiv = document.createElement('div');

            btnsDiv.style.cssText = 'display:flex;gap:8px;justify-content:flex-end;';

            var btnCancelSA = document.createElement('button');

            btnCancelSA.textContent = 'Cancelar';

            btnCancelSA.style.cssText = 'padding:8px 18px;background:white;border:1px solid #E0E0E0;border-radius:6px;cursor:pointer;';

            btnCancelSA.onclick = function(){ overlay.remove(); };

            var btnConfirmSA = document.createElement('button');

            btnConfirmSA.id = 'btn-confirm-send-all';

            btnConfirmSA.innerHTML = '&#9993; Enviar todos';

            btnConfirmSA.style.cssText = 'padding:8px 18px;background:#4CAF50;color:white;border:none;border-radius:6px;cursor:pointer;font-weight:600;';

            btnConfirmSA.onclick = doSendAll;

            btnsDiv.appendChild(btnCancelSA); btnsDiv.appendChild(btnConfirmSA);

            card.appendChild(hdrEl);

            // Seletor de perfil para envio em massa
            var saPerfilWrap = document.createElement('div');
            saPerfilWrap.style.cssText = 'margin-bottom:12px;padding-bottom:12px;border-bottom:1px solid #F0F0F0;';
            var saPerfilLbl = document.createElement('label');
            saPerfilLbl.style.cssText = 'font-size:13px;font-weight:600;display:block;margin-bottom:5px;';
            saPerfilLbl.textContent = 'Perfil do cliente (para todos)';
            var saPerfilSel = document.createElement('select');
            saPerfilSel.id = 'sendall-perfil-select';
            saPerfilSel.style.cssText = 'width:100%;padding:7px 10px;border:1px solid #E0E0E0;border-radius:6px;font-size:13px;background:white;';
            saPerfilSel.innerHTML = '<option value="recorrente">↻ Recorrente — Agenda disponível</option>'
                + '<option value="novo">★ Novo cliente — Apresentação PicPay AR</option>'
                + '<option value="taxa_zero">≈ Taxa zerada — Simulação sem compromisso</option>';
            saPerfilWrap.appendChild(saPerfilLbl); saPerfilWrap.appendChild(saPerfilSel);
            card.appendChild(saPerfilWrap);

            card.appendChild(tblDiv);

            card.appendChild(btnsDiv);



            overlay.appendChild(card);

            document.body.appendChild(overlay);

            // Event delegation para botoes + Adicionar
            overlay.addEventListener('click', function(ev) {
                var btn = ev.target.closest('.sendall-add-btn');
                if (!btn) return;
                var i = btn.getAttribute('data-idx');
                var listDiv = document.getElementById('sendall-list-' + i);
                if (!listDiv) return;
                var row = document.createElement('div');
                row.style.cssText = 'display:flex;gap:3px;align-items:center;';
                var ei = document.createElement('input');
                ei.type = 'email'; ei.className = 'sendall-email-input';
                ei.setAttribute('data-idx', i);
                ei.placeholder = 'destinatario@empresa.com.br';
                ei.style.cssText = 'flex:1;padding:5px 8px;border:1px solid #E0E0E0;border-radius:4px;font-size:12px;';
                var rm = document.createElement('button'); rm.type='button'; rm.innerHTML='&times;';
                rm.style.cssText = 'border:1px solid #E0E0E0;border-radius:4px;padding:3px 7px;cursor:pointer;background:none;color:#9E9E9E;flex-shrink:0;';
                rm.onclick = function() { if (listDiv.querySelectorAll('input').length > 1) listDiv.removeChild(row); };
                row.appendChild(ei); row.appendChild(rm); listDiv.appendChild(row);
                ei.focus();
            });



            // Botao fechar via closure (sem aspas simples no onclick inline)





            fetch('/email_destinatarios').then(function(r){ return r.json(); }).then(function(dest){

                _allEmpresasData.forEach(function(e, i) {

                    var saved = dest[(e.nome||'').toUpperCase()] || '';

                    var inp = document.getElementById('sendall-email-' + i);

                    if (inp && saved) inp.value = saved;

                });

            }).catch(function(){});

        }



        function doSendAll() {

            var btn = document.getElementById('btn-confirm-send-all');

            if (btn) { btn.disabled = true; btn.textContent = 'Enviando...'; }

            var promises = _allEmpresasData.map(function(e, i) {

                // Ler multiplos emails (novo) ou campo unico (legado)
        var multiIn = document.querySelectorAll('.sendall-email-input[data-idx="' + i + '"]');
        var inp = document.getElementById('sendall-email-' + i);

                var toEmails = multiIn.length > 0
            ? Array.from(multiIn).map(function(x){return x.value.trim();}).filter(Boolean)
            : (inp && inp.value.trim() ? [inp.value.trim()] : []);
        var toEmail = toEmails.join(',');

                var statusEl = document.getElementById('sendall-status-' + i);

                if (toEmails.length === 0) { if (statusEl) statusEl.innerHTML = '<span style="color:#9E9E9E;">Pulado</span>'; return Promise.resolve(); }

                if (statusEl) statusEl.innerHTML = '<span style="color:#1565C0;">Enviando...</span>';

                return fetch('/send_email', {

                    method: 'POST', headers: {'Content-Type': 'application/json'},

                    body: JSON.stringify({session_id: e.sid, empresa: e.nome, safe_name: e.safe,

                        to_email: toEmail, to_emails: toEmails, urs: e.urs, valor_total: e.valor, valor_operavel: e.operavel,
                        taxa_pct: e.taxa || 0, perfil_cliente: (document.getElementById('sendall-perfil-select')||{value:'recorrente'}).value})

                })

                .then(function(r){ return r.json(); })

                .then(function(d){

                    if (statusEl) statusEl.innerHTML = d.error

                        ? '<span style="color:#C62828;">&#10060; Erro</span>'

                        : '<span style="color:#2E7D32;">&#10003; Enviado</span>';

                })

                .catch(function(){ if (statusEl) statusEl.innerHTML = '<span style="color:#C62828;">Erro</span>'; });

            });

            Promise.all(promises).then(function(){ if (btn) { btn.disabled = false; btn.textContent = '\u2709 Enviar todos'; } });

        }





        // ================================================================

        // EMAIL COTACAO PERSONALIZADA

        // ================================================================

        function openEmailModalCustom(sid, safe, empresaEnc, urs, valorTotal, valorOp, taxaPct) {

            var empresa = decodeURIComponent(empresaEnc);

            fetch('/email_destinatarios')

            .then(function(r){ return r.json(); })

            .then(function(dest){

                var emailSalvo = dest[(empresa||'').toUpperCase()] || '';

                function fmtBRL(v) {

                    return 'R$ ' + parseFloat(v).toLocaleString('pt-BR', {minimumFractionDigits:2, maximumFractionDigits:2});

                }

                // pct: taxa da operacao (ex: 1.5%), nao ratio operavel/total

                var pct = (taxaPct !== undefined && taxaPct !== null) ? taxaPct : (valorTotal > 0 ? Math.round(valorOp/valorTotal*100) : 0);

                var existing = document.getElementById('emailModalCustom');

                if (existing) existing.remove();



                var overlay = document.createElement('div');

                overlay.id = 'emailModalCustom';

                overlay.style.cssText = 'position:fixed;top:0;left:0;width:100%;height:100%;background:rgba(0,0,0,0.5);z-index:9999;display:flex;align-items:center;justify-content:center;';



                var card = document.createElement('div');

                card.style.cssText = 'background:white;border-radius:12px;padding:28px;max-width:540px;width:92%;box-shadow:0 8px 32px rgba(0,0,0,0.2);';



                var hdr = document.createElement('div');

                hdr.style.cssText = 'display:flex;justify-content:space-between;align-items:center;margin-bottom:16px;';

                var h3 = document.createElement('h3');

                h3.style.cssText = 'margin:0;color:#1B5E20;font-size:16px;';

                h3.textContent = '\u2709 Enviar cota\u00e7\u00e3o personalizada';

                var btnX = document.createElement('button');

                btnX.innerHTML = '&times;';

                btnX.style.cssText = 'background:none;border:none;font-size:22px;cursor:pointer;color:#757575;';

                btnX.onclick = function(){ overlay.remove(); };

                hdr.appendChild(h3); hdr.appendChild(btnX);



                var info = document.createElement('div');

                info.style.cssText = 'background:#F9FBE7;border-radius:8px;padding:12px 16px;margin-bottom:16px;';

                info.innerHTML = '<strong style="font-size:14px;">' + empresa + '</strong><br>'

                    + '<span style="font-size:12px;color:#757575;">'

                    + urs.toLocaleString('pt-BR') + ' URs \u00a0|\u00a0 Agenda: ' + fmtBRL(valorTotal)

                    + (pct > 0 ? ' \u00a0|\u00a0 <span style="color:#1B5E20;font-weight:bold;">Taxa de opera\u00e7\u00e3o: ' + pct + '%</span>' : '')

                    + '</span>';



                // Perfil de cliente
                var perfilWrapC = document.createElement('div');
                perfilWrapC.style.marginBottom = '12px';
                perfilWrapC.innerHTML = '<label style="font-size:13px;font-weight:600;display:block;margin-bottom:4px;">Perfil do cliente</label><select id="custom-email-perfil-select" style="width:100%;padding:7px 10px;border:1px solid #E0E0E0;border-radius:6px;font-size:13px;"><option value="recorrente">Recorrente (j\u00e1 opera AR no PicPay)</option><option value="novo">Novo cliente (n\u00e3o opera ainda)</option><option value="taxa_zero">Taxa zero (agenda indicativa)</option></select>';

                var msgWrap = document.createElement('div');

                msgWrap.style.marginBottom = '12px';

                var msgLbl = document.createElement('label');

                msgLbl.style.cssText = 'font-size:13px;font-weight:600;display:block;margin-bottom:4px;';

                msgLbl.textContent = 'Mensagem personalizada';

                var msgTa = document.createElement('textarea');

                msgTa.id = 'customEmailMsg'; msgTa.rows = 3;

                msgTa.style.cssText = 'width:100%;box-sizing:border-box;padding:8px 12px;border:1px solid #E0E0E0;border-radius:6px;font-size:13px;resize:vertical;';

                msgTa.value = window._lastCustomMsg || 'Conforme a sua solicita\u00e7\u00e3o, segue em anexo a cota\u00e7\u00e3o atualizada da sua agenda de antecipa\u00e7\u00e3o de receb\u00edveis.';
                msgTa.addEventListener('input', function(){ window._lastCustomMsg = msgTa.value; });

                msgWrap.appendChild(msgLbl); msgWrap.appendChild(msgTa);



                var fldWrap = document.createElement('div');

                fldWrap.style.marginBottom = '12px';

                var lbl = document.createElement('label');

                lbl.style.cssText = 'font-size:13px;font-weight:600;display:block;margin-bottom:4px;';

                lbl.textContent = 'E-mail do destinat\u00e1rio';

                var inp = document.createElement('input');

                inp.type = 'email'; inp.id = 'emailToCustom'; inp.value = emailSalvo;

                inp.placeholder = 'contato@empresa.com.br';

                inp.style.cssText = 'width:100%;box-sizing:border-box;padding:8px 12px;border:1px solid #E0E0E0;border-radius:6px;font-size:14px;';

                fldWrap.appendChild(lbl); fldWrap.appendChild(inp);



                var statusDiv = document.createElement('div');

                statusDiv.id = 'emailStatusCustom';

                statusDiv.style.cssText = 'min-height:20px;margin-bottom:8px;font-size:13px;';



                var btns = document.createElement('div');

                btns.style.cssText = 'display:flex;gap:8px;justify-content:flex-end;';

                var btnCancel = document.createElement('button');

                btnCancel.textContent = 'Cancelar';

                btnCancel.style.cssText = 'padding:8px 18px;background:white;border:1px solid #E0E0E0;border-radius:6px;cursor:pointer;';

                btnCancel.onclick = function(){ overlay.remove(); };

                var btnSend = document.createElement('button');

                btnSend.innerHTML = '&#9993; Enviar';

                btnSend.style.cssText = 'padding:8px 18px;background:#4CAF50;color:white;border:none;border-radius:6px;cursor:pointer;font-weight:600;';

                btnSend.onclick = function(){

                    var toEmail = (document.getElementById('emailToCustom')||{}).value||'';

                    var msg = (document.getElementById('customEmailMsg')||{}).value||'';

                    var st = document.getElementById('emailStatusCustom');

                    if (!toEmail) { if(st) st.textContent = 'Informe o e-mail do destinat\u00e1rio.'; return; }

                    if (st) st.textContent = 'Enviando...';

                    fetch('/send_email', {

                        method: 'POST', headers: {'Content-Type': 'application/json'},

                        body: JSON.stringify({session_id: sid, empresa: empresa, safe_name: safe,

                            to_email: toEmail, urs: urs, valor_total: valorTotal, valor_operavel: valorOp,

                            taxa_pct: taxaPct || 0,

                            custom_message: msg, operator_email: getOperatorEmail(), perfil_cliente: (document.getElementById('custom-email-perfil-select')||{value:'recorrente'}).value, is_custom: true})

                    })

                    .then(function(r){ return r.json(); })

                    .then(function(d){

                        if (d.error) { if(st) st.textContent = 'Erro: ' + d.error; }

                        else { if(st) st.textContent = d.message||'Enviado com sucesso!'; setTimeout(function(){ overlay.remove(); }, 2000); }

                    })

                    .catch(function(e){ if(st) st.textContent = 'Erro de conex\u00e3o.'; });

                };

                btns.appendChild(btnCancel); btns.appendChild(btnSend);

                card.appendChild(hdr); card.appendChild(info); card.appendChild(perfilWrapC); card.appendChild(msgWrap);

                card.appendChild(fldWrap); card.appendChild(statusDiv); card.appendChild(btns);

                overlay.appendChild(card);

                document.body.appendChild(overlay);

            });

        }





        // Popular _allEmpresasData apos geracao

        function _updateEmpresasData(empresas, sid) {

            _allEmpresasData = (empresas||[]).map(function(e) {

                return {sid: sid, safe: e.safe_name, nome: e.nome,

                        urs: e.urs||0, valor: e.valor||0, operavel: e.valor_operavel||0,

                        taxa: e.taxa||0};

            });

            var btnSA = document.getElementById('btn-send-all');

            if (btnSA) btnSA.style.display = _allEmpresasData.length > 0 ? 'inline-block' : 'none';

        }




        function renderMissingPanel(mbEmp) {
            if (!mbEmp || Object.keys(mbEmp).length === 0) return '';
            var tUrs=0,tVal=0;
            Object.values(mbEmp).forEach(function(v){tUrs+=v.urs||0;tVal+=v.valor||0;});
            var h='<div style="background:#FFF8E1;border:1px solid #FFB300;border-radius:8px;padding:12px 16px;margin-top:10px;">';
            h+='<div style="display:flex;align-items:center;gap:8px;margin-bottom:8px;">';
            h+='<span style="font-size:13px;color:#E65100;font-weight:600;">🔒 Seller IDs n\u00e3o encontrados</span>';
            h+='<span style="background:#FF6F00;color:white;border-radius:12px;padding:2px 8px;font-size:11px;">'+tUrs.toLocaleString('pt-BR')+' URs &bull; '+formatBRL(tVal)+' n\u00e3o oper\u00e1veis</span></div>';
            h+='<table style="width:100%;border-collapse:collapse;font-size:12px;"><thead><tr style="background:#FFF3E0;">';
            h+='<th style="padding:5px 8px;text-align:left;">Empresa</th><th style="padding:5px 8px;text-align:right;">CNPJs s/ID</th><th style="padding:5px 8px;text-align:right;">URs</th><th style="padding:5px 8px;text-align:right;">Valor</th></tr></thead><tbody>';
            Object.entries(mbEmp).forEach(function(kv,i){
                var emp=kv[0],v=kv[1],bg=i%2?'background:#FFFDE7;':'';
                h+='<tr style="'+bg+'border-bottom:1px solid #FFE082;">';
                h+='<td style="padding:5px 8px;color:#E65100;font-weight:600;">'+emp+'</td>';
                h+='<td style="padding:5px 8px;text-align:right;">'+v.count+'</td>';
                h+='<td style="padding:5px 8px;text-align:right;">'+(v.urs||0).toLocaleString('pt-BR')+'</td>';
                h+='<td style="padding:5px 8px;text-align:right;">'+formatBRL(v.valor||0)+'</td></tr>';
            });
            h+='</tbody></table><p style="font-size:11px;color:#E65100;margin:6px 0 0;">Estas URs n\u00e3o entrar\u00e3o na Sele\u00e7\u00e3o de URs.</p></div>';
            return h;
        }


        // Listeners para botoes de selecao de CNPJ
        document.addEventListener('click', function(ev) {
            if (ev.target.id === 'cnpj-sel-all') {
                document.querySelectorAll('.cnpj-checkbox').forEach(function(cb){ cb.checked = true; });
            } else if (ev.target.id === 'cnpj-sel-none') {
                document.querySelectorAll('.cnpj-checkbox').forEach(function(cb){ cb.checked = false; });
            }
        });


        function selectAllCnpjs(selectAll) {
            var sel = document.getElementById('custom-cnpj-select');
            if (!sel) return;
            for (var i = 0; i < sel.options.length; i++) {
                sel.options[i].selected = selectAll;
            }
            updateCnpjCount();
        }

        function updateCnpjCount() {
            var sel = document.getElementById('custom-cnpj-select');
            var countEl = document.getElementById('cnpj-count');
            if (!sel || !countEl) return;
            var selected = Array.from(sel.selectedOptions).length;
            var total = sel.options.length;
            countEl.textContent = selected + '/' + total + ' selecionados';
        }


        function filterCnpjs(query) {
            var sel = document.getElementById('custom-cnpj-select');
            if (!sel) return;
            var q = (query || '').toLowerCase();
            for (var i = 0; i < sel.options.length; i++) {
                var opt = sel.options[i];
                var match = !q || opt.value.toLowerCase().indexOf(q) >= 0 || opt.text.toLowerCase().indexOf(q) >= 0;
                opt.style.display = match ? '' : 'none';
            }
        }


        function updateFiltersForCnpj() {
            var sel = document.getElementById('custom-cnpj-select');
            var empresa = document.getElementById('custom-empresa').value;
            if (!sel || !empresa || !sessionId) return;

            var selectedCnpjs = Array.from(sel.selectedOptions).map(function(o){ return o.value; });
            if (selectedCnpjs.length === 0 || selectedCnpjs.length === sel.options.length) {
                // Todos ou nenhum selecionado — nao filtrar (usa loadEmpresaFilters normal)
                return;
            }

            var cnpjParam = encodeURIComponent(selectedCnpjs.join(','));
            fetch('/get_datas?session_id=' + sessionId + '&empresa=' + encodeURIComponent(empresa) + '&cnpj=' + cnpjParam)
            .then(function(r){ return r.json(); })
            .then(function(data){
                // Atualizar adquirentes
                var adqWrap = document.getElementById('custom-adquirente-wrap');
                if (adqWrap) {
                    adqWrap.innerHTML = buildCheckboxes('adq', data.adquirentes, null);
                }
                // Atualizar arranjos
                var arrWrap = document.getElementById('custom-arranjo-wrap');
                if (arrWrap) {
                    var labelMap = {'ECC':'Elo','VCC':'Visa','MCC':'Master','ACC':'Amex'};
                    arrWrap.innerHTML = buildCheckboxes('arr', data.arranjos, labelMap);
                }
                // Atualizar datas
                var container = document.getElementById('datas-checkboxes');
                if (container && data.datas) {
                    var html = '';
                    data.datas.forEach(function(d) {
                        html += '<label style="display:inline-block;margin:3px 8px;font-size:13px;cursor:pointer;">';
                        html += '<input type="checkbox" value="' + d.data + '" checked onchange="updateDatasSummary()"> ';
                        html += d.data + ' <span style="color:#9E9E9E;">(' + d.urs + ' URs | ' + formatBRL(d.valor) + ')</span></label>';
                    });
                    container.innerHTML = html;
                    updateDatasSummary();
                }
                // Atualizar badge operavel para os CNPJs selecionados
                var _oe = document.getElementById('empresa-info-operavel');
                if (_oe && data.total_valor_elegivel > 0) {
                    _oe.textContent = 'Oper\u00e1vel (CNPJs sel.): ' + formatBRL(data.total_valor_elegivel);
                    _oe.style.display = '';
                }
                var _iv = document.getElementById('empresa-info-urs');
                if (_iv) _iv.textContent = (data.total_urs||0).toLocaleString('pt-BR') + ' URs';
                var _vv = document.getElementById('empresa-info-valor');
                if (_vv) _vv.textContent = 'Agenda: ' + formatBRL(data.total_valor||0);
            })
            .catch(function(){ /* silencioso */ });
        }

        checkServerStatus();



        checkDbStatus(false);



        loadHistory();



        // Drag & drop



        const ua = document.getElementById('upload-area');



        ua.addEventListener('dragover', e => { e.preventDefault(); ua.classList.add('dragover'); });



        ua.addEventListener('dragleave', () => ua.classList.remove('dragover'));



        ua.addEventListener('drop', e => { e.preventDefault(); ua.classList.remove('dragover'); handleFiles(e.dataTransfer.files); });



        document.getElementById('file-input').addEventListener('change', e => handleFiles(e.target.files));



        let sessionId = null;



        function getOperatorEmail() {



            const sel = document.getElementById('operador');



            if (sel.value === 'outro') return document.getElementById('operador-email').value;



            return sel.value;



        }



        function formatBRL(v) {



            return 'R$ ' + v.toLocaleString('pt-BR', {minimumFractionDigits: 2, maximumFractionDigits: 2});



        }



        function handleFiles(files) {



            if (!files || files.length === 0) return;



            const fd = new FormData();



            for (let i = 0; i < files.length; i++) {



                fd.append('files', files[i]);



            }



            const fileNames = Array.from(files).map(f => f.name).join(', ');



            document.getElementById('upload-status').innerHTML = '<div class="status status-info"><span class="loader"></span> Analisando ' + files.length + ' arquivo(s): ' + fileNames + '</div>';



            document.getElementById('upload-status').classList.remove('hidden');



            fetch('/upload', {method: 'POST', body: fd})



            .then(r => r.json())



            .then(data => {



                if (data.error) {



                    document.getElementById('upload-status').innerHTML = '<div class="status status-error">'+data.error+'</div>';



                    return;



                }



                sessionId = data.session_id;



                // Iniciar prefetch de sellers imediatamente em background



                startPrefetch(sessionId);



                const filesInfo = data.source_files ? ' (' + data.source_files.length + ' arquivo(s): ' + data.source_files.join(', ') + ')' : '';



                document.getElementById('upload-status').innerHTML = '<div class="status status-success">✅ '+data.total_urs.toLocaleString()+' URs encontradas em '+data.total_empresas+' empresas' + filesInfo + '</div>';



                // Build table



                let html = '<table><tr><th>Empresa</th><th class="text-right">CNPJs</th><th class="text-right">URs</th><th class="text-right">Valor Disponivel</th><th>Adquirentes</th></tr>';



                let tv = 0, tu = 0;



                data.empresas.forEach(e => {



                    tv += e.valor;



                    tu += e.urs;



                    html += '<tr><td><strong>'+e.nome+'</strong></td><td class="text-right">'+e.cnpjs+'</td><td class="text-right">'+e.urs.toLocaleString()+'</td><td class="text-right">'+formatBRL(e.valor)+'</td><td>'+e.adquirentes.map(a=>'<span class="chip">'+a+'</span>').join('')+'</td></tr>';



                });



                html += '<tr class="total-row"><td>TOTAL</td><td></td><td class="text-right">'+tu.toLocaleString()+'</td><td class="text-right">'+formatBRL(tv)+'</td><td></td></tr></table>';



                document.getElementById('resumo-content').innerHTML = html;



                document.getElementById('step-resumo').classList.remove('hidden');



                // Checar raizes desconhecidas



                if (data.unknown_raizes && Object.keys(data.unknown_raizes).length > 0) {



                    let formHtml = '<div class="status status-info mb-2">Encontrei raizes de CNPJ que nao reconheco. Informe o nome de cada empresa:</div>';



                    formHtml += '<div style="margin-top:12px;">';



                    for (const [raiz, info] of Object.entries(data.unknown_raizes)) {



                        formHtml += '<div class="form-row" style="margin-bottom:8px;align-items:center;">';



                        formHtml += '<div><strong>' + raiz + '</strong> <span style="color:#9E9E9E;">(' + info.cnpjs + ' CNPJs, ' + info.urs + ' URs, ex: ' + info.example + ')</span></div>';



                        formHtml += '<div><input type="text" id="raiz-' + raiz + '" placeholder="Nome da empresa" style="width:250px;"></div>';



                        formHtml += '</div>';



                    }



                    formHtml += '<button class="btn btn-primary mt-4" onclick="saveRaizes()">Salvar e continuar</button>';



                    formHtml += '</div>';



                    



                    document.getElementById('step-unknown').innerHTML = '<h2>Empresas nao identificadas</h2>' + formHtml;



                    document.getElementById('step-unknown').classList.remove('hidden');



                    // Nao mostrar params ainda



                } else {



                    document.getElementById('step-params').classList.remove('hidden');



                }



                // Populate custom empresa select



                const sel = document.getElementById('custom-empresa');



                sel.innerHTML = '';



                // Guardar mapa de valores para enriquecer depois com elegivel



                window._empresaValores = {};



                data.empresas.forEach(e => {



                    window._empresaValores[e.nome] = {total: e.valor, elegivel: null};



                    const opt = document.createElement('option');



                    opt.value = e.nome;



                    opt.text = e.nome + ' (R$ ' + (e.valor/1e6).toFixed(1) + 'M)';



                    sel.appendChild(opt);



                });



            });



        }



        function generateAll() {



            const taxa = parseFloat(document.getElementById('taxa').value);



            const di = parseFloat(document.getElementById('di-periodo').value) / 100;



            const email = getOperatorEmail();



            if (!taxa && taxa !== 0) { alert('Informe a taxa mensal'); return; }



            if (!email) { alert('Informe o operador'); return; }



            const btn = event.target;



            btn.disabled = true;



            btn.innerHTML = '<span class="loader"></span> Gerando...';

            // Se 'Cotacao Completa' marcada, gen_cotacao deve ser true
            var _genCompleto = !!(document.getElementById('gen-completo') && document.getElementById('gen-completo').checked);
            var _genCotacao  = document.getElementById('gen-cotacao').checked || _genCompleto;

            fetch('/generate', {



                method: 'POST',



                headers: {'Content-Type': 'application/json'},



                body: JSON.stringify({session_id: sessionId, taxa: taxa, di_periodo: di, operator_email: email,



                        gen_cotacao: _genCotacao,

                        gen_selecao: document.getElementById('gen-selecao').checked,

                        only_eligible: !_genCompleto,

                taxa_map: getTaxaMap()})



            })



            .then(r => r.json())



            .then(data => {



                btn.disabled = false;



                btn.innerHTML = '🏭 Gerar Cotacoes (Todas as Empresas)';



                if (data.error) { alert(data.error); return; }



                let sellerInfo = '';



                let sellerAlert = '';



                if (data.seller_error) {



                    sellerAlert = '<div class="status status-error mt-2">' + data.seller_error + '</div>';



                } else if (data.seller_ids_found !== undefined) {



                    sellerInfo = ' | Seller IDs: ' + data.seller_ids_found + ' encontrados';



                    if (data.seller_ids_missing > 0) {



                        sellerInfo += ', ' + data.seller_ids_missing + ' faltantes';



                    }



                }



                let html = '<div class="status status-success">✅ '+data.empresas.length+' cotações geradas!' + sellerInfo + '</div>';



                html += '<table class="mt-2"><tr><th>Empresa</th><th class="text-right">URs</th><th class="text-right">Valor Bruto</th><th class="text-right">Operável</th><th class="text-right">Taxa</th><th>Ações</th></tr>';



                data.empresas.forEach(e => {
                    var taxaLabel1 = e.taxa ? (e.taxa.toFixed(2) + '%') : '-';
                    var opLabel1   = e.valor_operavel > 0 ? formatBRL(e.valor_operavel) : '<span style="color:#9E9E9E;">-</span>';
                    html += '<tr>';
                    html += '<td><strong>'+e.nome+'</strong></td>';
                    html += '<td class="text-right">'+e.urs.toLocaleString()+'</td>';
                    html += '<td class="text-right">'+formatBRL(e.valor)+'</td>';
                    html += '<td class="text-right" style="color:#1B5E20;">'+opLabel1+'</td>';
                    html += '<td class="text-right" style="color:#1B5E20;font-weight:600;">'+taxaLabel1+'</td>';
                    html += '<td>';
                    html += '<a href="/download/'+sessionId+'/'+encodeURIComponent(e.safe_name)+'" class="btn btn-secondary" style="padding:6px 12px;font-size:12px;">⬇ Download</a>';





                

                                    html += '<button class="btn btn-email-action" data-sid="'+sessionId+'" data-safe="'+encodeURIComponent(e.safe_name)+'" data-nome="'+encodeURIComponent(e.nome)+'" data-urs="'+e.urs+'" data-valor="'+(e.valor||0)+'" data-operavel="'+(e.valor_operavel||0)+'" data-taxa="'+(e.taxa||0)+'" style="padding:6px 12px;font-size:12px;background:#E3F2FD;border:1px solid #90CAF9;border-radius:6px;cursor:pointer;margin-left:4px;">&#9993;&#65039; E-mail</button>';
                    html += '</td></tr>';
                });



                html += '</table>';



                document.getElementById('results-content').innerHTML = html + sellerAlert;



                document.getElementById('step-results').classList.remove('hidden');
                // Mostrar calculadora AR e disparar cálculo inicial
                var _arCard = document.getElementById('step-calc-ar');
                if (_arCard) { _arCard.classList.remove('hidden'); setTimeout(calcularAR, 300); }



                _updateEmpresasData(data.empresas, sessionId);
                if (data.missing_by_empresa && Object.keys(data.missing_by_empresa).length > 0) {
                    var _rcm1 = document.getElementById('results-content');
                    if (_rcm1) _rcm1.insertAdjacentHTML('beforeend', renderMissingPanel(data.missing_by_empresa));
                }



                var cbTaxa = document.getElementById('taxa-individual');

                if (cbTaxa && cbTaxa.checked) buildTaxasEmpresas();



                // Buscar seller IDs em background



                const sellerStatus = document.createElement('div');



                sellerStatus.className = 'status status-info mt-2';



                sellerStatus.innerHTML = '<span class="loader"></span> Buscando Seller IDs no Databricks...';



                document.getElementById('results-content').appendChild(sellerStatus);



                fetch('/fetch_sellers', {



                    method: 'POST',



                    headers: {'Content-Type': 'application/json'},



                    body: JSON.stringify({session_id: sessionId, taxa: taxa, di_periodo: di, operator_email: email,

                        only_eligible: !(document.getElementById('gen-completo') && document.getElementById('gen-completo').checked), taxa_map: getTaxaMap()})



                })



                .then(r2 => r2.json())



                .then(sellerData => {



                    if (sellerData.requires_eligibility_confirmation) {



                        // Esconder resultados preliminares (gerados sem elegibilidade)



                        const stepResults = document.getElementById('step-results');



                        if (stepResults) stepResults.classList.add('hidden');



                        sellerStatus.className = 'status mt-2';



                        sellerStatus.style.background = '#FFF8E1';



                        sellerStatus.style.color = '#E65100';



                        sellerStatus.style.border = '2px solid #F9A825';



                        sellerStatus.style.padding = '12px';



                        sellerStatus.innerHTML = '&#9888; <strong>' + sellerData.ineligible_sellers.length + ' seller(s) inelegiveis</strong> &mdash; os arquivos ainda n&atilde;o foram gerados. Confirme abaixo para gerar sem eles (ou inclui-los s&oacute; na cota&ccedil;&atilde;o).';



                        const fetchParams = {session_id: sessionId, taxa: taxa, di_periodo: di, operator_email: email, taxa_map: getTaxaMap()};



                        showEligibilityPanel(sellerData.ineligible_sellers, sessionId, fetchParams);



                    } else if (sellerData.seller_error) {



                        sellerStatus.className = 'status status-error mt-2';



                        sellerStatus.innerHTML = '<strong>Erro ao buscar Seller IDs:</strong> ' + sellerData.seller_error



                            + '<br><button class="btn btn-secondary" style="padding:4px 12px;font-size:12px;margin-top:8px;" '



                            + 'onclick="retrySellers()">Tentar novamente</button>';



                    } else {



                        sellerStatus.className = 'status status-success mt-2';



                        sellerStatus.innerHTML = 'Seller IDs: ' + sellerData.seller_ids_found + ' encontrados'



                            + (sellerData.seller_ids_missing > 0 ? ', ' + sellerData.seller_ids_missing + ' faltantes' : '')



                            + ' (arquivos atualizados)';



                        // Painel de inelegiveis por empresa



                        if (sellerData.inelig_by_empresa && Object.keys(sellerData.inelig_by_empresa).length > 0) {



                            const rc = document.getElementById('results-content');



                            if (rc) rc.insertAdjacentHTML('beforeend', renderIneligPanel(sellerData.inelig_by_empresa, sessionId));



                        }



                        // Painel de resumo por seller



                        if (sellerData.seller_summary && sellerData.seller_summary.length > 0) {



                            const rc = document.getElementById('results-content');



                            if (rc) rc.insertAdjacentHTML('beforeend', renderSellerSummary(sellerData.seller_summary));



                        }



                    }



                })



                .catch((err) => {



                    sellerStatus.className = 'status status-error mt-2';



                    sellerStatus.innerHTML = 'Erro ao buscar Seller IDs: falha na comunicacao com o servidor. '



                        + '<button class="btn btn-secondary" style="padding:4px 12px;font-size:12px;margin-left:8px;" '



                        + 'onclick="retrySellers()">Tentar novamente</button>';



                });



                document.getElementById('step-custom').classList.remove('hidden');



                if (data.ineligible_excluded) lastIneligibleExcluded = data.ineligible_excluded;



                loadHistory();



            });



        }



        function toggleDatasCustom() {



            const mode = document.getElementById('custom-datas-mode').value;



            document.getElementById('datas-range').classList.toggle('hidden', mode !== 'range');



            document.getElementById('datas-range').style.display = mode === 'range' ? 'flex' : 'none';



            document.getElementById('datas-select').classList.toggle('hidden', mode !== 'select');



        }



        function toggleAllDatas(checked) {



            document.querySelectorAll('#datas-checkboxes input[type=checkbox]').forEach(cb => cb.checked = checked);



            updateDatasSummary();



        }



        function updateDatasSummary() {



            const checked = document.querySelectorAll('#datas-checkboxes input:checked');



            const total = document.querySelectorAll('#datas-checkboxes input[type=checkbox]').length;



            const el = document.getElementById('datas-summary');



            if (checked.length === 0 || checked.length === total) {



                el.textContent = 'Todas as ' + total + ' datas selecionadas';



            } else {



                el.textContent = checked.length + ' de ' + total + ' datas selecionadas';



            }



        }



        function loadEmpresaFilters() {



            const empresa = document.getElementById('custom-empresa').value;



            if (!empresa || !sessionId) return;



            // Loading state



            document.getElementById('custom-adquirente-wrap').innerHTML = '<span style="color:#9E9E9E;font-size:12px;">Carregando...</span>';



            document.getElementById('custom-arranjo-wrap').innerHTML = '<span style="color:#9E9E9E;font-size:12px;">Carregando...</span>';



            fetch('/get_datas?session_id=' + sessionId + '&empresa=' + encodeURIComponent(empresa))



            .then(r => r.json())



            .then(data => {



                // Adquirentes checkboxes (multi-selecao)



                document.getElementById('custom-adquirente-wrap').innerHTML = buildCheckboxes('adq', data.adquirentes, null);



                // Arranjos checkboxes (multi-selecao)



                document.getElementById('custom-arranjo-wrap').innerHTML = buildCheckboxes('arr', data.arranjos,



                    {'ECC':'Elo','VCC':'Visa','MCC':'Master','ACC':'Amex'});

                // Popula filtro de CNPJs
                var cnpjGroup = document.getElementById('custom-cnpj-group');
                var cnpjSel   = document.getElementById('custom-cnpj-select');
                if (cnpjGroup && cnpjSel) {
                    if (data.cnpjs && data.cnpjs.length > 0) {
                        cnpjGroup.style.display = '';
                        cnpjSel.innerHTML = '';
                        data.cnpjs.forEach(function(c) {
                            var hasOp = c.agenda_operavel > 0;
                            var opt = document.createElement('option');
                            opt.value = c.cnpj;
                            opt.selected = hasOp;  // seleciona apenas os operaveis
                            opt.title = c.cnpj + ' | Total: ' + formatBRL(c.agenda_total) + ' | Operável: ' + formatBRL(c.agenda_operavel);
                            opt.style.color = hasOp ? '#212121' : '#C62828';
                            opt.textContent = c.cnpj
                                + '  │  ' + c.urs + ' URs'
                                + '  │  Total: ' + formatBRL(c.agenda_total)
                                + (hasOp
                                    ? '  │  Op: ' + formatBRL(c.agenda_operavel)
                                    : '  │  ⛔ Inoperável');
                            cnpjSel.appendChild(opt);
                        });
                        cnpjSel.size = Math.min(data.cnpjs.length, 6);
                        updateCnpjCount();
                    } else {
                        cnpjGroup.style.display = 'none';
                    }
            
        // Badges de totais abaixo do select de empresa
        var _infoDiv = document.getElementById('empresa-info');
        var _infoUrs = document.getElementById('empresa-info-urs');
        var _infoVal = document.getElementById('empresa-info-valor');
        var _infoOp  = document.getElementById('empresa-info-operavel');
        if (_infoDiv) _infoDiv.classList.remove('hidden');
        if (_infoUrs) _infoUrs.textContent = (data.total_urs||0).toLocaleString('pt-BR') + ' URs';
        if (_infoVal) _infoVal.textContent = 'Agenda: ' + formatBRL(data.total_valor||0);
        if (_infoOp) {
            var _opv = data.total_valor_elegivel || 0;
            if (_opv > 0 && _opv < (data.total_valor||0)) {
                _infoOp.textContent = 'Operável: ' + formatBRL(_opv);
                _infoOp.style.display = '';
            } else if (_opv > 0) {
                _infoOp.textContent = 'Operável: ' + formatBRL(_opv);
                _infoOp.style.display = '';
            } else {
                _infoOp.style.display = 'none';
            }
        }

    } else if (!data.has_eligibility) {



                    let infoEl = document.getElementById('empresa-elig-info');



                    if (!infoEl) {



                        infoEl = document.createElement('div');



                        infoEl.id = 'empresa-elig-info';



                        infoEl.style.cssText = 'margin-top:6px;font-size:12px;color:#9E9E9E;';



                        document.getElementById('custom-empresa').parentNode.appendChild(infoEl);



                    }



                    infoEl.innerHTML = 'Aguardando verificação de elegibilidade...';



                }



                // Datas (checkboxes)



                const container = document.getElementById('datas-checkboxes');



                let html = '';



                data.datas.forEach(d => {



                    html += '<label style="display:inline-block;margin:3px 8px;font-size:13px;cursor:pointer;">';



                    html += '<input type="checkbox" value="' + d.data + '" checked onchange="updateDatasSummary()"> ';



                    html += d.data + ' <span style="color:#9E9E9E;">(' + d.urs + ' URs | ' + formatBRL(d.valor) + ')</span></label><br>';



                });



                container.innerHTML = html;



                updateDatasSummary();



            });



        }



        function getSelectedDatas() {



            const mode = document.getElementById('custom-datas-mode').value;



            if (mode === 'todas') return {mode: 'todas'};



            if (mode === 'range') {



                return {mode: 'range', de: document.getElementById('custom-data-de').value, ate: document.getElementById('custom-data-ate').value};



            }



            if (mode === 'select') {



                const checked = [];



                document.querySelectorAll('#datas-checkboxes input:checked').forEach(cb => checked.push(cb.value));



                return {mode: 'select', datas: checked};



            }



        }



        // Sellers excluidos da ultima geracao (para mostrar no personalizada)



        let lastIneligibleExcluded = [];



        function _parseValorCustom(str) {



            if (!str || str.trim() === '') return 0;



            str = str.trim().toUpperCase().replace(/[R$\s]/g,'').replace(',','.');



            if (str.endsWith('B')) return parseFloat(str) * 1e9;



            if (str.endsWith('M')) return parseFloat(str) * 1e6;



            if (str.endsWith('K')) return parseFloat(str) * 1e3;



            return parseFloat(str) || 0;



        }



        function _setValorCustom(v) {



            var el = document.getElementById('custom-valor');



            if (!el) return;



            if (v === 0) { el.value = ''; }



            else if (v >= 1e6) { el.value = (v/1e6 % 1 === 0 ? (v/1e6).toFixed(0) : (v/1e6).toFixed(1)) + 'M'; }



            else { el.value = (v/1e3).toFixed(0) + 'K'; }



            _previewValor();



        }



        function _previewValor() {



            var el = document.getElementById('custom-valor');



            var prev = document.getElementById('custom-valor-preview');



            if (!el || !prev) return;



            var v = _parseValorCustom(el.value);



            prev.textContent = v > 0 ? '= ' + formatBRL(v) : '';



        }



        function generateCustom() {



            const empresa = document.getElementById('custom-empresa').value;



            const valor = _parseValorCustom(document.getElementById('custom-valor').value);



            // Ler checkboxes marcados (multi-selecao)



            const adqChecked = Array.from(document.querySelectorAll('#custom-adquirente-wrap input[type=checkbox]:checked')).map(function(c){return c.value;});



            const arrChecked = Array.from(document.querySelectorAll('#custom-arranjo-wrap input[type=checkbox]:checked')).map(function(c){return c.value;});



            const adquirente = adqChecked.join(',');



            const arranjo = arrChecked.join(',');



            const customTaxa = document.getElementById('custom-taxa').value;



            const taxa = customTaxa ? parseFloat(customTaxa) : parseFloat(document.getElementById('taxa').value);



            const di = parseFloat(document.getElementById('di-periodo').value) / 100;



            const email = getOperatorEmail();



            const datasFilter = getSelectedDatas();



            fetch('/generate_custom', {



                method: 'POST',



                headers: {'Content-Type': 'application/json'},



                body: JSON.stringify({session_id: sessionId, empresa: empresa, valor_alvo: valor,



                    adquirente: adquirente, arranjo: arranjo, taxa: taxa, di_periodo: di,



                    operator_email: email, datas_filter: datasFilter,
                    cnpj: (function(){ var s=document.getElementById('custom-cnpj-select'); return s ? Array.from(s.selectedOptions).map(function(o){return o.value;}).join(',') : ''; })()
                    })



            })



            .then(r => r.json())



            .then(data => {



                if (data.error) { alert(data.error); return; }



                let html = '<div class="status status-success">✅ '+data.urs.toLocaleString()+' URs | '+formatBRL(data.valor)+' | '+data.periodo+'</div>';

                if (data.ur_parcial && data.ur_parcial.disponivel_total > data.valor) {
                    var economia = data.ur_parcial.disponivel_total - data.valor;
                    html += '<div style="font-size:12px;color:#555;padding:4px 12px;margin-top:-6px;">'
                        + '📌 URs selecionadas somam ' + formatBRL(data.ur_parcial.disponivel_total)
                        + ' — valor efetivamente onerado: <strong>' + formatBRL(data.valor) + '</strong>'
                        + ' (cessão parcial de ' + formatBRL(data.ur_parcial.remanescente) + ' poupada na última UR)'
                        + '</div>';
                }



                if (data.aviso) {
                    html += '<div class="status mt-2" style="background:#FFF8E1;color:#E65100;border:1px solid #F9A825;padding:10px;font-size:12px;">'
                        + '⚠️ ' + data.aviso + '</div>';
                }

                if (data.ur_parcial) {
                    var up = data.ur_parcial;
                    html += '<div class="status mt-2" style="background:#E8F5E9;border:1px solid #A5D6A7;padding:12px 16px;font-size:12px;border-radius:8px;">'
                        + '<div style="font-weight:700;color:#1B5E20;margin-bottom:6px;">✂️ Cessão parcial aplicada na última UR</div>'
                        + '<table style="width:100%;border-collapse:collapse;font-size:12px;">'
                        + '<tr><td style="color:#555;padding:2px 0;width:140px;">ID da UR</td>'
                        + '<td style="font-family:monospace;color:#1B5E20;">' + up.receivable_id + '</td></tr>'
                        + '<tr><td style="color:#555;padding:2px 0;">Vencimento</td>'
                        + '<td>' + up.data_liquidacao + ' &nbsp;·&nbsp; ' + up.adquirente + ' / ' + up.arranjo + '</td></tr>'
                        + '<tr><td style="color:#555;padding:2px 0;">Disponível na UR</td>'
                        + '<td>' + formatBRL(up.disponivel) + '</td></tr>'
                        + '<tr><td style="color:#555;padding:2px 0;font-weight:600;">Valor cedido</td>'
                        + '<td style="font-weight:700;color:#2E7D32;">' + formatBRL(up.cedido) + '</td></tr>'
                        + '<tr><td style="color:#555;padding:2px 0;">Remanescente na UR</td>'
                        + '<td style="color:#757575;">' + formatBRL(up.remanescente) + '</td></tr>'
                        + '</table></div>';
                }



                if (lastIneligibleExcluded && lastIneligibleExcluded.length > 0) {



                    const excForEmp = lastIneligibleExcluded.filter(function(s) { return s.empresa === empresa; });



                    if (excForEmp.length > 0) {



                        html += '<div class="status mt-2" style="background:#FFF8E1;color:#E65100;border:1px solid #F9A825;padding:10px;font-size:12px;">'



                            + '⚠️ <strong>' + excForEmp.length + ' seller(s) inelegivel(is) excluído(s) desta empresa:</strong> '



                            + excForEmp.map(function(s){ return s.seller_id; }).join(', ') + '</div>';



                    }



                }



                var _empresaNome = encodeURIComponent(empresa || '');

                var _urs  = data.urs || 0;

                var _vt   = data.valor_total || data.valor || 0;

                var _vo   = data.valor_operavel || data.valor || 0;

                var _taxa = data.taxa || 0;

                html += '<div class="mt-2" style="display:flex;gap:8px;align-items:center;">'

                    + '<a href="/download/'+sessionId+'/custom" class="btn btn-download">⬇️ Download Personalizada</a>'

                    + '<button class="btn-email-action-custom" '

                    + 'data-sid="'+sessionId+'" data-safe="custom" '

                    + 'data-nome="'+_empresaNome+'" '

                    + 'data-urs="'+_urs+'" '

                    + 'data-valor="'+_vt+'" '

                    + 'data-operavel="'+_vo+'" '

                    + 'data-taxa="'+_taxa+'" '

                    + 'style="padding:8px 14px;font-size:12px;background:#E3F2FD;border:1px solid #90CAF9;border-radius:6px;cursor:pointer;">'

                    + '&#9993; Enviar por e-mail</button>'

                    + '</div>';



                document.getElementById('custom-result').innerHTML = html;



                document.getElementById('custom-result').classList.remove('hidden');



            });



        }



        function saveRaizes() {



            const inputs = document.querySelectorAll('[id^="raiz-"]');



            const raizes = {};



            let allFilled = true;



            inputs.forEach(inp => {



                const raiz = inp.id.replace('raiz-', '');



                const nome = inp.value.trim();



                if (!nome) { allFilled = false; inp.style.borderColor = '#C62828'; }



                else { inp.style.borderColor = '#4CAF50'; raizes[raiz] = nome; }



            });



            if (!allFilled) { alert('Preencha o nome de todas as empresas'); return; }



            fetch('/add_raizes', {



                method: 'POST',



                headers: {'Content-Type': 'application/json'},



                body: JSON.stringify({raizes: raizes})



            })



            .then(r => r.json())



            .then(data => {



                if (data.error) { alert(data.error); return; }



                document.getElementById('step-unknown').innerHTML = '<div class="status status-success">Empresas salvas! Serao lembradas nas proximas vezes.</div>';



                document.getElementById('step-params').classList.remove('hidden');



                



                // Reload do resumo com novos nomes - re-upload



                document.getElementById('file-input').dispatchEvent(new Event('change'));



            });



        }



        // Elegibilidade



        let eligibilityPanelParams = null;



        function showIneligibilityPanel(ineligibleSellers, idsFound, idsMissing, sessionId, params) {

            eligibilityPanelParams = { sessionId, params };



            // Detectar se e cenario de elegibilidade indisponivel

            var hasUnverified = ineligibleSellers.some(function(s) { return s.unverified; });

            var totalSellers  = ineligibleSellers.length;



            // --- Estado de paginacao ---

            var _page     = 0;

            var _pageSize = 25;



            function _renderPanel() {

                var start   = _page * _pageSize;

                var end     = _pageSize === 0 ? totalSellers : Math.min(start + _pageSize, totalSellers);

                var slice   = ineligibleSellers.slice(start, end);

                var totalPg = _pageSize === 0 ? 1 : Math.ceil(totalSellers / _pageSize);



                // Titulo e descricao adaptados

                var titulo = hasUnverified

                    ? '\u26a0\ufe0f Elegibilidade indispon\u00edvel'

                    : '\u26a0\ufe0f Sellers inelegi\u0301veis encontrados';

                var descricao = hasUnverified

                    ? 'N\u00e3o foi poss\u00edvel verificar a elegibilidade no Databricks. '

                      + 'Os <strong>' + totalSellers + ' sellers</strong> listados abaixo n\u00e3o puderam ser verificados. '

                      + 'Voc\u00ea pode gerar a cota\u00e7\u00e3o mesmo assim \u2014 as URs aparecer\u00e3o em vermelho indicando que o valor antecip\u00e1vel n\u00e3o foi confirmado.'

                    : 'Seller IDs encontrados: <strong>' + idsFound + '</strong>'

                      + (idsMissing > 0 ? ' \u2014 Sem cadastro: <strong>' + idsMissing + '</strong>' : '')

                      + '. Para gerar <strong>Cota\u00e7\u00e3o + Sele\u00e7\u00e3o de URs</strong>, decida abaixo o que fazer:';



                var html = '<div class="eligibility-panel" id="eligibilityPanel">';

                html += '<h3>' + titulo + '</h3>';

                html += '<p style="font-size:13px;color:#5D4037;margin-bottom:12px;">' + descricao + '</p>';



                // Controles de paginacao (topo)

                html += '<div style="display:flex;align-items:center;gap:12px;margin-bottom:12px;flex-wrap:wrap;">';

                html += '<span style="font-size:12px;color:#5D4037;">';

                if (_pageSize === 0) {

                    html += 'Exibindo todos ' + totalSellers + ' sellers';

                } else {

                    html += 'Exibindo ' + (start+1) + '\u2013' + end + ' de ' + totalSellers + ' sellers';

                }

                html += '</span>';

                html += '<div style="margin-left:auto;display:flex;align-items:center;gap:8px;">';

                html += '<label style="font-size:12px;color:#5D4037;">Por p\u00e1gina:</label>';

                html += '<select onchange="_ineligPageSize(parseInt(this.value))" style="font-size:12px;padding:3px 6px;border:1px solid #A5D6A7;border-radius:4px;">';

                [25, 50, 100].forEach(function(n) {

                    html += '<option value="' + n + '"' + (n === _pageSize ? ' selected' : '') + '>' + n + '</option>';

                });

                html += '<option value="0"' + (_pageSize === 0 ? ' selected' : '') + '>Todos</option>';

                html += '</select>';

                if (totalPg > 1) {

                    html += '<button onclick="_ineligPrev()" ' + (_page === 0 ? 'disabled' : '') + ' style="padding:3px 10px;font-size:12px;background:#E8F5E9;border:1px solid #A5D6A7;border-radius:4px;cursor:pointer;">&laquo; Ant</button>';

                    html += '<span style="font-size:12px;color:#5D4037;">' + (_page+1) + '/' + totalPg + '</span>';

                    html += '<button onclick="_ineligNext(' + totalPg + ')" ' + (_page >= totalPg-1 ? 'disabled' : '') + ' style="padding:3px 10px;font-size:12px;background:#E8F5E9;border:1px solid #A5D6A7;border-radius:4px;cursor:pointer;">Pr\u00f3x &raquo;</button>';

                }

                html += '</div></div>';



                // Tabela

                html += '<table class="eligibility-table">';

                html += '<thead><tr><th>S\u00f3 Cota\u00e7\u00e3o</th><th>Seller ID</th><th>Empresa</th><th>CNPJs</th><th>URs</th><th>Valor Agenda</th><th>Status</th></tr></thead><tbody>';



                slice.forEach(function(s) {

                    var cnpjsStr = s.cnpjs.slice(0,2).join(', ') + (s.cnpjs.length > 2 ? ' +' + (s.cnpjs.length-2) : '');

                    var badge;

                    if (s.unverified) {

                        badge = '<span style="background:#FF9800;color:#fff;padding:2px 8px;border-radius:3px;font-size:11px;font-weight:600;">Elegib. indispon\u00edvel</span>';

                    } else if (s.absent) {

                        badge = '<span class="badge-ausente">Ausente</span>';

                    } else {

                        badge = '<span class="badge-inelegivel">total = 0</span>';

                    }

                    html += '<tr>';

                    html += '<td style="text-align:center"><input type="checkbox" class="seller-checkbox" value="' + s.seller_id + '"></td>';

                    html += '<td><strong>' + s.seller_id + '</strong></td>';

                    html += '<td style="font-size:12px">' + (s.empresa || '') + '</td>';

                    html += '<td style="font-size:11px">' + cnpjsStr + '</td>';

                    html += '<td>' + s.urs.toLocaleString('pt-BR') + '</td>';

                    html += '<td>' + formatBRL(s.valor) + '</td>';

                    html += '<td>' + badge + '</td></tr>';

                });

                html += '</tbody></table>';



                // Controles de paginacao (rodape)

                if (totalPg > 1) {

                    html += '<div style="display:flex;justify-content:center;align-items:center;gap:10px;margin-top:12px;">';

                    html += '<button onclick="_ineligPrev()" ' + (_page === 0 ? 'disabled' : '') + ' style="padding:4px 14px;font-size:12px;background:#E8F5E9;border:1px solid #A5D6A7;border-radius:4px;cursor:pointer;">&laquo; Anterior</button>';

                    html += '<span style="font-size:12px;color:#5D4037;">';

                    html += 'P\u00e1gina ' + (_page+1) + ' de ' + totalPg;

                    html += '</span>';

                    html += '<button onclick="_ineligNext(' + totalPg + ')" ' + (_page >= totalPg-1 ? 'disabled' : '') + ' style="padding:4px 14px;font-size:12px;background:#E8F5E9;border:1px solid #A5D6A7;border-radius:4px;cursor:pointer;">Pr\u00f3xima &raquo;</button>';

                    html += '</div>';

                }



                // Nota e acoes

                html += '<p class="eligibility-note" style="margin-top:12px;">';

                if (hasUnverified) {

                    html += '\u2139\ufe0f <strong>S\u00f3 Cota\u00e7\u00e3o</strong> = entra na Cota\u00e7\u00e3o (URs em vermelho) mas n\u00e3o na Sele\u00e7\u00e3o de URs. <strong>Desmarcado</strong> = exclu\u00eddo de ambos.';

                } else {

                    html += '\u2139\ufe0f <strong>S\u00f3 Cota\u00e7\u00e3o</strong> = entra na Cota\u00e7\u00e3o mas n\u00e3o na Sele\u00e7\u00e3o de URs. <strong>Desmarcado</strong> = exclu\u00eddo de ambos.';

                }

                html += '</p>';

                html += '<div class="eligibility-actions">';

                html += '<button class="btn btn-primary" onclick="generateWithSellers()">Gerar Cota\u00e7\u00e3o + Sele\u00e7\u00e3o com Seller IDs</button>';

                html += '<button class="btn btn-secondary" onclick="selectAllEligibility(true)">Marcar todos (s\u00f3 cota\u00e7\u00e3o)</button>';

                html += '<button class="btn btn-secondary" onclick="selectAllEligibility(false)">Desmarcar todos (excluir)</button>';

                html += '</div></div>';



                var stepResults = document.getElementById('step-results');

                if (stepResults) {

                    var ex = document.getElementById('eligibilityPanel');

                    if (ex) ex.remove();

                    stepResults.insertAdjacentHTML('afterend', html);

                    document.getElementById('eligibilityPanel').scrollIntoView({behavior:'smooth', block:'start'});

                    // Re-attach event listeners apos re-render

                    document.querySelectorAll('.seller-checkbox').forEach(function(cb) {

                        cb.addEventListener('change', function() {});

                    });

                }

            }



            // Funcoes de paginacao expostas ao HTML

            window._ineligPrev = function() { if (_page > 0) { _page--; _renderPanel(); } };

            window._ineligNext = function(total) { if (_page < total-1) { _page++; _renderPanel(); } };

            window._ineligPageSize = function(n) { _pageSize = n; _page = 0; _renderPanel(); };



            _renderPanel();

        }



        function showEligibilityPanel(ineligibleSellers, sessionId, params) {



            showIneligibilityPanel(ineligibleSellers, 0, 0, sessionId, params);



        }



        function renderIneligPanel(ineligByEmpresa, sid) {



            if (!ineligByEmpresa || Object.keys(ineligByEmpresa).length === 0) return '';



            const empresas = Object.keys(ineligByEmpresa).sort();



            const uid = 'inelig_' + Date.now();



            let html = '<div class="mt-3" style="border:1px solid #FFCDD2;border-radius:8px;overflow:hidden;">';



            html += '<div style="background:#FFEBEE;padding:10px 14px;border-bottom:1px solid #FFCDD2;">';



            html += '<span style="font-weight:700;color:#C62828;font-size:13px;">&#9888; Sellers Inelegi&#769;veis</span>';



            html += '<span style="font-size:12px;color:#757575;margin-left:8px;">' + empresas.length + ' empresa(s) afetada(s)</span>';



            html += '</div>';



            empresas.forEach(function(emp) {



                const sellers = ineligByEmpresa[emp];



                const empUid = uid + '_' + emp.replace(/[^a-zA-Z0-9]/g, '_');



                const totalUrs = sellers.reduce(function(a, s) { return a + s.urs; }, 0);



                const totalVal = sellers.reduce(function(a, s) { return a + s.valor; }, 0);



                const safeName = emp.replace(/ /g, '_').replace(/[^a-zA-Z0-9_-]/g, '');



                html += '<div style="border-bottom:1px solid #FFCDD2;">';



                html += '<div data-toggle-target="' + empUid + '" class="seller-summary-toggle" style="padding:10px 14px;cursor:pointer;display:flex;align-items:center;justify-content:space-between;background:#FFF5F5;">';



                html += '<span style="font-weight:600;font-size:12px;color:#C62828;">' + emp + '</span>';



                html += '<span style="display:flex;align-items:center;gap:10px;">';



                html += '<span style="font-size:11px;color:#757575;">' + sellers.length + ' seller(s) &bull; ' + totalUrs.toLocaleString('pt-BR') + ' URs &bull; ' + formatBRL(totalVal) + '</span>';



                if (sid && safeName) {



                    html += '<a href="/download_inelegiveis/' + sid + '/' + safeName + '" onclick="event.stopPropagation()" class="btn btn-secondary" style="padding:3px 10px;font-size:11px;color:#C62828;border-color:#FFCDD2;">&#11015; CSV</a>';



                }



                html += '<span style="font-size:11px;color:#9E9E9E;">&#9662;</span>';



                html += '</span></div>';



                html += '<div id="' + empUid + '" class="hidden" style="padding:0 14px 10px;">';



                html += '<table style="width:100%;font-size:11px;border-collapse:collapse;margin-top:8px;">';



                html += '<tr style="background:#FFEBEE;"><th style="padding:5px 8px;text-align:left;">Seller ID</th><th style="padding:5px 8px;text-align:left;">CNPJs</th><th style="padding:5px 8px;text-align:right;">URs</th><th style="padding:5px 8px;text-align:right;">Valor Bruto</th></tr>';



                sellers.forEach(function(s) {



                    html += '<tr style="border-top:1px solid #FFCDD2;"><td style="padding:5px 8px;font-family:monospace;color:#B71C1C;">' + s.seller_id + '</td><td style="padding:5px 8px;color:#616161;font-size:10px;">' + (s.cnpjs || []).join(', ') + '</td><td style="padding:5px 8px;text-align:right;">' + (s.urs || 0).toLocaleString('pt-BR') + '</td><td style="padding:5px 8px;text-align:right;">' + formatBRL(s.valor || 0) + '</td></tr>';



                });



                html += '</table></div></div>';



            });



            html += '</div>';



            return html;



        }



        function renderSellerSummary(summary) {



            if (!summary || summary.length === 0) return '';



            const total = summary.length;



            const inelig = summary.filter(s => !s.eligible).length;



            const uid = 'ss_' + Date.now();



            let badge = inelig > 0



                ? '<span style="background:#FFCDD2;color:#B71C1C;border-radius:12px;padding:2px 10px;font-size:11px;margin-left:8px;">' + inelig + ' inelegi&#769;vel' + (inelig > 1 ? 'eis' : '') + '</span>'



                : '<span style="background:#C8E6C9;color:#1B5E20;border-radius:12px;padding:2px 10px;font-size:11px;margin-left:8px;">todos elegi&#769;veis</span>';



            let html = '<div class="mt-2" style="border:1px solid #E0E0E0;border-radius:6px;overflow:hidden;">';



            html += '<div data-toggle-target="' + uid + '" class="seller-summary-toggle" '



                 + 'style="background:#F5F5F5;padding:10px 14px;cursor:pointer;display:flex;align-items:center;justify-content:space-between;user-select:none;">';



            html += '<span style="font-weight:600;font-size:13px;color:#212121;">&#128202; Resumo por Seller <span style="font-weight:400;color:#757575;">(' + total + ' seller' + (total > 1 ? 's' : '') + ')</span>' + badge + '</span>';



            html += '<span style="color:#9E9E9E;font-size:12px;">clique para expandir &#9662;</span></div>';



            html += '<div id="' + uid + '" class="hidden">';



            html += '<table style="width:100%;font-size:12px;border-collapse:collapse;">';



            html += '<tr style="background:#EEEEEE;">'



                 + '<th style="padding:7px 10px;text-align:left;font-weight:600;">Seller ID</th>'



                 + '<th style="padding:7px 10px;text-align:left;font-weight:600;">Empresa</th>'



                 + '<th style="padding:7px 10px;text-align:left;font-weight:600;">CNPJs</th>'



                 + '<th style="padding:7px 10px;text-align:center;font-weight:600;">Status</th>'



                 + '<th style="padding:7px 10px;text-align:right;font-weight:600;">URs</th>'



                 + '<th style="padding:7px 10px;text-align:right;font-weight:600;">Valor Bruto</th>'



                 + '</tr>';



            summary.forEach(function(s) {



                const rowStyle = s.eligible



                    ? ''



                    : 'background:#FFEBEE;color:#B71C1C;';



                const statusBadge = s.eligible



                    ? '<span style="background:#E8F5E9;color:#1B5E20;border-radius:10px;padding:2px 8px;font-size:11px;font-weight:600;">&#10003; Elegi&#769;vel</span>'



                    : '<span style="background:#FFCDD2;color:#B71C1C;border-radius:10px;padding:2px 8px;font-size:11px;font-weight:600;">&#8855; Inelegi&#769;vel</span>';



                const cnpjsStr = (s.cnpjs || []).join(', ') || '&#8212;';



                html += '<tr style="border-top:1px solid #EEEEEE;' + rowStyle + '">'



                     + '<td style="padding:6px 10px;font-family:monospace;">' + s.seller_id + '</td>'



                     + '<td style="padding:6px 10px;">' + s.empresa + '</td>'



                     + '<td style="padding:6px 10px;font-size:11px;color:#616161;">' + cnpjsStr + '</td>'



                     + '<td style="padding:6px 10px;text-align:center;">' + statusBadge + '</td>'



                     + '<td style="padding:6px 10px;text-align:right;">' + (s.urs || 0).toLocaleString('pt-BR') + '</td>'



                     + '<td style="padding:6px 10px;text-align:right;">' + formatBRL(s.valor || 0) + '</td>'



                     + '</tr>';



            });



            html += '</table></div></div>';



            return html;



        }



        // Toggle handler para paineis seller-summary (delegado ao document)



        document.addEventListener('click', function(e) {



            var toggle = e.target.closest('.seller-summary-toggle');



            if (!toggle) return;



            var targetId = toggle.getAttribute('data-toggle-target');



            var el = document.getElementById(targetId);



            if (el) el.classList.toggle('hidden');



        });



        function generateWithSellers() {



            const p = eligibilityPanelParams ? eligibilityPanelParams.params : null;



            const sid = eligibilityPanelParams ? eligibilityPanelParams.sessionId : null;



            if (!p || !sid) return;



            const forceCotacaoOnly = [];



            document.querySelectorAll('.seller-checkbox:checked').forEach(cb => forceCotacaoOnly.push(cb.value));



            const onlyElig = false; // sempre gerar cotacao completa quando ha inelegiveis

            const payload = Object.assign({}, p, { session_id: sid, force_cotacao_only: forceCotacaoOnly, eligibility_confirmed: true, only_eligible: onlyElig, taxa_map: getTaxaMap() });



            const panel = document.getElementById('eligibilityPanel');



            if (panel) panel.innerHTML = '<p style="padding:16px;color:#1B5E20;font-weight:600;">⏳ Gerando Cotação + Seleção de URs com Seller IDs... aguarde.</p>';



            fetch('/fetch_sellers', {



                method: 'POST',



                headers: {'Content-Type': 'application/json'},



                body: JSON.stringify(payload)



            })



            .then(r => r.json())



            .then(function(data) {



                if (data.requires_eligibility_confirmation) {



                    showIneligibilityPanel(data.ineligible_sellers, data.seller_ids_found, data.seller_ids_missing, sid, p);



                    return;



                }



                if (panel) panel.remove();



                // Atualizar sellerStatus



                const sellerStatus = document.getElementById('sellerStatus');



                if (sellerStatus) {



                    sellerStatus.className = 'status status-success mt-2';



                    sellerStatus.style.cssText = '';



                    sellerStatus.innerHTML = '✅ Cotação + Seleção gerados! Seller IDs: ' + data.seller_ids_found + ' encontrados'



                        + (data.seller_ids_missing > 0 ? ', ' + data.seller_ids_missing + ' faltantes' : '');



                }



                // Mostrar step-results e atualizar tabela



                const stepResultsAfter = document.getElementById('step-results');



                if (stepResultsAfter) stepResultsAfter.classList.remove('hidden');



                if (data.empresas) {



                    let ineligInfo = '';



                    if (data.ineligible_excluded && data.ineligible_excluded.length > 0) {



                        ineligInfo = '<div class="status mt-2" style="background:#FFF8E1;color:#E65100;border:1px solid #F9A825;padding:10px;font-size:13px;">'



                            + '<strong>⚠️ ' + data.ineligible_excluded.length + ' seller(s) excluído(s) desta geração:</strong> '



                            + data.ineligible_excluded.map(function(s) { return '<strong>' + s.seller_id + '</strong> (' + s.empresa + ')'; }).join(', ')



                            + '</div>';



                    }



                    let html = '<div class="status status-success">✅ ' + data.empresas.length + ' empresas geradas com Seller IDs | ' + data.seller_ids_found + ' encontrados'



                        + (data.seller_ids_missing > 0 ? ', ' + data.seller_ids_missing + ' faltantes' : '') + '</div>';



                    html += ineligInfo;



                    html += '<table class="mt-2"><tr><th>Empresa</th><th class="text-right">URs</th><th class="text-right">Valor Bruto</th><th class="text-right">Operável</th><th class="text-right">Taxa</th><th>Ações</th></tr>';



                    data.empresas.forEach(function(e) {
                        var taxaLabel2 = e.taxa ? (parseFloat(e.taxa).toFixed(2) + '%') : '-';
                        var opLabel2   = e.valor_operavel > 0 ? formatBRL(e.valor_operavel) : '<span style="color:#9E9E9E;">-</span>';
                        html += '<tr>';
                        html += '<td><strong>' + e.nome + '</strong></td>';
                        html += '<td class="text-right">' + e.urs.toLocaleString('pt-BR') + '</td>';
                        html += '<td class="text-right">' + formatBRL(e.valor) + '</td>';
                        html += '<td class="text-right" style="color:#1B5E20;font-weight:600;">' + opLabel2 + '</td>';
                        html += '<td class="text-right" style="color:#1B5E20;font-weight:600;">' + taxaLabel2 + '</td>';
                        html += '<td>';



                        



                    html += '<a href="/download/' + sid + '/' + encodeURIComponent(e.safe_name) + '" class="btn btn-secondary" style="padding:6px 12px;font-size:12px;">⬇ Download</a>';





                    

                                            html += '<button class="btn btn-email-action" data-sid="'+sid+'" data-safe="'+encodeURIComponent(e.safe_name)+'" data-nome="'+encodeURIComponent(e.nome)+'" data-urs="'+e.urs+'" data-valor="'+(e.valor||0)+'" data-operavel="'+(e.valor_operavel||0)+'" data-taxa="'+(e.taxa||0)+'" style="padding:6px 12px;font-size:12px;background:#E3F2FD;border:1px solid #90CAF9;border-radius:6px;cursor:pointer;margin-left:4px;">&#9993;&#65039; E-mail</button>';
                        html += '</td></tr>';
                    });



                    html += '</table>';



                    const rc = document.getElementById('results-content');



                    if (rc) {



                        rc.innerHTML = html;



                        // Painel de inelegiveis por empresa



                        if (data.inelig_by_empresa && Object.keys(data.inelig_by_empresa).length > 0) {



                            rc.insertAdjacentHTML('beforeend', renderIneligPanel(data.inelig_by_empresa, sid));



                        }



                        // Painel de resumo por seller



                        if (data.seller_summary && data.seller_summary.length > 0) {



                            rc.insertAdjacentHTML('beforeend', renderSellerSummary(data.seller_summary));



                        }
                        if (data.missing_by_empresa && Object.keys(data.missing_by_empresa).length > 0) {
                            rc.insertAdjacentHTML('beforeend', renderMissingPanel(data.missing_by_empresa));
                        }



                    }



                }



                document.getElementById('step-custom').classList.remove('hidden');



                if (data.ineligible_excluded) lastIneligibleExcluded = data.ineligible_excluded;



                loadHistory();



            })



            .catch(function(err) {



                if (panel) panel.innerHTML = '<p style="padding:12px;color:#B71C1C;">Erro: ' + err + '</p>';



            });



        }



        function confirmEligibility() {



            generateWithSellers();



        }



        function selectAllEligibility(checked) {



            document.querySelectorAll('.seller-checkbox').forEach(cb => cb.checked = checked);



        }



        



        // Historico



        // Prefetch de sellers em background



        let prefetchSessionId = null;



        let prefetchInterval = null;



        let prefetchDone = false;



        let _prefetchStartTime = null;



        let _prefetchElapsedTimer = null;



        function startPrefetch(sid) {



            prefetchSessionId = sid;



            prefetchDone = false;



            _prefetchStartTime = Date.now();



            // Criar indicador de status abaixo do upload



            let indicator = document.getElementById('prefetch-indicator');



            if (!indicator) {



                indicator = document.createElement('div');



                indicator.id = 'prefetch-indicator';



                indicator.style.cssText = 'font-size:12px;color:#757575;margin-top:8px;padding:8px 12px;'



                    + 'background:#F5F5F5;border-radius:6px;border-left:3px solid #BDBDBD;'



                    + 'display:flex;align-items:center;gap:8px;';



                const uploadStatus = document.getElementById('upload-status');



                if (uploadStatus) uploadStatus.after(indicator);



                else {



                    const stepGen = document.getElementById('step-generate');



                    if (stepGen) stepGen.appendChild(indicator);



                }



            }



            indicator.style.borderLeftColor = '#BDBDBD';



            indicator.innerHTML = '<span class="loader" style="width:12px;height:12px;border-width:2px;"></span>'



                + ' <span id="prefetch-msg">Verificando sellers e elegibilidade... <strong>você já pode continuar</strong></span>'



                + ' <span id="prefetch-elapsed" style="margin-left:auto;color:#9E9E9E;">0s</span>';



            // Atualizar elapsed a cada segundo



            if (_prefetchElapsedTimer) clearInterval(_prefetchElapsedTimer);



            _prefetchElapsedTimer = setInterval(function() {



                const el = document.getElementById('prefetch-elapsed');



                if (el && _prefetchStartTime) {



                    const sec = Math.floor((Date.now() - _prefetchStartTime) / 1000);



                    el.textContent = sec + 's';



                }



            }, 1000);



            // Disparar prefetch



            fetch('/prefetch_sellers', {



                method: 'POST',



                headers: {'Content-Type': 'application/json'},



                body: JSON.stringify({session_id: sid})



            }).catch(() => {});



            // Monitorar progresso a cada 5s



            if (prefetchInterval) clearInterval(prefetchInterval);



            prefetchInterval = setInterval(checkPrefetchStatus, 5000);



            setTimeout(checkPrefetchStatus, 8000);



        }



        function checkPrefetchStatus() {



            if (!prefetchSessionId || prefetchDone) return;



            fetch('/prefetch_status?session_id=' + prefetchSessionId)



            .then(r => r.json())



            .then(function(data) {



                const indicator = document.getElementById('prefetch-indicator');



                const msg = document.getElementById('prefetch-msg');



                const elapsed = _prefetchStartTime



                    ? Math.floor((Date.now() - _prefetchStartTime) / 1000) + 's'



                    : '';



                if (data.status === 'done') {



                    prefetchDone = true;



                    if (prefetchInterval) { clearInterval(prefetchInterval); prefetchInterval = null; }



                    if (_prefetchElapsedTimer) { clearInterval(_prefetchElapsedTimer); _prefetchElapsedTimer = null; }



                    if (indicator) {



                        if (data.error && !data.seller_ids_found) {



                            indicator.style.borderLeftColor = '#E65100';



                            indicator.innerHTML = '⚠️ <span>Sellers: erro (' + data.error.substring(0,80) + ')</span>'



                                + '<span style="margin-left:auto;color:#9E9E9E;">' + elapsed + '</span>';



                        } else {



                            indicator.style.borderLeftColor = '#4CAF50';



                            indicator.innerHTML = '✅ <span><strong>' + (data.seller_ids_found || 0)



                                + ' Seller IDs</strong> verificados — elegibilidade pronta</span>'



                                + '<span style="margin-left:auto;color:#9E9E9E;">' + elapsed + '</span>';



                            indicator.style.color = '#2E7D32';



                            // Recarregar empresa-info se usuario ja selecionou empresa



                            const empSel = document.getElementById('custom-empresa');



                            if (empSel && empSel.value) loadEmpresaFilters();



                        }



                    }



                } else if (data.status === 'running') {



                    if (indicator) {



                        const sec = _prefetchStartTime



                            ? Math.floor((Date.now() - _prefetchStartTime) / 1000)



                            : 0;



                        const est = sec < 30 ? ' (est. ~60s)' : sec < 90 ? ' (est. ~2min)' : ' (aguarde...)';



                        if (msg) msg.innerHTML = 'Verificando sellers e elegibilidade' + est



                            + ' — <strong>você já pode continuar</strong>';



                    }



                }



            }).catch(() => {});



        }



        function loadHistory() {



            document.getElementById('history-sync-status').innerHTML = '<span class="loader" style="width:14px;height:14px;"></span> Carregando...';



            fetch('/history')



            .then(r => r.json())



            .then(data => {



                const el = document.getElementById('history-content');



                document.getElementById('history-sync-status').innerHTML = data.length + ' registros';



                if (!data || data.length === 0) {



                    el.innerHTML = '<p style="color:#9E9E9E;">Nenhuma cotacao gerada ainda.</p>';



                    return;



                }



                let html = '<table><tr><th>Data</th><th>Operador</th><th>Maquina</th><th>Tipo</th><th>Empresas</th><th class="text-right">URs</th><th class="text-right">Valor</th><th>Taxa</th><th>DI</th><th>Acoes</th></tr>';



                data.forEach(h => {



                    const empresas = (h.empresas || []).join(', ');



                    const empresasShort = empresas.length > 40 ? empresas.substring(0, 37) + '...' : empresas;



                    html += '<tr>';



                    html += '<td style="white-space:nowrap;">' + h.timestamp + '</td>';



                    html += '<td><strong>' + (h.operador || '').split('@')[0] + '</strong></td>';



                    html += '<td style="font-size:11px;color:#757575;">' + (h.usuario_maquina || '') + '</td>';



                    html += '<td><span class="badge ' + (h.tipo === 'Completa' ? 'badge-green' : 'badge-blue') + '">' + (h.tipo || '') + '</span></td>';



                    html += '<td style="font-size:11px;" title="' + empresas + '">' + empresasShort + '</td>';



                    html += '<td class="text-right">' + (h.total_urs || 0).toLocaleString() + '</td>';



                    html += '<td class="text-right">' + formatBRL(h.total_valor || 0) + '</td>';



                    html += '<td>' + (h.taxa || 0) + '%</td>';



                    html += '<td>' + ((h.di_periodo || 0) * 100).toFixed(1) + '%</td>';



                    html += '<td>';



                    if (h.session_id) {



                        html += '<a href="/download_all/' + h.session_id + '" class="btn btn-secondary" style="padding:4px 10px;font-size:11px;">Download</a>';



                    }



                    html += '</td></tr>';



                });



                html += '</table>';



                el.innerHTML = html;



            })



            .catch(() => {



                document.getElementById('history-sync-status').innerHTML = 'Erro ao carregar';



            });



        }



        function retrySellers() {



            const taxa = parseFloat(document.getElementById('taxa').value);



            const di = parseFloat(document.getElementById('di-periodo').value) / 100;



            const email = getOperatorEmail();



            const sellerStatus = document.querySelector('.status.status-error.mt-2') 



                || document.querySelector('.status.mt-2');



            if (sellerStatus) {



                sellerStatus.className = 'status status-info mt-2';



                sellerStatus.innerHTML = '<span class="loader"></span> Tentando novamente buscar Seller IDs...';



            }



            fetch('/fetch_sellers', {



                method: 'POST',



                headers: {'Content-Type': 'application/json'},



                body: JSON.stringify({session_id: sessionId, taxa: taxa, di_periodo: di, operator_email: email})



            })



            .then(r => r.json())



            .then(sellerData => {



                if (sellerData.seller_error) {



                    sellerStatus.className = 'status status-error mt-2';



                    sellerStatus.innerHTML = '<strong>Erro ao buscar Seller IDs:</strong> ' + sellerData.seller_error



                        + '<br><button class="btn btn-secondary" style="padding:4px 12px;font-size:12px;margin-top:8px;" '



                        + 'onclick="retrySellers()">Tentar novamente</button>';



                } else {



                    sellerStatus.className = 'status status-success mt-2';



                    sellerStatus.innerHTML = 'Seller IDs: ' + sellerData.seller_ids_found + ' encontrados'



                        + (sellerData.seller_ids_missing > 0 ? ', ' + sellerData.seller_ids_missing + ' faltantes' : '')



                        + ' (arquivos atualizados)';



                }



            })



            .catch(() => {



                sellerStatus.className = 'status status-error mt-2';



                sellerStatus.innerHTML = 'Falha na comunicacao. Verifique sua conexao/VPN. '



                    + '<button class="btn btn-secondary" style="padding:4px 12px;font-size:12px;margin-top:8px;" '



                    + 'onclick="retrySellers()">Tentar novamente</button>';



            });



        }



        
        // ── CALCULADORA AR (por empresa) ──────────────────────────────────
        var _arTaxaMap = {};

        function _arGetTaxaMap() {
            var map = {};
            document.querySelectorAll('.ar-taxa-emp').forEach(function(inp) {
                var nome = inp.dataset.empresa;
                var v = parseFloat(inp.value);
                if (nome && !isNaN(v) && v > 0) map[nome] = v / 100.0;
            });
            return map;
        }

        var _arDebounceTimers = {};
        function debounceCalcAR(empresa) {
            var key = empresa || '__all__';
            clearTimeout(_arDebounceTimers[key]);
            _arDebounceTimers[key] = setTimeout(function(){ calcularAR(); }, 500);
        }

        function setArTaxaAll(pct) {
            document.querySelectorAll('.ar-taxa-emp').forEach(function(inp){ inp.value = pct.toFixed(2); });
            calcularAR();
        }

        function calcularAR() {
            if (!sessionId) return;
            var elRes = document.getElementById('ar-results');
            if (!elRes) return;
            elRes.innerHTML = '<span class="loader" style="width:14px;height:14px;border-width:2px;display:inline-block;"></span> <span style="font-size:12px;color:#555;">Calculando...</span>';

            var taxaMap     = _arGetTaxaMap();
            var taxaDefault = parseFloat((document.getElementById('ar-taxa-default') || {}).value || '1.37') / 100.0;

            fetch('/calcular_ar', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({session_id: sessionId, taxa_map: taxaMap, taxa_default: taxaDefault})
            })
            .then(function(r){ return r.json(); })
            .then(function(d) {
                if (d.erro) { elRes.innerHTML = '<span style="color:#C62828;font-size:13px;">Erro: ' + d.erro + '</span>'; return; }

                function fBRL(v)  { return 'R$\u00a0' + (v||0).toLocaleString('pt-BR', {minimumFractionDigits:0, maximumFractionDigits:0}); }
                function fPct(v)  { return ((v||0)*100).toFixed(2) + '%'; }
                function fPct4(v) { return ((v||0)*100).toFixed(4) + '%'; }

                var empresas = d.empresas || [];

                // Inicializar inputs na div oculta (para _arGetTaxaMap funcionar)
                var arInp = document.getElementById('ar-inputs');
                if (arInp && !arInp.querySelector('.ar-taxa-emp')) {
                    arInp.innerHTML = empresas.map(function(e) {
                        return '<input type="number" class="ar-taxa-emp" data-empresa="' + (e.nome||'').replace(/"/g,'&quot;') + '" value="' + (((e.taxa||0.0137)*100).toFixed(2)) + '" min="0.01" max="50" step="0.01">';
                    }).join('');
                }

                // Tabela unificada: uma linha por empresa
                var html = '<div style="overflow-x:auto;">' +
                    '<table style="width:100%;border-collapse:collapse;font-size:12px;">' +
                    '<thead><tr style="background:#E8F5E9;">' +
                    '<th style="padding:7px 10px;text-align:left;border:1px solid #C8E6C9;white-space:nowrap;">Decis\u00e3o</th>' +
                    '<th style="padding:7px 10px;text-align:left;border:1px solid #C8E6C9;">Empresa</th>' +
                    '<th style="padding:7px 10px;text-align:center;border:1px solid #C8E6C9;white-space:nowrap;">Taxa %a.m.</th>' +
                    '<th style="padding:7px 10px;text-align:right;border:1px solid #C8E6C9;white-space:nowrap;">Rec. Bruta</th>' +
                    '<th style="padding:7px 10px;text-align:right;border:1px solid #C8E6C9;white-space:nowrap;">COF</th>' +
                    '<th style="padding:7px 10px;text-align:right;border:1px solid #C8E6C9;white-space:nowrap;">Custo CERC</th>' +
                    '<th style="padding:7px 10px;text-align:right;border:1px solid #C8E6C9;white-space:nowrap;">Margem</th>' +
                    '<th style="padding:7px 10px;text-align:right;border:1px solid #C8E6C9;white-space:nowrap;">% CDI</th>' +
                    '<th style="padding:7px 10px;text-align:right;border:1px solid #C8E6C9;white-space:nowrap;">ROIC a.a.</th>' +
                    '<th style="padding:7px 10px;text-align:right;border:1px solid #C8E6C9;white-space:nowrap;">Cria\u00e7\u00e3o Valor</th>' +
                    '</tr></thead><tbody>';

                // Preservar taxas digitadas pelo usuario (nao sobrescrever com valor do servidor)
                var _taxasAtuais = {};
                document.querySelectorAll('.ar-taxa-emp').forEach(function(inp){ var v = parseFloat(inp.value); if (inp.dataset.empresa && !isNaN(v) && v > 0) _taxasAtuais[inp.dataset.empresa] = v; });
                empresas.forEach(function(e, i) {
                    var semCor    = e.semaforo_cor || '#388E3C';
                    var semTxt    = e.semaforo     || '-';
                    var rowBg     = i % 2 === 0 ? 'white' : '#F9FBE7';
                    var margCor   = (e.margem||0) >= 0 ? '#2E7D32' : '#C62828';
                    var cvCor     = (e.criacao_valor||0) >= 0 ? '#1B5E20' : '#C62828';
                    var cercVal      = e.custos_cerc || 0;
                    var cercBatch    = e.cerc_batch_carrego || 0;
                    var cercReg      = e.cerc_registro || 0;
                    var cercTarifa   = e.cerc_tarifa || 0;
                    var taxaPct   = (_taxasAtuais.hasOwnProperty(e.nome||'') ? _taxasAtuais[e.nome||''] : (e.taxa||0.0137)*100).toFixed(2);

                    html += '<tr style="background:' + rowBg + ';">' +
                        '<td style="padding:7px 10px;border:1px solid #E0E0E0;text-align:center;">' +
                            '<span style="background:' + semCor + ';color:white;font-size:11px;font-weight:700;padding:3px 10px;border-radius:4px;display:inline-block;min-width:62px;text-align:center;">' + semTxt + '</span>' +
                        '</td>' +
                        '<td style="padding:7px 10px;border:1px solid #E0E0E0;color:#1B5E20;font-weight:600;">' + (e.nome||'-') + '</td>' +
                        '<td style="padding:7px 10px;border:1px solid #E0E0E0;text-align:center;">' +
                            '<input type="number" class="ar-taxa-emp" data-empresa="' + (e.nome||'').replace(/"/g,'&quot;') + '"' +
                            ' value="' + taxaPct + '" min="0.01" max="50" step="0.01" oninput="debounceCalcAR()"' +
                            ' style="width:70px;padding:3px 6px;border:1px solid #ccc;border-radius:4px;font-size:12px;text-align:center;">' +
                        '</td>' +
                        '<td style="padding:7px 10px;border:1px solid #E0E0E0;text-align:right;color:#1B5E20;font-weight:600;">' + fBRL(e.receita_bruta) + '</td>' +
                        '<td style="padding:7px 10px;border:1px solid #E0E0E0;text-align:right;color:#C62828;">' + fBRL(e.cof_total) + '</td>' +
                        '<td style="padding:7px 10px;border:1px solid #E0E0E0;text-align:right;color:' + (cercVal < 0 ? '#C62828' : '#757575') + ';cursor:default;" title="Batch+Carrego: ' + fBRL(cercBatch) + ' | Registro: ' + fBRL(cercReg) + ' | Tarifa URs: ' + fBRL(cercTarifa) + '">' + (cercVal !== 0 ? fBRL(cercVal) : '<span style="color:#bbb;">—</span>') + '</td>' +
                        '<td style="padding:7px 10px;border:1px solid #E0E0E0;text-align:right;font-weight:700;color:' + margCor + ';">' + fBRL(e.margem) + '</td>' +
                        '<td style="padding:7px 10px;border:1px solid #E0E0E0;text-align:right;">' + fPct(e.pct_cdi) + '</td>' +
                        '<td style="padding:7px 10px;border:1px solid #E0E0E0;text-align:right;font-weight:600;">' + fPct(e.roic_aa) + '</td>' +
                        '<td style="padding:7px 10px;border:1px solid #E0E0E0;text-align:right;font-weight:700;color:' + cvCor + ';">' + fPct4(e.criacao_valor) + '</td>' +
                    '</tr>';
                });
                html += '</tbody></table></div>';

                if (d.aviso_curvas) html += '<div style="color:#F57C00;font-size:11px;margin-top:6px;">\u26a0 ' + d.aviso_curvas + '</div>';
                elRes.innerHTML = html;
            })
            .catch(function(err) {
                elRes.innerHTML = '<span style="color:#C62828;font-size:13px;">Erro de comunica\u00e7\u00e3o: ' + err + '</span>';
            });
        }
        // ── FIM CALCULADORA AR ─────────────────────────────────────────────

function downloadAll() {



            window.location.href = '/download_all/' + sessionId;



        }



        // ================================================================

        // E-MAIL: modal + config

        // ================================================================

        function openEmailModal(sid, safeName, empresaEnc, urs, valorTotal, valorOp, taxaPct) {

            var empresa = decodeURIComponent(empresaEnc);

            var safe    = decodeURIComponent(safeName);

            fetch('/email_destinatarios')

            .then(function(r){ return r.json(); })

            .then(function(dest){

                var _esr = dest[empresa.toUpperCase()];
                var emailSalvo = Array.isArray(_esr) ? _esr : (_esr ? [_esr] : []);

                function fmtBRL(v) {

                    return 'R$ ' + parseFloat(v).toLocaleString('pt-BR', {minimumFractionDigits:2, maximumFractionDigits:2});

                }

                var pct = (taxaPct !== undefined && taxaPct !== null) ? taxaPct : (valorTotal > 0 ? Math.round(valorOp/valorTotal*100) : 0);



                // Overlay

                var overlay = document.createElement('div');

                overlay.id = 'emailModal';

                overlay.style.cssText = 'position:fixed;top:0;left:0;width:100%;height:100%;background:rgba(0,0,0,0.5);z-index:9999;display:flex;align-items:center;justify-content:center;';



                // Card

                var card = document.createElement('div');

                card.style.cssText = 'background:white;border-radius:12px;padding:28px;max-width:520px;width:90%;box-shadow:0 8px 32px rgba(0,0,0,0.2);';



                // Header

                var hdr = document.createElement('div');

                hdr.style.cssText = 'display:flex;justify-content:space-between;align-items:center;margin-bottom:16px;';

                var h3 = document.createElement('h3');

                h3.style.cssText = 'margin:0;color:#1B5E20;font-size:16px;';

                h3.textContent = 'Enviar Cotação por E-mail';

                var btnClose = document.createElement('button');

                btnClose.innerHTML = '&times;';

                btnClose.style.cssText = 'background:none;border:none;font-size:20px;cursor:pointer;color:#757575;';

                btnClose.onclick = function(){ overlay.remove(); };

                hdr.appendChild(h3);

                hdr.appendChild(btnClose);



                // Info empresa

                var info = document.createElement('div');

                info.style.cssText = 'background:#F9FBE7;border-radius:8px;padding:12px 16px;margin-bottom:16px;';

                var taxaInfo = '';

                if (pct > 0) {

                    taxaInfo = ' &nbsp;|&nbsp; <span style="color:#1B5E20;font-weight:bold;">Taxa de opera\u00e7\u00e3o: ' + pct + '%</span>';

                }

                info.innerHTML =

                    '<strong style="font-size:14px;">' + empresa + '</strong><br>'

                    + '<span style="font-size:12px;color:#757575;">'

                    + urs.toLocaleString('pt-BR') + ' URs'

                    + ' &nbsp;|&nbsp; Agenda: ' + fmtBRL(valorTotal)

                    + taxaInfo

                    + '</span>';



                // Seletor de perfil do cliente
                var perfilWrap = document.createElement('div');
                perfilWrap.style.cssText = 'margin-bottom:10px;';
                var perfilLbl = document.createElement('label');
                perfilLbl.style.cssText = 'font-size:13px;font-weight:600;display:block;margin-bottom:4px;';
                perfilLbl.textContent = 'Perfil do cliente';
                var perfilSel = document.createElement('select');
                perfilSel.id = 'email-perfil-select';
                perfilSel.style.cssText = 'width:100%;padding:7px 10px;border:1px solid #E0E0E0;border-radius:6px;font-size:13px;background:white;';
                perfilSel.innerHTML = '<option value="recorrente">↻ Recorrente — Agenda disponível</option>'
                    + '<option value="novo">★ Novo cliente — Apresentação PicPay AR</option>'
                    + '<option value="taxa_zero">≈ Taxa zerada — Simulação sem compromisso</option>';
                perfilWrap.appendChild(perfilLbl); perfilWrap.appendChild(perfilSel);

                // Campo destinatarios (multiplos)
        var fldWrap = document.createElement('div');
        fldWrap.style.marginBottom = '12px';
        var lbl = document.createElement('label');
        lbl.style.cssText = 'font-size:13px;font-weight:600;display:block;margin-bottom:6px;';
        lbl.textContent = 'Destinatários';
        var emailListDiv = document.createElement('div');
        emailListDiv.style.cssText = 'display:flex;flex-direction:column;gap:5px;margin-bottom:5px;';
        function addEmailRow(v) {
            var row = document.createElement('div'); row.style.cssText = 'display:flex;gap:5px;';
            var ei = document.createElement('input');
            ei.type = 'email'; ei.className = 'email-to-input'; ei.value = v || '';
            ei.placeholder = 'destinatario@empresa.com.br';
            ei.style.cssText = 'flex:1;padding:7px 10px;border:1px solid #E0E0E0;border-radius:6px;font-size:13px;';
            var rm = document.createElement('button');
            rm.innerHTML = '&times;'; rm.type = 'button';
            rm.title = 'Remover';
            rm.style.cssText = 'border:1px solid #E0E0E0;border-radius:6px;padding:4px 8px;cursor:pointer;background:none;color:#9E9E9E;';
            rm.onclick = function() { if (emailListDiv.children.length > 1) row.remove(); };
            row.appendChild(ei); row.appendChild(rm);
            emailListDiv.appendChild(row);
        }
        // Pre-preencher com emails salvos (ou 1 campo vazio)
        var initEmails = emailSalvo.length > 0 ? emailSalvo : [''];
        initEmails.forEach(function(e) { addEmailRow(e); });
        // Botao adicionar
        var btnAdd = document.createElement('button');
        btnAdd.type = 'button'; btnAdd.textContent = '+ Adicionar destinatário';
        btnAdd.style.cssText = 'width:100%;border:1px dashed #90CAF9;color:#1565C0;background:none;border-radius:6px;padding:5px;font-size:12px;cursor:pointer;';
        btnAdd.onclick = function() { addEmailRow(''); };
        fldWrap.appendChild(lbl);
        fldWrap.appendChild(emailListDiv);
        fldWrap.appendChild(btnAdd);
        var inp = {value: ''};  // compatibilidade legada



                // Preview assunto/anexo

                var preview = document.createElement('div');

                preview.style.cssText = 'background:#F5F5F5;border-radius:6px;padding:10px 12px;margin-bottom:16px;font-size:12px;color:#616161;';

                preview.innerHTML = '<strong>Assunto:</strong> [PicPay AR] Agenda de Antecipação de Recebíveis — ' + empresa + '<br>'

                    + '<strong>Anexo:</strong> Cotacao_COMPLETO_' + safe + '.xlsx';



                // Status

                var statusDiv = document.createElement('div');

                statusDiv.id = 'emailStatus';

                statusDiv.style.cssText = 'min-height:20px;margin-bottom:8px;font-size:13px;';



                // Botoes

                var btns = document.createElement('div');

                btns.style.cssText = 'display:flex;gap:8px;justify-content:flex-end;';

                var btnCancel = document.createElement('button');

                btnCancel.textContent = 'Cancelar';

                btnCancel.style.cssText = 'padding:8px 18px;background:white;border:1px solid #E0E0E0;border-radius:6px;cursor:pointer;';

                btnCancel.onclick = function(){ overlay.remove(); };

                var btnSend = document.createElement('button');

                btnSend.innerHTML = '&#9993; Enviar';

                btnSend.style.cssText = 'padding:8px 18px;background:#4CAF50;color:white;border:none;border-radius:6px;cursor:pointer;font-weight:600;';

                btnSend.onclick = function(){ doSendEmail(sid, safe, empresaEnc, urs, valorTotal, valorOp, taxaPct||0); };

                btns.appendChild(btnCancel);

                btns.appendChild(btnSend);



                // Montar card

                card.appendChild(hdr);

                card.appendChild(info);
                card.appendChild(perfilWrap);

                card.appendChild(fldWrap);

                // Campo mensagem personalizada
                var msgWrap = document.createElement('div');
                msgWrap.style.marginBottom = '12px';
                var msgLbl = document.createElement('label');
                msgLbl.style.cssText = 'font-size:13px;font-weight:600;display:block;margin-bottom:4px;';
                msgLbl.textContent = 'Mensagem personalizada (opcional)';
                var msgTa = document.createElement('textarea');
                msgTa.id = 'emailMsg'; msgTa.rows = 3;
                msgTa.style.cssText = 'width:100%;box-sizing:border-box;padding:8px 12px;border:1px solid #E0E0E0;border-radius:6px;font-size:13px;resize:vertical;';
                msgTa.placeholder = 'Deixe em branco para usar o texto padrão. Exemplo: Conforme solicitado, segue cotação atualizada.';
                msgTa.value = window._lastEmailMsg || '';
                msgTa.addEventListener('input', function(){ window._lastEmailMsg = msgTa.value; });
                msgWrap.appendChild(msgLbl); msgWrap.appendChild(msgTa);
                card.appendChild(msgWrap);
                card.appendChild(preview);

                card.appendChild(statusDiv);

                card.appendChild(btns);

                overlay.appendChild(card);

                document.body.appendChild(overlay);

            });

        }



        function doSendEmail(sid, safe, empresaEnc, urs, valorTotal, valorOp, taxaPct) {

            var empresa  = decodeURIComponent(empresaEnc);

            var emailInputs = document.querySelectorAll('.email-to-input');
            var toEmails = emailInputs.length > 0
                ? Array.from(emailInputs).map(function(e){return e.value.trim();}).filter(Boolean)
                : ((document.getElementById('emailTo')||{}).value||'').split(',').map(function(s){return s.trim();}).filter(Boolean);
            var toEmail = toEmails.join(',');

            var statusEl = document.getElementById('emailStatus');

            if (!toEmail) {

                if (statusEl) statusEl.textContent = 'Informe o e-mail do destinatário.';

                return;

            }

            if (statusEl) statusEl.textContent = 'Enviando...';

            fetch('/send_email', {

                method: 'POST',

                headers: {'Content-Type': 'application/json'},

                body: JSON.stringify({session_id:sid, empresa:empresa, safe_name:safe,

                    to_email:toEmail, to_emails:toEmails, urs:urs, valor_total:valorTotal, valor_operavel:valorOp, taxa_pct:taxaPct||0, operator_email:getOperatorEmail(), custom_message:(document.getElementById('emailMsg')||{value:''}).value||'', perfil_cliente:(document.getElementById('email-perfil-select')||{value:'recorrente'}).value})

            })

            .then(function(r){ return r.json(); })

            .then(function(d){

                if (d.error) {

                    if (statusEl) statusEl.textContent = 'Erro: ' + d.error;

                } else {

                    if (statusEl) statusEl.textContent = d.message || 'Enviado com sucesso!';

                    setTimeout(function(){ var m=document.getElementById('emailModal'); if(m) m.remove(); }, 2000);

                }

            })

            .catch(function(e){ if(statusEl) statusEl.textContent = 'Erro de conexao: ' + e; });

        }



        function saveEmailConfig() {
            var user = (document.getElementById('cfg-email-user')||{}).value||'';
            var pass = (document.getElementById('cfg-email-pass')||{}).value||'';
            var name = (document.getElementById('cfg-email-name')||{}).value||'PicPay AR';
            var statusEl = document.getElementById('cfg-email-status');
            if (!user||!pass) { if(statusEl) statusEl.innerHTML='<span style="color:#C62828;">Preencha o e-mail e o App Password.</span>'; return; }
            if(statusEl) statusEl.innerHTML='<span style="color:#555;">Salvando e testando conexão...</span>';
            // 1. Salvar config
            fetch('/setup_email', {
                method:'POST', headers:{'Content-Type':'application/json'},
                body: JSON.stringify({smtp_user:user, smtp_pass:pass, display_name:name})
            })
            .then(function(r){ return r.json(); })
            .then(function(d){
                if (d.error) {
                    if(statusEl) statusEl.innerHTML='<span style="color:#C62828;">&#10060; Erro ao salvar: '+d.error+'</span>';
                    return;
                }
                // 2. Testar conexão automaticamente
                if(statusEl) statusEl.innerHTML='<span style="color:#1565C0;">Testando conexão com o servidor de e-mail...</span>';
                fetch('/test_smtp')
                .then(function(r){ return r.json(); })
                .then(function(t){
                    var portasOk = (t.connectivity||[]).filter(function(c){ return c.ok; });
                    var login = t.login || {};
                    if (login.ok) {
                        if(statusEl) statusEl.innerHTML='<span style="color:#2E7D32;font-weight:600;">&#10003; Conexão OK — e-mail pronto para envio</span>';
                        var pe = document.getElementById('cfg-email-pass'); if(pe) pe.value='';
                    } else if (portasOk.length === 0) {
                        if(statusEl) statusEl.innerHTML='<span style="color:#C62828;font-weight:600;">&#10060; Rede bloqueada — entre em contato com o suporte TI</span><br>'
                            +'<span style="font-size:11px;color:#555;">O firewall desta m\u00e1quina est\u00e1 bloqueando o envio de e-mails (portas 587 e 465). Fale com o TI para liberar ou use outro dispositivo.</span>';
                    } else if (!login.ok) {
                        var errMsg = (login.error||'').toLowerCase();
                        var dica = errMsg.indexOf('username') >= 0 || errMsg.indexOf('password') >= 0 || errMsg.indexOf('535') >= 0
                            ? 'Verifique se o App Password está correto. Gere um novo em myaccount.google.com > Segurança > Senhas de app.'
                            : 'Não foi possível autenticar. Verifique o e-mail e o App Password.';
                        if(statusEl) statusEl.innerHTML='<span style="color:#C62828;font-weight:600;">&#10060; Falha na autenticação</span><br>'
                            +'<span style="font-size:11px;color:#555;">'+dica+'</span>';
                    }
                })
                .catch(function(){ if(statusEl) statusEl.innerHTML='<span style="color:#C62828;">Não foi possível testar a conexão.</span>'; });
            })
            .catch(function(e){ if(statusEl) statusEl.innerHTML='<span style="color:#C62828;">Erro: '+e+'</span>'; });
        }





        // Aba Envios no historico

        function loadEmailHistory() {

            fetch('/email_history?limit=100')

            .then(function(r){ return r.json(); })

            .then(function(hist){

                var cont = document.getElementById('email-history-content');

                if (!cont) return;

                if (!hist || hist.length === 0) {

                    cont.innerHTML = '<p style="color:#9E9E9E;text-align:center;padding:20px;">Nenhum e-mail enviado ainda.</p>';

                    return;

                }

                var html = '<table style="width:100%;border-collapse:collapse;">'

                    + '<thead><tr style="background:#E8F5E9;">'

                    + '<th style="padding:8px;text-align:left;font-size:12px;">Data/Hora</th>'

                    + '<th style="padding:8px;text-align:left;font-size:12px;">Empresa</th>'

                    + '<th style="padding:8px;text-align:left;font-size:12px;">Destinat\u00e1rio</th>'

                    + '<th style="padding:8px;text-align:right;font-size:12px;">Valor Oper\u00e1vel</th>'

                    + '<th style="padding:8px;text-align:center;font-size:12px;">Status</th>'

                    + '</tr></thead><tbody>';

                hist.forEach(function(e, i) {

                    var st = e.status === 'enviado'

                        ? '<span style="color:#2E7D32;font-weight:600;">\u2705 Enviado</span>'

                        : '<span style="color:#C62828;">\u274c Erro</span>';

                    var bg = i % 2 === 0 ? '' : 'background:#F9FBE7;';

                    html += '<tr style="' + bg + '">'

                        + '<td style="padding:7px 8px;font-size:12px;">' + (e.timestamp||'') + '</td>'

                        + '<td style="padding:7px 8px;font-size:12px;font-weight:600;">' + (e.empresa||'') + '</td>'

                        + '<td style="padding:7px 8px;font-size:12px;">' + (e.to_email||'') + '</td>'

                        + '<td style="padding:7px 8px;font-size:12px;text-align:right;">'

                        + (e.valor_operavel ? 'R$ ' + parseFloat(e.valor_operavel).toLocaleString('pt-BR',{minimumFractionDigits:2}) : '-')

                        + '</td>'

                        + '<td style="padding:7px 8px;text-align:center;">' + st + '</td>'

                        + '</tr>';

                });

                html += '</tbody></table>';

                cont.innerHTML = html;

            });

        }



    </script>



</body>



</html>



'''



# ==============================================================================



# ROUTES



# ==============================================================================



@app.route('/')



def index():
    from flask import make_response
    resp = make_response(render_template_string(HTML_TEMPLATE, operadores=OPERADORES, app_version=APP_VERSION))
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    resp.headers['Pragma'] = 'no-cache'
    resp.headers['Expires'] = '0'
    return resp



@app.route('/upload', methods=['POST'])



def upload():



    files = request.files.getlist('files')



    if not files:



        # Fallback pra campo 'file' (compatibilidade)



        f = request.files.get('file')



        if f and f.filename:



            files = [f]



    



    if not files or all(not f.filename for f in files):



        return jsonify({'error': 'Nenhum arquivo enviado'})



    sid = datetime.now().strftime('%Y%m%d_%H%M%S')



    



    # Salvar e parsear todos os arquivos



    all_records = []



    file_names = []



    for i, f in enumerate(files):



        if not f.filename:



            continue



        fpath = os.path.join(UPLOAD_DIR, f'{sid}_{i}.csv')



        f.save(fpath)



        file_names.append(f.filename)



        recs = parse_agenda(fpath)



        if recs:



            all_records.extend(recs)



    



    # Salvar path do primeiro arquivo (pra compatibilidade) e lista de todos



    main_path = os.path.join(UPLOAD_DIR, f'{sid}_merged.csv')



    



    records = all_records



    if not records:



        return jsonify({'error': 'Nenhum registro encontrado nos arquivos'})



    raiz_map = load_raizes()



    empresas = analyze_records(records, raiz_map)



    emp_list = []



    for name in sorted(empresas.keys(), key=lambda n: empresas[n]['valor'], reverse=True):



        e = empresas[name]



        emp_list.append({



            'nome': name,



            'cnpjs': len(e['cnpjs']),



            'urs': e['urs'],



            'valor': e['valor'],



            'adquirentes': sorted(e['adquirentes']),



            'arranjos': sorted(e['arranjos'])



        })



    # Salvar paths dos arquivos originais (nao mergear - manter formato original)



    saved_paths = []



    for i, f_item in enumerate(files):



        if not f_item.filename:



            continue



        fpath_i = os.path.join(UPLOAD_DIR, f'{sid}_{i}.csv')



        # Arquivo ja foi salvo antes, so guardar o path



        if os.path.exists(fpath_i):



            saved_paths.append(fpath_i)



    



    # Save session com lista de arquivos



    session_data = {'files': saved_paths, 'session_id': sid, 'source_files': file_names, 'total_records': len(records)}



    with open(os.path.join(UPLOAD_DIR, f'{sid}.json'), 'w') as jf:



        json.dump(session_data, jf)



    # Detectar raizes desconhecidas (otimizado - single pass)



    all_raizes_in_file = set(r.get('raiz', '') for r in records)



    unknown_raiz_set = set(r for r in all_raizes_in_file if r and r not in raiz_map)



    



    # Tentar buscar nomes automaticamente no Salesforce



    if unknown_raiz_set:



        # Montar mapa de CNPJs reais por raiz para busca mais precisa
        _cnpjs_pr = {}
        for _r2 in records:
            _raiz2 = _r2.get('raiz', '')
            _cnpj2 = _r2.get('cnpj', '')
            if _raiz2 in unknown_raiz_set and _cnpj2:
                _cnpjs_pr.setdefault(_raiz2, []).append(_cnpj2)
        sf_names = fetch_empresa_names(unknown_raiz_set, _cnpjs_pr)



        if sf_names:



            # Salvar as novas raizes automaticamente



            for raiz, nome in sf_names.items():



                raiz_map[raiz] = nome



            with open(RAIZES_PATH, 'w', encoding='utf-8') as f:



                json.dump(raiz_map, f, indent=2, ensure_ascii=False)



            # Recalcular desconhecidas



            unknown_raiz_set = set(r for r in all_raizes_in_file if r and r not in raiz_map)



            # Refazer analise com novos nomes



            empresas = analyze_records(records, raiz_map)



            emp_list = []



            for name in sorted(empresas.keys(), key=lambda n: empresas[n]['valor'], reverse=True):



                e = empresas[name]



                emp_list.append({



                    'nome': name,



                    'cnpjs': len(e['cnpjs']),



                    'urs': e['urs'],



                    'valor': e['valor'],



                    'adquirentes': sorted(e['adquirentes']),



                    'arranjos': sorted(e['arranjos'])



                })



    unknown_raizes = {}



    if unknown_raiz_set:



        raiz_stats = {}



        for r in records:



            raiz = r.get('raiz', '')



            if raiz in unknown_raiz_set:



                if raiz not in raiz_stats:



                    raiz_stats[raiz] = {'cnpjs': set(), 'urs': 0}



                raiz_stats[raiz]['cnpjs'].add(r.get('cnpj', ''))



                raiz_stats[raiz]['urs'] += 1



        for raiz in sorted(raiz_stats.keys()):



            s = raiz_stats[raiz]



            unknown_raizes[raiz] = {



                'cnpjs': len(s['cnpjs']),



                'urs': s['urs'],



                'example': sorted(s['cnpjs'])[0] if s['cnpjs'] else ''



            }



    return jsonify({



        'session_id': sid,



        'total_urs': len(records),



        'total_empresas': len(empresas),



        'empresas': emp_list,



        'unknown_raizes': unknown_raizes,



        'source_files': file_names



    })



@app.route('/generate', methods=['POST'])



def generate():



    data = request.json

    _t0_gen = time.time()  # telemetria: marca início



    sid = data['session_id']



    taxa_pct = data['taxa']



    taxa_map = data.get('taxa_map') or {}   # {empresa: taxa_pct} — opcional



    di_periodo = data.get('di_periodo', 0.1465)



    email = data['operator_email']



    with open(os.path.join(UPLOAD_DIR, f'{sid}.json')) as jf:



        sess = json.load(jf)



    # Parsear todos os arquivos da sessao



    file_paths = sess.get('files', [sess['file']] if 'file' in sess else [])



    records = []



    for fp in file_paths:



        if os.path.exists(fp):



            records.extend(parse_agenda(fp))



    raiz_map = load_raizes()



    taxa_nominal = taxa_pct / 100.0



    # Group by empresa



    raiz_to_emp = {}



    emp_raizes = defaultdict(list)



    for raiz, nome in raiz_map.items():



        nu = nome.upper().strip()



        if nu not in emp_raizes:



            emp_raizes[nu] = {'nome': nome, 'raizes': []}



        emp_raizes[nu]['raizes'].append(raiz)



    for nu, info in emp_raizes.items():



        for raiz in info['raizes']:



            raiz_to_emp[raiz] = info['nome']



    for r in records:



        raiz = r.get('raiz', '')



        if raiz not in raiz_to_emp:



            raiz_to_emp[raiz] = f'RAIZ_{raiz}'



    empresa_records = {}



    for r in records:



        emp = raiz_to_emp.get(r.get('raiz', ''), 'OUTROS')



        if emp not in empresa_records:



            empresa_records[emp] = []



        empresa_records[emp].append(r)



    seller_map = {}



    seller_error = None



    out_dir = os.path.join(OUTPUT_DIR, sid)



    # Limpar pasta anterior pra nao misturar arquivos de gerações diferentes



    if os.path.exists(out_dir):



        shutil.rmtree(out_dir, ignore_errors=True)



    os.makedirs(out_dir, exist_ok=True)



    gen_cotacao = data.get('gen_cotacao', True)



    gen_selecao = data.get('gen_selecao', True)



    # Salvar opcoes pra reuso no fetch_sellers (write atomico evita arquivo vazio)



    opts_path = os.path.join(UPLOAD_DIR, f'{sid}_opts.json')



    try:



        _opts_data = json.dumps({'gen_cotacao': gen_cotacao, 'gen_selecao': gen_selecao, 'only_eligible': bool(locals().get('only_eligible', True)), 'taxa': taxa_pct, 'di_periodo': di_periodo, 'email': email})



        with open(opts_path, 'w') as f:



            f.write(_opts_data)



    except Exception as _e:



        import logging as _log



        _log.warning(f'Nao foi possivel salvar opts.json: {_e}')



    result_empresas = []



    total_missing = 0



    for emp_nome, emp_recs in sorted(empresa_records.items()):



        safe = ''.join(c for c in emp_nome.replace(' ', '_').replace('/', '-') if c.isalnum() or c in '_-')



        emp_dir = os.path.join(out_dir, safe)



        os.makedirs(emp_dir, exist_ok=True)



        # /generate: sem elegibilidade — gera tudo com emp_recs



        if gen_cotacao:



            cot_path = os.path.join(emp_dir, f'Cotacao_Elegiveis_{safe}.xlsx')



            taxa_nominal_emp = (taxa_map.get(emp_nome, taxa_pct) / 100.0) if taxa_map else taxa_nominal



            generate_cotacao(emp_recs, emp_nome, taxa_nominal_emp, di_periodo, seller_map, cot_path)



        if gen_selecao:



            sel_path = os.path.join(emp_dir, f'Selecao_URs_{safe}.csv')



            generate_selecao(emp_recs, taxa_pct, email, seller_map, sel_path)



        result_empresas.append({



            'nome': emp_nome,



            'safe_name': safe,



            'urs': len(emp_recs),



            'valor': sum(r.get('disponivel', 0) for r in emp_recs),



            'valor_operavel': sum(r.get('disponivel', 0) for r in emp_recs),



            'taxa': round(taxa_map.get(emp_nome, taxa_pct) if taxa_map else taxa_pct, 4)



        })



    # Salvar no histórico



    import socket



    hostname = socket.gethostname()



    operador_sel = email.split('@')[0] if email else 'desconhecido'



    save_history_entry({



        'timestamp': datetime.now().strftime('%d/%m/%Y %H:%M'),



        'operador': email,



        'usuario_maquina': hostname,



        'tipo': 'Completa',



        'empresas': [e['nome'] for e in result_empresas],



        'total_urs': sum(e['urs'] for e in result_empresas),



        'total_valor': sum(e['valor'] for e in result_empresas),



        'taxa': taxa_pct,



        'di_periodo': di_periodo,



        'session_id': sid



    })

    # Telemetria: registra 1 evento por empresa gerada
    if _TELEMETRIA_OK:
        _dur_gen = round(time.time() - _t0_gen, 1)
        for _e in result_empresas:
            _usage_log.registrar(
                operador=email or '',
                evento='generate',
                empresa=_e.get('nome', ''),
                urs=_e.get('urs', 0),
                valor_bruto=_e.get('valor', 0.0),
                valor_operavel=_e.get('valor_operavel', 0.0),
                taxa=_e.get('taxa', 0.0),
                duracao_s=_dur_gen,
                status='ok',
                session_id=sid,
            )

    return jsonify({



        'empresas': result_empresas,



        'seller_ids_found': len(seller_map) // 2,



        'seller_ids_missing': total_missing,



        'seller_error': seller_error



    })



@app.route('/fetch_sellers', methods=['POST'])



def fetch_sellers_route():



    """Busca seller IDs no Databricks e regera os arquivos."""



    data = request.json



    sid = data['session_id']



    taxa_pct  = data['taxa']



    taxa_map  = data.get('taxa_map') or {}   # {empresa: taxa_pct}



    di_periodo = data.get('di_periodo', 0.1465)



    email = data['operator_email']



    with open(os.path.join(UPLOAD_DIR, f'{sid}.json')) as jf:



        sess = json.load(jf)



    # Parsear todos os arquivos da sessao



    file_paths = sess.get('files', [sess['file']] if 'file' in sess else [])



    records = []

    # Cache de records parseados (evita re-parsear a cada chamada)
    _rc_path = os.path.join(UPLOAD_DIR, f'{sid}_records_cache.json')
    _rc_valid = False
    if os.path.exists(_rc_path) and file_paths:
        try:
            _rc_mtime = os.path.getmtime(_rc_path)
            _fp_mtime = max((os.path.getmtime(fp) for fp in file_paths if os.path.exists(fp)), default=0)
            if _rc_mtime > _fp_mtime:
                with open(_rc_path, 'r', encoding='utf-8') as _rcf:
                    records = json.load(_rcf)
                _rc_valid = True
        except Exception:
            pass

    if not _rc_valid:

        for fp in file_paths:



            if os.path.exists(fp):



                records.extend(parse_agenda(fp))

        # Salvar cache
        try:
            with open(_rc_path, 'w', encoding='utf-8') as _rcf:
                json.dump(records, _rcf, ensure_ascii=False)
        except Exception:
            pass



    raiz_map = load_raizes()



    taxa_nominal = taxa_pct / 100.0



    all_raizes = set(r.get('raiz', '') for r in records)



    # Verificar cache de sellers (evita rebuscar no Databricks se ja foi feito)



    sellers_cache_path = os.path.join(UPLOAD_DIR, f'{sid}_sellers.json')



    prefetch_cache_path = os.path.join(UPLOAD_DIR, f'{sid}_prefetch.json')



    seller_map_raw = None



    seller_error = None



    # Usar cache se disponivel (prefetch ja terminou ou geracao anterior)



    for cache_path in [sellers_cache_path, prefetch_cache_path]:



        if os.path.exists(cache_path):



            try:



                with open(cache_path, 'r') as f:



                    cached = json.load(f)



                if cached.get('status') == 'done' and cached.get('seller_map'):



                    seller_map_raw = cached['seller_map']



                    seller_error = cached.get('error')



                    break



                elif cached.get('status') == 'running':



                    # Prefetch ainda em andamento &#8212; aguardar ate 90s



                    for _ in range(30):



                        time.sleep(3)



                        with open(cache_path, 'r') as f:



                            cached = json.load(f)



                        if cached.get('status') == 'done':



                            seller_map_raw = cached.get('seller_map', {})



                            seller_error = cached.get('error')



                            break



                    break



            except Exception:



                pass



    # Se nao tem cache, buscar agora (fallback)



    if seller_map_raw is None:



        seller_map_raw, seller_error = fetch_seller_ids(all_raizes)



    # Normalizar



    normalized_sm = {}



    for k, v in (seller_map_raw or {}).items():



        normalized_sm[normalize_cnpj(k)] = str(v)



        normalized_sm[k] = str(v)



    seller_map = normalized_sm



    # Salvar cache final



    with open(sellers_cache_path, 'w', encoding='utf-8') as f:



        json.dump({'status': 'done', 'seller_map': seller_map, 'error': seller_error}, f)



    # Carregar opcoes de geracao salvas



    opts_path = os.path.join(UPLOAD_DIR, f'{sid}_opts.json')



    gen_cotacao = True



    gen_selecao = True



    # only_eligible: request tem prioridade sobre opts salvos



    # Se o operador confirma no painel de inelegiveis, JS envia only_eligible=false



    _oe_from_request = 'only_eligible' in data

    # Quando operador confirma elegibilidade, forcar only_eligible=False

    if data.get('eligibility_confirmed', False):

        only_eligible = False

        _oe_from_request = True





    only_eligible = data.get('only_eligible', True)  # False = gerar tb Cotacao_COMPLETO com inelegiveis em vermelho



    if os.path.exists(opts_path) and os.path.getsize(opts_path) > 0:



        try:



            with open(opts_path, 'r') as f:



                opts = json.load(f)



            gen_cotacao = opts.get('gen_cotacao', True)



            gen_selecao = opts.get('gen_selecao', True)



            # So usar opts se request nao enviou only_eligible explicitamente



            if not _oe_from_request:



                only_eligible = opts.get('only_eligible', only_eligible)



        except Exception:



            pass  # opts corrompido ? usa defaults acima



    # Regerar arquivos com seller IDs



    raiz_to_emp = {}



    emp_raizes = defaultdict(list)



    for raiz, nome in raiz_map.items():



        nu = nome.upper().strip()



        if nu not in emp_raizes:



            emp_raizes[nu] = {'nome': nome, 'raizes': []}



        emp_raizes[nu]['raizes'].append(raiz)



    for nu, info in emp_raizes.items():



        for raiz in info['raizes']:



            raiz_to_emp[raiz] = info['nome']



    for r in records:



        raiz = r.get('raiz', '')



        if raiz not in raiz_to_emp:



            raiz_to_emp[raiz] = f'RAIZ_{raiz}'



    empresa_records = {}



    for r in records:



        emp = raiz_to_emp.get(r.get('raiz', ''), 'OUTROS')



        if emp not in empresa_records:



            empresa_records[emp] = []



        empresa_records[emp].append(r)



    out_dir = os.path.join(OUTPUT_DIR, sid)



    # Garantir que out_dir existe antes de qualquer gravacao de arquivo



    os.makedirs(out_dir, exist_ok=True)



    all_cnpjs = set(r.get('cnpj', '') for r in records)



    missing_count = sum(1 for c in all_cnpjs if c not in seller_map and c.lstrip('0') not in seller_map)



    # Gerar seller_ids_faltantes



    if missing_count > 0:



        missing_dict = {c: "" for c in sorted(all_cnpjs) if c not in seller_map and c.lstrip('0') not in seller_map}



        with open(os.path.join(out_dir, 'seller_ids_faltantes.json'), 'w', encoding='utf-8') as f:



            json.dump(missing_dict, f, indent=2)



    # --- ELEGIBILIDADE ---



    # Sellers forçados pelo operador (inelegiveis mas aceitos so na cotacao)



    force_cotacao_only = set(data.get('force_cotacao_only', []))



    eligibility_confirmed = bool(data.get('eligibility_confirmed', False))



    # Verificar elegibilidade em paralelo com o processamento dos records



    # (usa cache de eligibility se disponivel)



    eligibility_cache_path = os.path.join(UPLOAD_DIR, f'{sid}_eligibility.json')



    eligibility = {}



    unique_seller_ids = list(set(v for v in seller_map.values() if v))



    if os.path.exists(eligibility_cache_path):



        try:



            with open(eligibility_cache_path, 'r') as f:



                eligibility = json.load(f)



        except Exception:



            eligibility = {}



    eligibility_fetch_failed = False



    if not eligibility and unique_seller_ids:



        import threading as _threading



        eligibility_result = {}



        eligibility_done = _threading.Event()



        def _fetch_elig():



            try:



                eligibility_result.update(fetch_eligibility(unique_seller_ids))



            except Exception:



                pass



            finally:



                eligibility_done.set()



        t = _threading.Thread(target=_fetch_elig, daemon=True)



        t.start()



        eligibility_done.wait(timeout=120)



        eligibility = eligibility_result



        if not eligibility:



            # Databricks falhou ou nao retornou dados - nao salvar cache vazio



            eligibility_fetch_failed = True



        else:



            # Salvar cache apenas se tem dados reais



            try:



                with open(eligibility_cache_path, 'w') as f:



                    json.dump(eligibility, f)



            except Exception:



                pass



    # Detectar se tabela de elegibilidade esta vazia/indisponivel



    # Se >90% dos sellers sao 'absent', a tabela provavelmente esta em manutencao



    if eligibility and unique_seller_ids:



        absent_count = sum(1 for v in eligibility.values() if v.get('absent', False))



        absent_ratio = absent_count / len(unique_seller_ids)



        if absent_ratio > 0.90:



            # Tabela vazia ou em manutencao - nao bloquear operacao



            eligibility = {}  # tratar como se nao tivesse verificado



            eligibility_fetch_failed = True



            seller_error = (seller_error or '') + (



                f' | Tabela de elegibilidade indisponivel ({absent_count}/{len(unique_seller_ids)} sellers ausentes).'



                ' Verificacao ignorada para nao bloquear a operacao.'



            )



    # Mapear seller_id -> cnpjs para calcular URs/valor dos inelegiveis



    sid_to_cnpjs = {}



    for cnpj, s_id in seller_map.items():



        if len(cnpj) == 14:  # so cnpjs normalizados



            if s_id not in sid_to_cnpjs:



                sid_to_cnpjs[s_id] = []



            sid_to_cnpjs[s_id].append(cnpj)



    # Salvar cache de elegibilidade para generate_custom reutilizar



    # NOME DIFERENTE do cache nativo ({sid}_eligibility.json) para nao conflitar



    try:



        _elig_cache = {'eligibility': eligibility, 'sid_to_cnpjs': sid_to_cnpjs}



        _elig_cache_path = os.path.join(UPLOAD_DIR, f'{sid}_elig_custom.json')



        with open(_elig_cache_path, 'w', encoding='utf-8') as _ef:



            json.dump(_elig_cache, _ef, ensure_ascii=False)



    except Exception:



        pass



    # Identificar inelegiveis que nao foram forcados pelo operador



    ineligible_sellers = []



    ineligible_cnpjs = set()



    for s_id, elig in eligibility.items():



        if not elig.get('eligible', False) and s_id not in force_cotacao_only:



            cnpjs_do_sid = sid_to_cnpjs.get(s_id, [])



            # Calcular URs e valor



            urs_count = sum(1 for r in records if seller_map.get(r.get('cnpj', ''), '') == s_id)



            valor_total = sum(r.get('disponivel', 0) for r in records if seller_map.get(r.get('cnpj', ''), '') == s_id)



            # Descobrir empresa do seller



            empresa_do_sid = 'Desconhecida'



            for cnpj_sid in cnpjs_do_sid:



                raiz_sid = cnpj_sid[:8] if len(cnpj_sid) >= 8 else ''



                if raiz_sid in raiz_to_emp:



                    empresa_do_sid = raiz_to_emp[raiz_sid]



                    break



            ineligible_sellers.append({



                'seller_id': s_id,



                'empresa': empresa_do_sid,



                'cnpjs': cnpjs_do_sid,



                'urs': urs_count,



                'valor': valor_total,



                'absent': elig.get('absent', False)



            })



            ineligible_cnpjs.update(cnpjs_do_sid)





    # Populat ineligible_cnpjs para TODOS os inelegiveis (incluindo cotacao_only)
    for _s_id2, _elig2 in eligibility.items():
        if not _elig2.get('eligible', False):
            ineligible_cnpjs.update(sid_to_cnpjs.get(_s_id2, []))    # Se ha inelegiveis novos (nao forcados) e operador ainda nao confirmou, pedir confirmacao



    if ineligible_sellers and not eligibility_confirmed:



        return jsonify({



            'requires_eligibility_confirmation': True,



            'ineligible_sellers': ineligible_sellers,



            'seller_ids_found': len(seller_map) // 2,



            'seller_ids_missing': missing_count,



            'seller_error': seller_error



        })



    # Quando elegibilidade falhou: mostrar todos os sellers como 'nao verificado'



    # O operador pode confirmar e gerar mesmo assim



    if eligibility_fetch_failed and not eligibility_confirmed and not ineligible_sellers:



        for s_id, cnpjs_list in sid_to_cnpjs.items():



            if s_id in force_cotacao_only:



                continue



            urs_count = sum(1 for r in records if seller_map.get(r.get('cnpj', ''), '') == s_id)



            valor_total = sum(r.get('disponivel', 0) for r in records if seller_map.get(r.get('cnpj', ''), '') == s_id)



            empresa_do_sid = 'Desconhecida'



            for cnpj_sid in cnpjs_list:



                raiz_sid = cnpj_sid[:8] if len(cnpj_sid) >= 8 else ''



                if raiz_sid in raiz_to_emp:



                    empresa_do_sid = raiz_to_emp[raiz_sid]



                    break



            ineligible_sellers.append({



                'seller_id': s_id,



                'empresa': empresa_do_sid,



                'cnpjs': cnpjs_list,



                'urs': urs_count,



                'valor': valor_total,



                'absent': False,



                'unverified': True



            })



        if ineligible_sellers:



            return jsonify({



                'requires_eligibility_confirmation': True,



                'ineligible_sellers': ineligible_sellers,



                'seller_ids_found': len(seller_map) // 2,



                'seller_ids_missing': missing_count,



                'seller_error': seller_error



            })



    # Avisar se verificacao de elegibilidade nao foi possivel



    if eligibility_fetch_failed and not eligibility_confirmed:



        seller_error = (seller_error or '') + ' | Verificacao de elegibilidade indisponivel (Databricks desconectado) - sellers nao foram verificados.'



    # CNPJs que vao apenas para cotacao (forcados pelo operador)



    cotacao_only_cnpjs = set()



    for s_id in force_cotacao_only:



        cotacao_only_cnpjs.update(sid_to_cnpjs.get(s_id, []))



    # --- FIM ELEGIBILIDADE ---



    # Garantir que out_dir existe (pode ter sido limpo entre chamadas)



    os.makedirs(out_dir, exist_ok=True)



    def _gerar_empresa(args):



        emp_nome, emp_recs = args



        safe = ''.join(c for c in emp_nome.replace(' ', '_').replace('/', '-') if c.isalnum() or c in '_-')



        emp_dir = os.path.join(out_dir, safe)



        os.makedirs(emp_dir, exist_ok=True)



        # CNPJs inelegiveis desta empresa



        emp_inelig_cnpjs  = set()

        emp_missing_cnpjs = set()
        emp_cotacao_only_cnpjs = set()



        for r in emp_recs:



            c = r.get('cnpj', '')



            if c in ineligible_cnpjs or c in cotacao_only_cnpjs:



                emp_inelig_cnpjs.add(c)

            if c in cotacao_only_cnpjs:

                emp_cotacao_only_cnpjs.add(c)



            if not seller_map.get(c, ''):



                emp_missing_cnpjs.add(c)



        # Filtro duplo: via ineligible_cnpjs (pre-computado) + via eligibility direto (defesa)



        def _is_eligible_cnpj(cnpj):



            if cnpj in ineligible_cnpjs or cnpj in cotacao_only_cnpjs:



                return False



            s_id = seller_map.get(cnpj, '')

            # Sem seller_id: nao podemos operar — tratar como inelegivel
            if not s_id:

                return False

            if eligibility:

                return eligibility.get(str(s_id), {}).get('eligible', True)

            return True



        _recs_cotacao = [r for r in emp_recs if _is_eligible_cnpj(r.get('cnpj', ''))]



        _recs_selecao = [r for r in _recs_cotacao



                        if seller_map.get(r.get('cnpj', ''), seller_map.get(r.get('cnpj_original', ''), ''))]  # so URs com seller_id



        if gen_cotacao:



            cot_path = os.path.join(emp_dir, f'Cotacao_Elegiveis_{safe}.xlsx')



            taxa_nominal_emp = (taxa_map.get(emp_nome, taxa_pct) / 100.0) if taxa_map else taxa_nominal



            generate_cotacao(_recs_cotacao, emp_nome, taxa_nominal_emp, di_periodo, seller_map, cot_path,

                            missing_cnpjs={r.get('cnpj','') for r in _recs_cotacao if not seller_map.get(r.get('cnpj',''),'')})



            if emp_inelig_cnpjs or emp_missing_cnpjs or emp_cotacao_only_cnpjs:

                cot_completo_path = os.path.join(emp_dir, f'Cotacao_COMPLETO_{safe}.xlsx')

                # Gerar COMPLETO em background para nao bloquear o retorno ao usuario
                def _gen_completo(_recs=list(emp_recs), _nome=str(emp_nome),
                                  _taxa=float(taxa_nominal_emp), _di=float(di_periodo),
                                  _sm=dict(seller_map), _path=str(cot_completo_path),
                                  _inelig=frozenset(emp_inelig_cnpjs | emp_cotacao_only_cnpjs),
                                  _miss=frozenset(emp_missing_cnpjs)):
                    try:
                        generate_cotacao(_recs, _nome, _taxa, _di, _sm, _path,
                                         ineligible_cnpjs=set(_inelig),
                                         missing_cnpjs=set(_miss))
                    except Exception as _e:
                        import logging as _lg, traceback as _tb; _lg.error(f'COMPLETO {_nome}: {_e}\n{_tb.format_exc()}')

                import threading as _thr
                _thr.Thread(target=_gen_completo, daemon=True).start()



        if gen_selecao:



            sel_path = os.path.join(emp_dir, f'Selecao_URs_{safe}.csv')



            taxa_pct_emp = taxa_map.get(emp_nome, taxa_pct) if taxa_map else taxa_pct



            generate_selecao(_recs_selecao, taxa_pct_emp, email, seller_map, sel_path)



        if emp_inelig_cnpjs:



            inelig_csv_path = os.path.join(emp_dir, f'Inelegiveis_{safe}.csv')



            generate_inelegiveis_csv(emp_inelig_cnpjs, seller_map, sid_to_cnpjs, emp_recs,



                                     raiz_to_emp, inelig_csv_path)



        return emp_nome, safe



    from concurrent.futures import ThreadPoolExecutor, as_completed as _as_completed



    _n_workers = min(2, max(1, len(empresa_records)))  # 2 workers: equilibrio entre paralelismo e conten??o GIL do openpyxl



    with ThreadPoolExecutor(max_workers=_n_workers) as _pool:



        _futures = {_pool.submit(_gerar_empresa, item): item[0]



                    for item in sorted(empresa_records.items())}



        for _fut in _as_completed(_futures):



            try:



                _fut.result()



            except Exception as _e:



                import logging as _log



                _log.error(f'Erro gerando empresa {_futures[_fut]}: {_e}')



    # Montar lista de empresas pra o frontend atualizar a tabela



    raiz_to_emp_fs = {}



    for raiz, nome in raiz_map.items():



        raiz_to_emp_fs[raiz] = nome



    for r in records:



        raiz = r.get('raiz', '')



        if raiz not in raiz_to_emp_fs:



            raiz_to_emp_fs[raiz] = f'RAIZ_{raiz}'



    # emp_result: apenas URs de sellers elegiveis com seller_id



    # (mesmo conjunto da Cotacao Elegiveis e Selecao de URs)



    emp_result = {}



    for r in records:



        cnpj = r.get('cnpj', '')



        # Excluir inelegiveis e cotacao_only



        if cnpj in ineligible_cnpjs or cnpj in cotacao_only_cnpjs:



            continue



        # Excluir sem seller_id (nao entram na selecao de URs)



        if not seller_map.get(cnpj, seller_map.get(r.get('cnpj_original',''), '')):



            continue



        emp = raiz_to_emp_fs.get(r.get('raiz', ''), 'OUTROS')



        if emp not in emp_result:



            emp_result[emp] = {'urs': 0, 'valor': 0.0}



        emp_result[emp]['urs'] += 1



        emp_result[emp]['valor'] += r.get('disponivel', 0)



    empresas_list = []



    for nome, stats in sorted(emp_result.items()):



        safe = ''.join(c for c in nome.replace(' ', '_').replace('/', '-') if c.isalnum() or c in '_-')



        taxa_emp = round(taxa_map.get(nome, taxa_pct) if taxa_map else taxa_pct, 4)



        # valor_bruto = todas as URs da empresa (inclusive inelegiveis)
        _valor_bruto = sum(r.get('disponivel',0) for r in records if raiz_to_emp_fs.get(r.get('raiz',''),'') == nome)
        _urs_bruto   = sum(1 for r in records if raiz_to_emp_fs.get(r.get('raiz',''),'') == nome)
        empresas_list.append({
            'nome': nome, 'safe_name': safe,
            'urs': _urs_bruto,
            'valor': _valor_bruto,
            'valor_operavel': stats['valor'],
            'urs_operavel': stats['urs'],
            'taxa': taxa_emp
        })



    # Montar lista de sellers definitivamente excluidos (inelegiveis sem cotacao)



    ineligible_excluded = []



    for s_id, elig in eligibility.items():



        if not elig.get('eligible', False) and s_id not in force_cotacao_only:



            cnpjs_do_sid = sid_to_cnpjs.get(s_id, [])



            empresa_do_sid = 'Desconhecida'



            for cnpj_sid in cnpjs_do_sid:



                raiz_sid = cnpj_sid[:8] if len(cnpj_sid) >= 8 else ''



                if raiz_sid in raiz_to_emp:



                    empresa_do_sid = raiz_to_emp[raiz_sid]



                    break



            ineligible_excluded.append({'seller_id': s_id, 'empresa': empresa_do_sid})



    # Montar seller_summary para exibicao na tela



    # Defensive: se eligibility foi lido do cache no formato errado, extrair sub-objeto



    _elig_map = eligibility if 'eligibility' not in eligibility else eligibility.get('eligibility', {})



    seller_summary = []



    for _sid, _elig in sorted(_elig_map.items()):



        _cnpjs = sid_to_cnpjs.get(_sid, [])



        _urs = sum(1 for r in records if r.get('cnpj','') in _cnpjs)



        _valor = sum(r.get('disponivel', 0) for r in records if r.get('cnpj','') in _cnpjs)



        _emp = 'Desconhecida'



        for _c in _cnpjs:



            _r8 = _c[:8] if len(_c) >= 8 else ''



            if _r8 in raiz_to_emp:



                _emp = raiz_to_emp[_r8]



                break



        seller_summary.append({



            'seller_id': _sid,



            'empresa': _emp,



            'cnpjs': _cnpjs,



            'eligible': _elig.get('eligible', False) if isinstance(_elig, dict) else False,



            'urs': _urs,



            'valor': round(_valor, 2)



        })



    # Mapa empresa -> sellers inelegiveis para painel UX



    inelig_by_empresa = {}



    for _s in seller_summary:



        if not _s['eligible']:



            _emp = _s['empresa']



            if _emp not in inelig_by_empresa:



                inelig_by_empresa[_emp] = []



            inelig_by_empresa[_emp].append({'seller_id': _s['seller_id'], 'cnpjs': _s['cnpjs'], 'urs': _s['urs'], 'valor': _s['valor']})




    # CNPJs sem seller_id por empresa
    _miss_set = {c for c in all_cnpjs
                 if not seller_map.get(c,'') and not seller_map.get(c.lstrip('0'),'')}
    missing_by_empresa = {}
    for _r_ms in records:
        _c_ms = _r_ms.get('cnpj','')
        if _c_ms not in _miss_set: continue
        _emp_ms = raiz_to_emp.get(_r_ms.get('raiz',''), 'OUTROS')
        if _emp_ms not in missing_by_empresa:
            missing_by_empresa[_emp_ms] = {'count':0,'urs':0,'valor':0.0,'_cnpjs':set()}
        if _c_ms not in missing_by_empresa[_emp_ms]['_cnpjs']:
            missing_by_empresa[_emp_ms]['count'] += 1
            missing_by_empresa[_emp_ms]['_cnpjs'].add(_c_ms)
        missing_by_empresa[_emp_ms]['urs']   += 1
        missing_by_empresa[_emp_ms]['valor'] += _r_ms.get('disponivel',0)
    missing_by_empresa = {
        k: {'count':v['count'],'urs':v['urs'],'valor':v['valor']}
        for k,v in missing_by_empresa.items()
    }

    return jsonify({



        'seller_ids_found': len([v for v in seller_map.values() if v]),



        'seller_ids_missing': missing_count,



        'seller_error': seller_error,



        'empresas': empresas_list,



        'ineligible_excluded': ineligible_excluded,



        'seller_summary': seller_summary,



        'inelig_by_empresa': inelig_by_empresa,

        'missing_by_empresa': missing_by_empresa



    })



@app.route('/confirm_eligibility', methods=['POST'])



def confirm_eligibility():



    """Recebe decisao do operador sobre sellers inelegiveis.



    



    Payload:



        session_id: str



        taxa: float



        di_periodo: float



        operator_email: str



        force_cotacao_only: list[str]  &#8212; seller_ids que o operador quer manter so na cotacao



    



    Repassa para /fetch_sellers com os sellers forcados.



    """



    req = request.json



    req['force_cotacao_only'] = req.get('force_cotacao_only', [])



    # Redirecionar logica para fetch_sellers_route com os dados atualizados



    with app.test_request_context(json=req):



        from flask import request as inner_req



        pass



    # Chamar diretamente a funcao com os dados



    import flask



    with app.test_request_context('/fetch_sellers', method='POST', json=req):



        return fetch_sellers_route()



@app.route('/prefetch_sellers', methods=['POST'])



def prefetch_sellers():



    """Inicia busca de seller IDs em background logo apos o upload.



    Retorna imediatamente &#8212; resultado salvo em cache para uso posterior.



    """



    import threading as _threading



    data = request.json



    sid = data.get('session_id')



    if not sid:



        return jsonify({'status': 'error', 'message': 'session_id obrigatorio'})



    cache_path = os.path.join(UPLOAD_DIR, f'{sid}_prefetch.json')



    # Nao iniciar se ja esta rodando ou concluido



    if os.path.exists(cache_path):



        try:



            with open(cache_path, 'r') as f:



                cached = json.load(f)



            return jsonify({'status': cached.get('status', 'unknown')})



        except Exception:



            pass



    # Marcar como running



    with open(cache_path, 'w') as f:



        json.dump({'status': 'running'}, f)



    def _do_prefetch():



        try:



            # Carregar records da sessao



            sess_path = os.path.join(UPLOAD_DIR, f'{sid}.json')



            if not os.path.exists(sess_path):



                with open(cache_path, 'w') as f:



                    json.dump({'status': 'done', 'seller_map': {}, 'error': 'Sessao nao encontrada'}, f)



                return



            with open(sess_path, 'r') as f:



                sess = json.load(f)



            file_paths = sess.get('files', [])



            records = []



            for fp in file_paths:



                if os.path.exists(fp):



                    records.extend(parse_agenda(fp))



            all_raizes = set(r.get('raiz', '') for r in records)



            seller_map_raw, seller_error = fetch_seller_ids(all_raizes)



            # Normalizar



            normalized = {}



            for k, v in seller_map_raw.items():



                normalized[normalize_cnpj(k)] = str(v)



                normalized[k] = str(v)



            # Salvar resultado



            with open(cache_path, 'w') as f:



                json.dump({'status': 'done', 'seller_map': normalized, 'error': seller_error}, f)



            # Ja iniciar eligibility em paralelo



            unique_sids = list(set(v for v in normalized.values() if v))



            if unique_sids:



                eligibility_cache = os.path.join(UPLOAD_DIR, f'{sid}_eligibility.json')



                if not os.path.exists(eligibility_cache):



                    try:



                        eligibility = fetch_eligibility(unique_sids)



                        with open(eligibility_cache, 'w') as f:



                            json.dump(eligibility, f)



                    except Exception:



                        pass



        except Exception as e:



            with open(cache_path, 'w') as f:



                json.dump({'status': 'done', 'seller_map': {}, 'error': str(e)}, f)



    t = _threading.Thread(target=_do_prefetch, daemon=True)



    t.start()



    return jsonify({'status': 'started'})



@app.route('/prefetch_status', methods=['GET'])



def prefetch_status():



    """Retorna status do prefetch de sellers."""



    sid = request.args.get('session_id')



    if not sid:



        return jsonify({'status': 'error'})



    for cache_name in [f'{sid}_sellers.json', f'{sid}_prefetch.json']:



        cache_path = os.path.join(UPLOAD_DIR, cache_name)



        if os.path.exists(cache_path):



            try:



                with open(cache_path, 'r') as f:



                    cached = json.load(f)



                status = cached.get('status', 'unknown')



                seller_map = cached.get('seller_map', {})



                found = len([v for v in seller_map.values() if v and len(str(v)) > 3]) // 2



                return jsonify({



                    'status': status,



                    'seller_ids_found': found,



                    'error': cached.get('error')



                })



            except Exception:



                pass



    return jsonify({'status': 'not_started'})



@app.route('/add_raizes', methods=['POST'])



def add_raizes():



    """Salva novas raizes informadas pelo operador."""



    data = request.json



    novas = data.get('raizes', {})  # {raiz: nome_empresa}



    if not novas:



        return jsonify({'error': 'Nenhuma raiz informada'})



    



    raiz_map = load_raizes()



    for raiz, nome in novas.items():



        raiz_map[raiz.zfill(8)] = nome.strip()



    



    with open(RAIZES_PATH, 'w', encoding='utf-8') as f:



        json.dump(raiz_map, f, indent=2, ensure_ascii=False)



    



    return jsonify({'status': 'ok', 'total': len(raiz_map)})



@app.route('/get_datas')



def get_datas():



    sid = request.args.get('session_id')



    empresa = request.args.get('empresa')



    with open(os.path.join(UPLOAD_DIR, f'{sid}.json')) as jf:



        sess = json.load(jf)



    # Parsear todos os arquivos da sessao



    file_paths = sess.get('files', [sess['file']] if 'file' in sess else [])



    records = []



    for fp in file_paths:



        if os.path.exists(fp):



            records.extend(parse_agenda(fp))



    raiz_map = load_raizes()



    emp_raizes = set()



    for raiz, nome in raiz_map.items():



        if nome == empresa:



            emp_raizes.add(raiz)



    filtered = [r for r in records if r.get('raiz', '') in emp_raizes]

    # Filtro opcional por CNPJs (para atualizar adquirentes/arranjos por CNPJ)
    cnpj_filter_gd = request.args.get('cnpj', '')
    if cnpj_filter_gd:
        _cnpjs_gd = [c.strip() for c in cnpj_filter_gd.split(',') if c.strip()]
        if _cnpjs_gd:
            filtered = [r for r in filtered if r.get('cnpj', '') in _cnpjs_gd]

    _seller_map_gd = {}
    _sc_p = os.path.join(UPLOAD_DIR, f'{sid}_sellers.json')
    if os.path.exists(_sc_p):
        try:
            with open(_sc_p,'r') as _sf: _seller_map_gd = json.load(_sf).get('seller_map',{})
        except Exception: pass



    # Carregar inelegiveis para calcular valor operavel



    _inelig_cnpjs_gd = set()



    _elig_path_gd = os.path.join(UPLOAD_DIR, f'{sid}_elig_custom.json')



    if os.path.exists(_elig_path_gd):



        try:



            with open(_elig_path_gd, 'r', encoding='utf-8') as _ef:



                _ed = json.load(_ef)



            _elig = _ed.get('eligibility', {})



            _s2c = _ed.get('sid_to_cnpjs', {})



            for _sid2, _e in _elig.items():



                if not _e.get('eligible', False):



                    _inelig_cnpjs_gd.update(_s2c.get(_sid2, []))



        except Exception:



            pass



    datas_info = defaultdict(lambda: {'urs': 0, 'valor': 0.0})



    adquirentes = set()



    arranjos = set()



    for r in filtered:



        ds = r.get('data_liquidacao', '')



        if ds:



            datas_info[ds]['urs'] += 1



            datas_info[ds]['valor'] += r.get('disponivel', 0)



        adq = r.get('adquirente', '').strip()



        if adq:



            adquirentes.add(adq)



        arr = r.get('arranjo', '').strip()



        if arr:



            arranjos.add(arr)



    result_datas = []



    for ds in sorted(datas_info.keys(), key=lambda d: parse_date(d) or datetime.max):



        result_datas.append({



            'data': ds,



            'urs': datas_info[ds]['urs'],



            'valor': datas_info[ds]['valor']



        })



    filtered_elig = [
        r for r in filtered
        if r.get('cnpj','') not in _inelig_cnpjs_gd
        and _seller_map_gd.get(r.get('cnpj',''), _seller_map_gd.get(r.get('cnpj_original',''),''))
    ]



    _cnpj_map = {}
    for _r in filtered:
        _c = _r.get('cnpj', '')
        if not _c: continue
        if _c not in _cnpj_map:
            _cnpj_map[_c] = {'cnpj': _c, 'agenda_total': 0.0, 'agenda_operavel': 0.0, 'urs': 0}
        _cnpj_map[_c]['agenda_total'] += _r.get('disponivel', 0)
        _cnpj_map[_c]['urs'] += 1
        if _c not in _inelig_cnpjs_gd and _seller_map_gd.get(_c, ''):
            _cnpj_map[_c]['agenda_operavel'] += _r.get('disponivel', 0)
    _cnpjs_list = sorted(_cnpj_map.values(), key=lambda x: -x['agenda_total'])

    return jsonify({



        'datas': result_datas,



        'adquirentes': sorted(adquirentes),



        'arranjos': sorted(arranjos),



        'total_urs': len(filtered),



        'total_valor': sum(r.get('disponivel', 0) for r in filtered),



        'total_valor_elegivel': sum(r.get('disponivel', 0) for r in filtered_elig),



        'total_urs_elegivel': len(filtered_elig),



        'has_eligibility': len(_inelig_cnpjs_gd) > 0



,

        'cnpjs': _cnpjs_list
    })



@app.route('/download_inelegiveis/<sid>/<safe_name>')



def download_inelegiveis(sid, safe_name):



    from flask import send_file



    path = os.path.join(OUTPUT_DIR, sid, safe_name, f'Inelegiveis_{safe_name}.csv')



    if not os.path.exists(path):



        return jsonify({'error': 'Arquivo nao encontrado'}), 404



    return send_file(path, as_attachment=True, download_name=f'Inelegiveis_{safe_name}.csv', mimetype='text/csv')



@app.route('/generate_custom', methods=['POST'])



def generate_custom():



    import traceback as _tb, logging as _lg



    data = request.json



    sid = data['session_id']



    empresa = data['empresa']



    valor_alvo = data.get('valor_alvo', 0)



    adquirente = data.get('adquirente', '')



    arranjo = data.get('arranjo', '')



    taxa_pct = data['taxa']



    di_periodo = data.get('di_periodo', 0.1465)



    email = data['operator_email']



    only_eligible = True  # personalizada sempre opera apenas elegiveis



    with open(os.path.join(UPLOAD_DIR, f'{sid}.json')) as jf:



        sess = json.load(jf)



    # Parsear todos os arquivos da sessao



    file_paths = sess.get('files', [sess['file']] if 'file' in sess else [])



    records = []



    for fp in file_paths:



        if os.path.exists(fp):



            records.extend(parse_agenda(fp))



    raiz_map = load_raizes()



    taxa_nominal = taxa_pct / 100.0



    # Filter by empresa



    emp_raizes = set()



    for raiz, nome in raiz_map.items():



        if nome == empresa:



            emp_raizes.add(raiz)



    filtered = [r for r in records if r.get('raiz', '') in emp_raizes]



    # Filtro adquirente: aceita string unica ou multi-valores separados por virgula



    if adquirente:



        adq_list = [a.strip().upper() for a in adquirente.split(',') if a.strip()]



        if adq_list:



            filtered = [r for r in filtered if any(a in r.get('adquirente', '').upper() for a in adq_list)]



    # Filtro arranjo: aceita string unica ou multi-valores separados por virgula



    if arranjo:



        arr_list = [a.strip().upper() for a in arranjo.split(',') if a.strip()]



        if arr_list:



            filtered = [r for r in filtered if r.get('arranjo', '').upper() in arr_list]



    # Filtro por CNPJ
    cnpj_filter = data.get('cnpj', '')
    if cnpj_filter:
        _cnpj_list = [c.strip() for c in cnpj_filter.split(',') if c.strip()]
        if _cnpj_list:
            filtered = [r for r in filtered if r.get('cnpj', '') in _cnpj_list]

    # Filter by datas



    datas_filter = data.get('datas_filter', {})



    if datas_filter and datas_filter.get('mode') != 'todas':



        mode = datas_filter.get('mode', 'todas')



        if mode == 'range':



            de = parse_date(datas_filter.get('de', ''))



            ate = parse_date(datas_filter.get('ate', ''))



            if de and ate:



                filtered = [r for r in filtered if parse_date(r.get('data_liquidacao', ''))



                           and de <= parse_date(r['data_liquidacao']) <= ate]



            elif de:



                filtered = [r for r in filtered if parse_date(r.get('data_liquidacao', ''))



                           and parse_date(r['data_liquidacao']) >= de]



            elif ate:



                filtered = [r for r in filtered if parse_date(r.get('data_liquidacao', ''))



                           and parse_date(r['data_liquidacao']) <= ate]



        elif mode == 'select':



            selected_datas = set(datas_filter.get('datas', []))



            if selected_datas:



                filtered = [r for r in filtered if r.get('data_liquidacao', '') in selected_datas]



    if not filtered:



        return jsonify({'error': 'Nenhuma UR encontrada com esses filtros'})



    out_dir = os.path.join(OUTPUT_DIR, sid, 'custom')



    if os.path.exists(out_dir):



        for f_item in os.listdir(out_dir):



            try:



                os.remove(os.path.join(out_dir, f_item))



            except Exception:



                pass



    os.makedirs(out_dir, exist_ok=True)



    # Carregar seller IDs salvos da geracao anterior



    seller_map = {}



    for cache_name in [f'{sid}_sellers.json', f'{sid}_prefetch.json']:



        sellers_path = os.path.join(UPLOAD_DIR, cache_name)



        if os.path.exists(sellers_path):



            try:



                with open(sellers_path, 'r', encoding='utf-8') as f:



                    cached = json.load(f)



                if isinstance(cached, dict) and cached.get('seller_map'):



                    seller_map = cached['seller_map']



                    break



                elif isinstance(cached, dict) and cached.get('status') == 'done':



                    seller_map = cached.get('seller_map', {})



                    break



            except Exception:



                pass



    safe = ''.join(c for c in empresa.replace(' ', '_') if c.isalnum() or c in '_-')



    suffix = f'_{int(valor_alvo/1e6)}M' if valor_alvo > 0 else ''



    # Carregar inelegiveis ANTES do valor_alvo para nao contar URs que serao excluidas



    custom_inelig_cnpjs = set()



    elig_cache_path = os.path.join(UPLOAD_DIR, f'{sid}_elig_custom.json')



    if os.path.exists(elig_cache_path):



        try:



            with open(elig_cache_path, 'r', encoding='utf-8') as _ef:



                _elig_data = json.load(_ef)



            _elig = _elig_data.get('eligibility', {})



            _sid_to_cnpjs = _elig_data.get('sid_to_cnpjs', {})



            for _sid, _e in _elig.items():



                if not _e.get('eligible', False):



                    custom_inelig_cnpjs.update(_sid_to_cnpjs.get(_sid, []))



        except Exception:



            pass



    # Filtrar inelegiveis antes de aplicar valor_alvo



    inelig_in_filter = {c for c in custom_inelig_cnpjs if any(r.get('cnpj','') == c for r in filtered)} if custom_inelig_cnpjs else set()



    filtered_no_inelig = [
            r for r in filtered
            if r.get('cnpj','') not in custom_inelig_cnpjs
            and seller_map.get(r.get('cnpj',''), seller_map.get(r.get('cnpj_original',''),''))  # excluir sem seller_id
        ] if custom_inelig_cnpjs else filtered



    # Aplicar valor_alvo APENAS nas URs elegiveis



    valor_alvo_aviso = None  # aviso de diferenca para o frontend

    if valor_alvo > 0:

        filtered_no_inelig.sort(key=lambda r: parse_date(r.get('data_liquidacao', '')) or datetime.max)

        selected = []
        acum = 0.0

        for r in filtered_no_inelig:
            val_ur = r.get('disponivel', 0)
            restante = round(valor_alvo - acum, 2)

            if restante <= 0:
                break  # alvo já atingido exatamente

            if acum + val_ur <= valor_alvo:
                # UR cabe inteira — sem alteração
                r = dict(r)
                r.pop('_valor_cedido', None)
                selected.append(r)
                acum += val_ur

            else:
                # UR ultrapassa o alvo — ceder apenas o valor restante necessário
                r = dict(r)
                r['_valor_cedido'] = restante  # cessão parcial da UR
                selected.append(r)
                acum = valor_alvo
                break

        filtered_no_inelig = selected



        # Avisar diferenca entre alvo e o que foi selecionado



        if acum < valor_alvo * 0.95:



            # Agenda elegivel menor que o alvo solicitado



            valor_alvo_aviso = (



                f'Agenda elegível insuficiente: R$ {acum:,.2f} disponivel '



                f'(solicitado R$ {valor_alvo:,.2f}). '



                f'Todas as URs elegíveis foram incluídas.'



            )



        elif acum > valor_alvo * 1.05:



            # Estouro acima do alvo



            valor_alvo_aviso = (



                f'Valor selecionado R$ {acum:,.2f} excede o alvo de '



                f'R$ {valor_alvo:,.2f} em {((acum/valor_alvo)-1)*100:.1f}%.'



            )



    if not filtered_no_inelig:



        return jsonify({'error': 'Nenhuma UR elegivel encontrada com esses filtros'})



    cot_path = os.path.join(out_dir, f'Cotacao_{safe}{suffix}.xlsx')  # variavel de compatibilidade



    # URs com seller_id: base comum para Cotacao Elegiveis e Selecao de URs



    # Garante que os dois arquivos tenham exatamente o mesmo conjunto de URs



    filtered_com_seller = [r for r in filtered_no_inelig



                           if seller_map.get(r.get('cnpj',''), seller_map.get(r.get('cnpj_original',''),''))]



    # Cotacao Elegiveis (arquivo principal) - apenas URs com seller_id



    cot_path = os.path.join(out_dir, f'Cotacao_Elegiveis_{safe}{suffix}.xlsx')



    generate_cotacao(filtered_com_seller, empresa, taxa_nominal, di_periodo, seller_map, cot_path,

                    missing_cnpjs={r.get('cnpj','') for r in filtered_com_seller if not seller_map.get(r.get('cnpj',''),'')})



    # Cotacao COMPLETA com inelegiveis em vermelho (so se only_eligible=False)



    if not only_eligible and inelig_in_filter:



        cot_completo_path = os.path.join(out_dir, f'Cotacao_COMPLETO_{safe}{suffix}.xlsx')



        generate_cotacao(filtered, empresa, taxa_nominal, di_periodo, seller_map,



                         cot_completo_path, ineligible_cnpjs=inelig_in_filter,

                         missing_cnpjs={r.get('cnpj','') for r in filtered if not seller_map.get(r.get('cnpj',''),'')})



    sel_path = os.path.join(out_dir, f'Selecao_URs_{safe}{suffix}.csv')



    generate_selecao(filtered_com_seller, taxa_pct, email, seller_map, sel_path)



    # Lista de inelegiveis



    if inelig_in_filter:



        _sid_to_cnpjs_custom = {}



        for _c, _s in seller_map.items():



            if _s:



                _sid_to_cnpjs_custom.setdefault(_s, []).append(_c)



        _raiz_to_emp_custom = {raiz: nome for raiz, nome in load_raizes().items()}



        inelig_csv_path = os.path.join(out_dir, f'Inelegiveis_{safe}{suffix}.csv')



        generate_inelegiveis_csv(inelig_in_filter, seller_map, _sid_to_cnpjs_custom,



                                 filtered_no_inelig, _raiz_to_emp_custom, inelig_csv_path)



    datas = sorted(



        set(r.get('data_liquidacao', '') for r in filtered_no_inelig if r.get('data_liquidacao', '')),



        key=lambda d: parse_date(d) or datetime.max



    )



    periodo_str = f'{datas[0]} a {datas[-1]}' if len(datas) > 1 else (datas[0] if datas else 'N/A')



    import socket



    save_history_entry({



        'timestamp': datetime.now().strftime('%d/%m/%Y %H:%M'),



        'operador': email,



        'usuario_maquina': socket.gethostname(),



        'tipo': f'Personalizada - {empresa}',



        'empresas': [empresa],



        'total_urs': len(filtered_no_inelig),



        'total_valor': sum(r.get('disponivel', 0) for r in filtered_no_inelig),



        'taxa': taxa_pct,



        'di_periodo': di_periodo,



        'filtros': {



            'valor_alvo': valor_alvo,



            'adquirente': adquirente,



            'arranjo': arranjo,



            'periodo': periodo_str



        },



        'session_id': sid



    })



    periodo = periodo_str



    # Valor bruto real: usar _valor_cedido na UR parcial (não o disponivel cheio)
    _valor_bruto = sum(
        r.get('_valor_cedido', r.get('disponivel', 0))
        for r in filtered_no_inelig
    )

    # Identificar UR com cessão parcial (se houver)
    _ur_parcial = None
    for _r in filtered_no_inelig:
        if '_valor_cedido' in _r:
            _ur_parcial = {
                'receivable_id':  _r.get('receivable_id', ''),
                'data_liquidacao': _r.get('data_liquidacao', ''),
                'adquirente':      _r.get('adquirente', ''),
                'arranjo':         _r.get('arranjo', ''),
                'disponivel':      round(_r.get('disponivel', 0), 2),
                'cedido':          round(_r['_valor_cedido'], 2),
                'remanescente':    round(_r.get('disponivel', 0) - _r['_valor_cedido'], 2),
                'disponivel_total': round(sum(r.get('disponivel',0) for r in filtered_no_inelig), 2),
            }
            break



    resp = {



        'urs': len(filtered_no_inelig),



        'valor': _valor_bruto,



        'valor_total':    _valor_bruto,



        'valor_operavel': _valor_bruto,



        'taxa': round(taxa_pct, 4),



        'periodo': periodo



    }



    if valor_alvo_aviso:
        resp['aviso'] = valor_alvo_aviso

    if _ur_parcial:
        resp['ur_parcial'] = _ur_parcial

    return jsonify(resp)



@app.route('/download/<sid>/<empresa>')



def download(sid, empresa):

    emp_dir = os.path.join(OUTPUT_DIR, sid, empresa)

    if not os.path.exists(emp_dir):
        return 'Nao encontrado', 404

    # Cotacao_COMPLETO_ e gerada em thread background ? aguarda ate 15s antes de zipar
    _prefixes = ('Cotacao_COMPLETO_', 'Cotacao_Elegiveis_')
    def _tem_cotacao():
        try:
            return any(f.startswith(_prefixes) and f.endswith('.xlsx')
                       for f in os.listdir(emp_dir))
        except Exception:
            return False
    _w = 0
    while not _tem_cotacao() and _w < 15:
        time.sleep(1); _w += 1

    buf = BytesIO()



    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:



        for root, dirs, files in os.walk(emp_dir):



            for file in files:



                fp = os.path.join(root, file)



                zf.write(fp, file)



    buf.seek(0)



    return send_file(buf, mimetype='application/zip', as_attachment=True,



                     download_name=f'{empresa}.zip')



@app.route('/download_all/<sid>')



def download_all(sid):



    out_dir = os.path.join(OUTPUT_DIR, sid)



    if not os.path.exists(out_dir):



        return 'Nao encontrado', 404



    buf = BytesIO()



    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:



        for root, dirs, files in os.walk(out_dir):



            for file in files:



                fp = os.path.join(root, file)



                arcname = os.path.relpath(fp, out_dir)



                zf.write(fp, arcname)



    buf.seek(0)



    return send_file(buf, mimetype='application/zip', as_attachment=True,



                     download_name=f'AR_{sid}.zip')



# ==============================================================================



# OAUTH ROUTES



# ==============================================================================



@app.route('/api/version')



def api_version():



    return jsonify({'version': APP_VERSION})



@app.route('/api/update/<filename>')



def api_update_file(filename):



    """Serve arquivos atualizados (webapp.py, raizes_conhecidas.json)."""



    allowed = ['webapp.py', 'raizes_conhecidas.json']



    if filename not in allowed:



        return 'Not found', 404



    fpath = os.path.join(BASE_DIR, filename)



    if os.path.exists(fpath):



        return send_file(fpath, as_attachment=True)



    return 'Not found', 404



@app.route('/restart', methods=['POST'])
def restart_app():
    """Reinicia o processo do app para carregar código novo do disco."""
    import threading, time, subprocess

    def _do_restart():
        time.sleep(1)
        try:
            if getattr(sys, 'frozen', False):
                # Rodando como .exe — relançar o executável
                exe = sys.executable
                subprocess.Popen([exe], creationflags=0x00000008)  # DETACHED_PROCESS
            else:
                # Rodando como script Python
                subprocess.Popen([sys.executable] + sys.argv)
            # Encerrar o processo atual após lançar o novo
            os._exit(0)
        except Exception:
            os._exit(0)

    threading.Thread(target=_do_restart, daemon=True).start()
    return jsonify({'status': 'restarting'})


@app.route('/server_status')



def server_status():



    server_url = os.environ.get('HISTORY_SERVER_URL', '').strip()



    connected = False



    if server_url and http_requests:



        try:



            r = http_requests.get(f'{server_url}/api/history', timeout=5, verify=False)



            connected = r.status_code == 200



        except Exception:



            pass



    return jsonify({



        'connected': connected,



        'url': server_url if server_url else None,



        'is_server': not bool(server_url)  # se nao tem URL, ele eh o servidor



    })



@app.route('/check_update')
def check_update():
    # Usa updater.py (GitHub) se disponível, fallback para HISTORY_SERVER_URL legado
    if _UPDATER_OK:
        return jsonify(_updater.get_status())
    server_url = os.environ.get('HISTORY_SERVER_URL', '').strip()
    if not server_url or not http_requests:
        return jsonify({'update_available': False, 'local_version': APP_VERSION})
    try:
        r = http_requests.get(f'{server_url}/api/version', timeout=5, verify=False)
        if r.status_code == 200:
            remote = r.json().get('version', APP_VERSION)
            return jsonify({'update_available': remote != APP_VERSION, 'local_version': APP_VERSION, 'remote_version': remote})
    except Exception:
        pass
    return jsonify({'update_available': False, 'local_version': APP_VERSION})


@app.route('/apply_update', methods=['POST'])
def apply_update():
    # Usa updater.py (GitHub) se disponível
    if _UPDATER_OK:
        status = _updater.get_status()
        if status.get('has_update'):
            _updater.start_download_and_update()
            return jsonify({'status': 'download iniciado'})
        return jsonify({'status': 'nenhuma atualização disponível'})
    server_url = os.environ.get('HISTORY_SERVER_URL', '').strip()



    if not server_url or not http_requests:



        return jsonify({'error': 'Servidor central nao configurado'})



    try:



        for filename in ['webapp.py', 'raizes_conhecidas.json']:



            r = http_requests.get(f'{server_url}/api/update/{filename}', timeout=30, verify=False)



            if r.status_code == 200:



                fpath = os.path.join(BASE_DIR, filename)



                with open(fpath, 'wb') as f:



                    f.write(r.content)



        # Recarregar versao



        return jsonify({'status': 'ok', 'version': 'atualizado'})



    except Exception as e:



        return jsonify({'error': str(e)})



@app.route('/oauth/login')



def oauth_login():



    """Reconecta ao Databricks via CLI (databricks auth token).



    Tenta buscar o token do CLI sem interacao do usuario.



    Se o CLI nao tiver sessao ativa, retorna instrucoes para reconectar manualmente.



    """



    token = get_token_via_cli()



    if token:



        save_oauth_token({'access_token': token, 'token_type': 'Bearer'})



        return """



        <html><body style="font-family:sans-serif;text-align:center;padding:60px;background:#f9fafb;">



        <h1 style="color:#1B5E20;">Conectado ao Databricks!</h1>



        <p style="color:#374151;">Token obtido via Databricks CLI com sucesso.</p>



        <script>setTimeout(function(){ window.location.href='/'; }, 2000);</script>



        <p><a href="/" style="color:#4CAF50;font-size:16px;">Voltar para o SimplificaE</a></p>



        </body></html>



        """



    else:



        # Detectar se o CLI esta instalado



        import subprocess as _sp



        cli_installed = False



        cli_paths = [



            os.path.join(os.environ.get('LOCALAPPDATA', ''),



                'Microsoft', 'WinGet', 'Packages',



                'Databricks.DatabricksCLI_Microsoft.Winget.Source_8wekyb3d8bbwe',



                'databricks.exe'),



            'databricks', 'databricks.exe'



        ]



        found_cli = None



        for _p in cli_paths:



            try:



                r = _sp.run([_p, 'version'], capture_output=True, timeout=5)



                if r.returncode == 0:



                    cli_installed = True



                    found_cli = _p



                    break



            except Exception:



                continue



        if not cli_installed:



            return """



        <html><body style="font-family:sans-serif;padding:40px;background:#f9fafb;max-width:680px;margin:auto;">



        <h2 style="color:#B45309;">Databricks CLI nao instalado</h2>



        <p style="color:#374151;">O Databricks CLI precisa estar instalado para conectar ao Databricks.</p>



        <p style="color:#374151;"><strong>Passo 1:</strong> Abra o <strong>PowerShell</strong> e execute:</p>



        <pre style="background:#1f2937;color:#f9fafb;padding:16px;border-radius:8px;font-size:13px;">winget install --id Databricks.DatabricksCLI -e --source winget</pre>



        <p style="color:#374151;"><strong>Passo 2:</strong> Apos instalar, execute:</p>



        <pre style="background:#1f2937;color:#f9fafb;padding:16px;border-radius:8px;font-size:13px;">databricks auth login --host https://picpay-principal.cloud.databricks.com</pre>



        <p style="color:#6B7280;font-size:13px;">Uma janela do navegador vai abrir para login. Apos concluir, volte aqui e clique em Tentar Novamente.</p>



        <a href="/oauth/login" style="display:inline-block;margin-top:16px;padding:10px 24px;background:#21A366;color:white;border-radius:6px;text-decoration:none;font-weight:600;">Tentar Novamente</a>



        &nbsp;



        <a href="/" style="display:inline-block;margin-top:16px;padding:10px 24px;background:#6B7280;color:white;border-radius:6px;text-decoration:none;">Voltar</a>



        </body></html>



            """



        else:



            cmd = f'& "{found_cli}" auth login --host {DATABRICKS_HOST}'



            return f"""



        <html><body style="font-family:sans-serif;padding:40px;background:#f9fafb;max-width:680px;margin:auto;">



        <h2 style="color:#B45309;">Sessao Databricks expirada</h2>



        <p style="color:#374151;">O token do Databricks expirou. Para reconectar, abra o <strong>PowerShell</strong> e execute:</p>



        <pre style="background:#1f2937;color:#f9fafb;padding:16px;border-radius:8px;font-size:13px;overflow-x:auto;">{cmd}</pre>



        <p style="color:#6B7280;font-size:13px;">Uma janela do navegador vai abrir para login. Apos concluir, clique em Tentar Novamente.</p>



        <a href="/oauth/login" style="display:inline-block;margin-top:16px;padding:10px 24px;background:#21A366;color:white;border-radius:6px;text-decoration:none;font-weight:600;">Tentar Novamente</a>



        &nbsp;



        <a href="/" style="display:inline-block;margin-top:16px;padding:10px 24px;background:#6B7280;color:white;border-radius:6px;text-decoration:none;">Voltar</a>



        </body></html>



            """



@app.route('/oauth/callback')



def oauth_callback():



    """Recebe callback do Databricks e troca code por tokens."""



    code = request.args.get('code')



    state = request.args.get('state')



    error = request.args.get('error')



    if error:



        return f"<h2>Erro OAuth</h2><p>{error}: {request.args.get('error_description', '')}</p><a href='/'>Voltar</a>"



    if state != session.get('oauth_state'):



        return "<h2>Erro</h2><p>State invalido</p><a href='/'>Voltar</a>"



    code_verifier = session.get('oauth_verifier', '')



    try:



        resp = http_requests.post(OAUTH_TOKEN_URL, data={



            'grant_type': 'authorization_code',



            'client_id': OAUTH_CLIENT_ID,



            'code': code,



            'redirect_uri': OAUTH_REDIRECT_URI,



            'code_verifier': code_verifier,



        }, timeout=15)



        if resp.status_code == 200:



            token_data = resp.json()



            save_oauth_token(token_data)



            return """



            <html><body style="font-family:sans-serif;text-align:center;padding:60px;">



            <h1 style="color:#1B5E20;">Conectado ao Databricks!</h1>



            <p>Token OAuth salvo com sucesso. O refresh automático está ativo.</p>



            <p><a href="/" style="color:#4CAF50;font-size:18px;">Voltar para o SimplificaÊ</a></p>



            </body></html>



            """



        else:



            return f"<h2>Erro ao obter token</h2><pre>{resp.text}</pre><a href='/'>Voltar</a>"



    except Exception as e:



        return f"<h2>Erro</h2><p>{e}</p><a href='/'>Voltar</a>"



@app.route('/oauth/status')



def oauth_status():



    """Retorna status da conexao.



    



    ?full=0 (padrao): check leve &#8212; so verifica se token existe localmente. Rapido.



    ?full=1: executa SELECT 1 real no Databricks para confirmar conectividade.



    """



    from flask import request as _req



    full_check = _req.args.get('full', '0') == '1'



    token_data = load_oauth_token()



    has_oauth = bool(token_data and 'access_token' in token_data)



    has_pat = bool(DATABRICKS_TOKEN)



    connected = False



    method = 'none'



    error_detail = ''



    if not full_check:



        # Zero HTTP &#8212; so verifica token local, instantaneo



        if has_oauth:



            connected = True



            method = 'oauth'



        elif has_pat:



            connected = True



            method = 'pat'



        else:



            error_detail = 'Sem token configurado'



    else:



        # Verificacao real: SELECT 1 no Databricks



        token = get_databricks_token()



        if token:



            method = 'oauth' if has_oauth else 'pat'



            if http_requests:



                try:



                    resp = http_requests.post(



                        f'{DATABRICKS_HOST}/api/2.0/sql/statements/',



                        headers={'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'},



                        json={'warehouse_id': DATABRICKS_WAREHOUSE_ID, 'statement': 'SELECT 1', 'wait_timeout': '10s'},



                        timeout=15



                    )



                    if resp.status_code == 200:



                        connected = True



                    elif resp.status_code == 403:



                        error_detail = 'Token expirado ou invalido'



                    else:



                        error_detail = f'Erro {resp.status_code}'



                except Exception:



                    error_detail = 'Sem conexao com Databricks'



        else:



            error_detail = 'Sem token'



    return jsonify({



        'connected': connected,



        'method': method,



        'has_oauth': has_oauth,



        'has_pat': has_pat,



        'error': error_detail,



        'full_check': full_check,



        'saved_at': token_data.get('saved_at', '') if token_data else ''



    })



@app.route('/oauth/disconnect')



def oauth_disconnect():



    """Remove token OAuth salvo."""



    if os.path.exists(OAUTH_TOKEN_FILE):



        os.remove(OAUTH_TOKEN_FILE)



    return jsonify({'status': 'disconnected'})



# ==============================================================================



# HISTÓRICO



# ==============================================================================



HISTORY_FILE = os.path.join(BASE_DIR, 'historico.json')



def load_history():



    if os.path.exists(HISTORY_FILE):



        try:



            with open(HISTORY_FILE, 'r', encoding='utf-8') as f:



                return json.load(f)



        except:



            return []



    return []




# ---------------------------------------------------------------
# SINCRONIZACAO DE HISTORICO COM GITHUB
# ---------------------------------------------------------------
GITHUB_TOKEN   = os.environ.get('GITHUB_METRICS_TOKEN', '')
GITHUB_REPO    = 'gercy-junior/simplificae-ar'
GITHUB_API     = 'https://api.github.com'

def _push_history_github(entry):
    """
    Envia o registro de historico para o GitHub em background.
    Cria/atualiza arquivo: metrics/historico_{operador}.json
    Falha silenciosamente se sem internet ou sem token.
    """
    try:
        import json as _json, base64, ssl, urllib.request, re as _re, socket as _sock

        # Verificar conectividade rapida
        try:
            _sock.create_connection(('api.github.com', 443), timeout=3)
        except Exception:
            return  # sem internet, nao bloquear

        token = GITHUB_TOKEN
        if not token:
            # Tentar ler do .env
            env_path = os.path.join(BASE_DIR, '.env')
            if not os.path.exists(env_path):
                env_path = os.path.join(BASE_DIR, '_internal', '.env')
            if os.path.exists(env_path):
                with open(env_path, 'r', encoding='utf-8', errors='ignore') as _f:
                    for _line in _f:
                        if _line.startswith('GITHUB_METRICS_TOKEN='):
                            token = _line.split('=', 1)[1].strip()
                            break

        if not token:
            return  # sem token, nao sincronizar

        # Nome do arquivo por operador (sanitizar email)
        operador = entry.get('operador', 'desconhecido')
        safe_op  = _re.sub(r'[^a-zA-Z0-9._-]', '_', operador)
        filename = f'metrics/historico_{safe_op}.json'
        api_url  = f'{GITHUB_API}/repos/{GITHUB_REPO}/contents/{filename}'

        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE

        headers = {
            'Authorization': f'token {token}',
            'Accept': 'application/vnd.github.v3+json',
            'Content-Type': 'application/json',
            'User-Agent': 'SimplificaE-Metrics/1.0'
        }

        # Tentar GET para obter SHA e conteudo atual
        existing_data = []
        sha = None
        try:
            req = urllib.request.Request(api_url, headers=headers)
            with urllib.request.urlopen(req, context=ctx, timeout=8) as r:
                resp = _json.loads(r.read())
                sha  = resp.get('sha', '')
                content_b64 = resp.get('content', '').replace('\n', '')
                if content_b64:
                    existing_data = _json.loads(base64.b64decode(content_b64).decode('utf-8'))
        except Exception:
            pass  # arquivo nao existe ainda

        # Adicionar novo registro
        if not isinstance(existing_data, list):
            existing_data = []
        existing_data.insert(0, entry)
        existing_data = existing_data[:500]  # max 500 registros por operador

        # Codificar conteudo
        new_content = base64.b64encode(
            _json.dumps(existing_data, ensure_ascii=False, indent=2).encode('utf-8')
        ).decode('utf-8')

        # PUT para criar/atualizar
        body = {
            'message': f'metrics: cotacao {entry.get("tipo", "")} por {operador}',
            'content': new_content
        }
        if sha:
            body['sha'] = sha

        req2 = urllib.request.Request(
            api_url,
            data=_json.dumps(body).encode('utf-8'),
            headers=headers,
            method='PUT'
        )
        with urllib.request.urlopen(req2, context=ctx, timeout=10) as r:
            pass  # sucesso

    except Exception:
        pass  # silencioso — nao impacta o usuario

def _push_history_background(entry):
    """Dispara sync em thread daemon (nao bloqueia o usuario)."""
    import threading as _thr
    _thr.Thread(target=_push_history_github, args=(entry,), daemon=True).start()

def save_history_entry(entry):



    # Salvar local



    history = load_history()



    history.insert(0, entry)



    history = history[:200]



    with open(HISTORY_FILE, 'w', encoding='utf-8') as f:



        json.dump(history, f, indent=2, ensure_ascii=False)



    



    # Enviar pro servidor central (se configurado)



    server_url = os.environ.get('HISTORY_SERVER_URL', '').strip()



    if server_url and http_requests:



        try:



            http_requests.post(f'{server_url}/api/history', json=entry, timeout=5, verify=False)



        except Exception:



            pass  # falha silenciosa - nao bloqueia a operacao

    # Sincronizar com GitHub em background (silencioso)
    _push_history_background(entry)



@app.route('/calcular_ar', methods=['POST'])
def calcular_ar():
    """
    Calcula indicadores da Calculadora AR por empresa para a sessão ativa.
    Body JSON: {
        "session_id": "...",
        "taxa_map": {"Empresa A": 0.0137, "Empresa B": 0.0122},  // taxa por empresa
        "taxa_default": 0.0137  // usada se empresa não estiver no taxa_map
    }
    Retorna: { "empresas": [ {nome, taxa, resultado...}, ... ], "aviso_curvas": ... }
    """
    try:
        data = request.get_json(force=True) or {}
        session_id  = data.get('session_id') or request.args.get('session_id')
        taxa_map    = data.get('taxa_map', {}) or {}
        taxa_default = float(data.get('taxa_default', 0.0137) or 0.0137)
        custos_cerc_total = float(data.get('custos_cerc', 0) or 0)

        if not session_id:
            return jsonify({'erro': 'session_id não informado'}), 400

        sess_path = os.path.join(UPLOAD_DIR, f'{session_id}.json')
        if not os.path.exists(sess_path):
            return jsonify({'erro': 'Sessão não encontrada.'}), 404

        with open(sess_path) as jf:
            sess = json.load(jf)

        file_paths = sess.get('files', [sess['file']] if 'file' in sess else [])
        records = []
        for fp in file_paths:
            if os.path.exists(fp):
                records.extend(parse_agenda(fp))

        if not records:
            return jsonify({'erro': 'Sessão sem URs carregadas.'}), 404

        # Agrupar records por empresa usando raiz_map
        raiz_map  = load_raizes()
        emp_map   = {}  # {nome_empresa: [records]}
        for r in records:
            raiz = r.get('raiz', '')
            nome = raiz_map.get(raiz, raiz or 'Desconhecido')
            emp_map.setdefault(nome, []).append(r)

        # Se não tem raiz_map, tratar tudo como uma empresa
        if not raiz_map or len(emp_map) <= 1:
            taxa_am = taxa_default
            resultado = calcular_indicadores_ar(records, taxa_am)
            nome_unico = list(emp_map.keys())[0] if emp_map else 'Todos'
            return jsonify({
                'empresas': [{'nome': nome_unico, 'taxa': taxa_am, **resultado}],
                'aviso_curvas': resultado.get('aviso_curvas')
            })

        # Calcular por empresa — filtra grupos sem raiz CNPJ válida
        resultados = []
        aviso_global = None
        # Volume total para distribuicao proporcional de custos CERC
        _vol_total = sum(r.get('disponivel', 0) for recs in emp_map.values() for r in recs if r.get('disponivel', 0) > 0) or 1.0
        for nome, recs in sorted(emp_map.items()):
            # Ignorar grupos que parecem cabeçalho duplicado (raiz não numérica)
            if nome and not nome.replace(' ','').replace('-','').isdigit():
                # Nome resolvido — ok
                pass
            elif nome and len(nome.replace(' ','')) < 6:
                continue  # raiz muito curta — lixo do CSV
            taxa_am = float(taxa_map.get(nome, taxa_default))
            if taxa_am <= 0 or taxa_am > 0.5:
                taxa_am = taxa_default
            _vol_emp = sum(r.get('disponivel', 0) for r in recs if r.get('disponivel', 0) > 0)
            _cerc_emp = custos_cerc_total * (_vol_emp / _vol_total) if _vol_total > 0 else 0.0
            res = calcular_indicadores_ar(recs, taxa_am)  # CERC calculado automaticamente da planilha
            if res.get('erro') and res['erro'] == 'Nenhuma UR válida com valor disponível.':
                continue  # Pular empresa sem URs válidas silenciosamente
            if res.get('aviso_curvas'):
                aviso_global = res['aviso_curvas']
            resultados.append({'nome': nome, 'taxa': taxa_am, **res})

        return jsonify({'empresas': resultados, 'aviso_curvas': aviso_global})

    except Exception as e:
        app.logger.exception('Erro em /calcular_ar')
        return jsonify({'erro': str(e)}), 500


@app.route('/history')



def get_history():



    server_url = os.environ.get('HISTORY_SERVER_URL', '').strip()



    if server_url and http_requests:



        try:



            r = http_requests.get(f'{server_url}/api/history', timeout=10, verify=False)



            if r.status_code == 200:



                return jsonify(r.json())



        except Exception:



            pass



    # Fallback: historico local



    return jsonify(load_history())



@app.route('/api/history', methods=['GET', 'POST'])



def api_history():



    """API pra servidor central - recebe e serve historico de todos."""



    if request.method == 'POST':



        entry = request.json



        if entry:



            save_history_entry(entry)



        return jsonify({'status': 'ok'})



    return jsonify(load_history())



def _warmup_databricks():



    """Acorda o warehouse Databricks em background no startup.



    Elimina o cold start da primeira operacao real.



    """



    import threading as _t



    def _do():



        try:



            token = get_databricks_token()



            if not token or not DATABRICKS_WAREHOUSE_ID or not http_requests:



                return



            headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}



            resp = http_requests.post(



                f"{DATABRICKS_HOST}/api/2.0/sql/statements/",



                headers=headers,



                json={"warehouse_id": DATABRICKS_WAREHOUSE_ID,



                      "statement": "SELECT 1",



                      "wait_timeout": "30s"},



                timeout=35



            )



            state = resp.json().get("status", {}).get("state", "")



            print(f"[warmup] Databricks warehouse: {state}")



        except Exception as e:



            print(f"[warmup] Databricks indisponivel: {e}")



    _t.Thread(target=_do, daemon=True).start()



# =========================================================================

# E-MAIL: config, envio e historico

# =========================================================================

EMAIL_CONFIG_FILE = os.path.join(_DATA_DIR, 'email_config.json')

EMAIL_HISTORY_FILE = os.path.join(_DATA_DIR, 'email_history.json')

EMAIL_DEST_FILE = os.path.join(_DATA_DIR, 'email_destinatarios.json')


# Gmail OAuth2 via HTTPS (fallback quando SMTP bloqueado por firewall)
GMAIL_TOKEN_FILE = os.path.join(_DATA_DIR, 'gmail_token.json')
GMAIL_SCOPES     = ['https://www.googleapis.com/auth/gmail.send']
_GMAIL_CLIENT_ID     = os.environ.get('GMAIL_CLIENT_ID', '')
_GMAIL_CLIENT_SECRET = os.environ.get('GMAIL_CLIENT_SECRET', '')
_GMAIL_REDIRECT_URI  = 'http://localhost:5000/gmail/callback'

def _gmail_token_load():
    if not os.path.exists(GMAIL_TOKEN_FILE):
        return None
    try:
        with open(GMAIL_TOKEN_FILE, 'r') as f:
            return json.load(f)
    except Exception:
        return None

def _gmail_token_save(data):
    with open(GMAIL_TOKEN_FILE, 'w') as f:
        json.dump(data, f, indent=2)

def _gmail_refresh_token(token_data):
    try:
        import urllib.request as _ur, urllib.parse as _up
        body = _up.urlencode({
            'client_id':     _GMAIL_CLIENT_ID,
            'client_secret': _GMAIL_CLIENT_SECRET,
            'refresh_token': token_data.get('refresh_token', ''),
            'grant_type':    'refresh_token',
        }).encode()
        req = _ur.Request('https://oauth2.googleapis.com/token', data=body,
                          headers={'Content-Type': 'application/x-www-form-urlencoded'})
        resp = json.loads(_ur.urlopen(req, timeout=15).read())
        if 'access_token' in resp:
            token_data['access_token'] = resp['access_token']
            if 'refresh_token' in resp:
                token_data['refresh_token'] = resp['refresh_token']
            _gmail_token_save(token_data)
            return token_data
    except Exception:
        pass
    return None

def _gmail_send_api(msg_obj):
    token = _gmail_token_load()
    if not token:
        return 'Gmail nao autorizado. Clique em Autorizar Gmail na secao de e-mail.'
    try:
        import urllib.request as _ur, base64 as _b64, time as _t
        exp = token.get('expires_at', 0)
        if exp and _t.time() > exp - 60:
            token = _gmail_refresh_token(token)
            if not token:
                return 'Sessao Gmail expirada. Clique em Autorizar Gmail novamente.'
        at = token['access_token']
        raw = _b64.urlsafe_b64encode(msg_obj.as_bytes()).decode()
        body = json.dumps({'raw': raw}).encode()
        req = _ur.Request(
            'https://gmail.googleapis.com/gmail/v1/users/me/messages/send',
            data=body,
            headers={'Authorization': 'Bearer ' + at, 'Content-Type': 'application/json'},
            method='POST'
        )
        json.loads(_ur.urlopen(req, timeout=20).read())
        return None
    except Exception as e:
        err = str(e)
        if '401' in err or 'invalid_grant' in err.lower():
            try: os.remove(GMAIL_TOKEN_FILE)
            except: pass
            return 'Sessao Gmail expirada. Clique em Autorizar Gmail novamente.'
        return 'Erro Gmail API: ' + err



def _smtp_send(cfg, msg_obj, to_list):
    """Envia e-mail tentando automaticamente 587/STARTTLS e 465/SSL.
    Ignora a porta salva na config — testa na ordem que funcionar na rede atual.
    Salva a porta que funcionou no email_config.json para proximas chamadas.
    Retorna None se OK, ou string de erro se falhou nas duas portas.
    """
    host = cfg.get('smtp_host', 'smtp.gmail.com')
    user = cfg['smtp_user']
    pwd  = cfg['smtp_pass']

    saved_port = int(cfg.get('smtp_port', 587))
    attempts = [(465, 'ssl'), (587, 'starttls')] if saved_port == 465 else [(587, 'starttls'), (465, 'ssl')]

    last_err = None
    for port, mode in attempts:
        try:
            if mode == 'ssl':
                with smtplib.SMTP_SSL(host, port, timeout=20) as srv:
                    srv.login(user, pwd)
                    srv.send_message(msg_obj, user, to_list)
            else:
                with smtplib.SMTP(host, port, timeout=20) as srv:
                    srv.ehlo(); srv.starttls(); srv.ehlo()
                    srv.login(user, pwd)
                    srv.send_message(msg_obj, user, to_list)
            if port != saved_port:
                try:
                    cfg2 = load_email_config()
                    cfg2['smtp_port'] = port
                    with open(EMAIL_CONFIG_FILE, 'w') as f:
                        import json as _json; _json.dump(cfg2, f, indent=2)
                except Exception:
                    pass
            return None
        except Exception as e:
            last_err = str(e)
            continue
    # SMTP falhou em todas as portas — tentar Gmail API via HTTPS/443
    if _gmail_token_load():
        gmail_err = _gmail_send_api(msg_obj)
        if gmail_err is None:
            return None  # enviado via Gmail API
        return gmail_err
    return last_err


def load_email_config():

    if os.path.exists(EMAIL_CONFIG_FILE):

        try:

            with open(EMAIL_CONFIG_FILE, 'r', encoding='utf-8') as f:

                return json.load(f)

        except Exception:

            pass

    return {}



def load_email_history():

    if os.path.exists(EMAIL_HISTORY_FILE):

        try:

            with open(EMAIL_HISTORY_FILE, 'r', encoding='utf-8') as f:

                return json.load(f)

        except Exception:

            pass

    return []



def save_email_history_entry(entry):

    hist = load_email_history()

    hist.insert(0, entry)

    hist = hist[:500]

    with open(EMAIL_HISTORY_FILE, 'w', encoding='utf-8') as f:

        json.dump(hist, f, indent=2, ensure_ascii=False)



@app.route('/setup_email', methods=['POST'])

def setup_email():

    data = request.json or {}

    cfg = {

        # strip() + replace \xa0 (non-breaking space) por espaço normal
        # Evita erro ascii ao copiar senha de PDF/Word/e-mail
        'smtp_user':     data.get('smtp_user', '').strip().replace('\xa0', ' '),
        'smtp_pass':     data.get('smtp_pass', '').strip().replace('\xa0', ' '),

        'smtp_host':     data.get('smtp_host', 'smtp.gmail.com'),

        'smtp_port':     int(data.get('smtp_port', 587)),

        'display_name':  data.get('display_name', 'PicPay AR').strip()

    }

    if not cfg['smtp_user'] or not cfg['smtp_pass']:

        return jsonify({'error': 'E-mail e App Password sao obrigatorios'}), 400

    with open(EMAIL_CONFIG_FILE, 'w', encoding='utf-8') as f:

        json.dump(cfg, f, indent=2)

    return jsonify({'status': 'ok', 'user': cfg['smtp_user']})



@app.route('/test_smtp', methods=['GET'])
def test_smtp_route():
    '''Testa conectividade SMTP e retorna diagnostico detalhado.'''
    import socket as _sock
    cfg = load_email_config()
    host = cfg.get('smtp_host', 'smtp.gmail.com')
    results = []
    for port in [587, 465, 25, 443, 80]:
        try:
            s = _sock.create_connection((host, port), timeout=5)
            s.close()
            results.append({'port': port, 'ok': True, 'error': None})
        except Exception as e:
            results.append({'port': port, 'ok': False, 'error': str(e)})
    # Testar login se credenciais estiverem configuradas
    login_result = None
    if cfg.get('smtp_user') and cfg.get('smtp_pass'):
        port_ok = next((r['port'] for r in results if r['ok']), None)
        if port_ok:
            try:
                if port_ok == 465:
                    with smtplib.SMTP_SSL(host, port_ok, timeout=10) as srv:
                        srv.login(cfg['smtp_user'], cfg['smtp_pass'])
                else:
                    with smtplib.SMTP(host, port_ok, timeout=10) as srv:
                        srv.ehlo(); srv.starttls(); srv.ehlo()
                        srv.login(cfg['smtp_user'], cfg['smtp_pass'])
                login_result = {'ok': True, 'port': port_ok, 'user': cfg['smtp_user'][:15]+'...'}
            except Exception as e:
                login_result = {'ok': False, 'port': port_ok, 'error': str(e)}
        else:
            login_result = {'ok': False, 'error': 'Nenhuma porta acessivel'}
    return jsonify({
        'host': host,
        'connectivity': results,
        'login': login_result,
        'configured': bool(cfg.get('smtp_user') and cfg.get('smtp_pass'))
    })


@app.route('/email_config')

def get_email_config():

    cfg = load_email_config()

    return jsonify({

        'smtp_user':    cfg.get('smtp_user', ''),

        'display_name': cfg.get('display_name', ''),

        'configured':   bool(cfg.get('smtp_pass'))

    })



@app.route('/gmail/status')
def gmail_status():
    token = _gmail_token_load()
    if not token:
        return jsonify({'authorized': False})
    email = token.get('email', '')
    import time as _t
    exp = token.get('expires_at', 0)
    valid = not exp or _t.time() < exp
    return jsonify({'authorized': True, 'email': email, 'valid': valid})

@app.route('/gmail/auth')
def gmail_auth():
    import urllib.parse as _up
    params = {
        'client_id':     _GMAIL_CLIENT_ID,
        'redirect_uri':  _GMAIL_REDIRECT_URI,
        'response_type': 'code',
        'scope':         ' '.join(GMAIL_SCOPES),
        'access_type':   'offline',
        'prompt':        'consent',
    }
    url = 'https://accounts.google.com/o/oauth2/v2/auth?' + _up.urlencode(params)
    return redirect(url)

@app.route('/gmail/callback')
def gmail_callback():
    code = request.args.get('code')
    error = request.args.get('error')
    if error or not code:
        return '<h3>Autorização cancelada.</h3><p>Feche esta aba e tente novamente.</p>'
    try:
        import urllib.request as _ur, urllib.parse as _up, time as _t
        body = _up.urlencode({
            'code':          code,
            'client_id':     _GMAIL_CLIENT_ID,
            'client_secret': _GMAIL_CLIENT_SECRET,
            'redirect_uri':  _GMAIL_REDIRECT_URI,
            'grant_type':    'authorization_code',
        }).encode()
        req = _ur.Request('https://oauth2.googleapis.com/token', data=body,
                          headers={'Content-Type': 'application/x-www-form-urlencoded'})
        resp = json.loads(_ur.urlopen(req, timeout=15).read())
        if 'access_token' not in resp:
            return '<h3>Erro na autorização.</h3><p>' + str(resp) + '</p>'
        # Buscar e-mail do usuário
        req2 = _ur.Request('https://www.googleapis.com/oauth2/v2/userinfo',
                           headers={'Authorization': 'Bearer ' + resp['access_token']})
        user_info = json.loads(_ur.urlopen(req2, timeout=10).read())
        expires_at = _t.time() + int(resp.get('expires_in', 3600))
        token_data = {
            'access_token':  resp['access_token'],
            'refresh_token': resp.get('refresh_token', ''),
            'expires_at':    expires_at,
            'email':         user_info.get('email', ''),
        }
        _gmail_token_save(token_data)
        email = token_data['email']
        return (
            '<html><body style="font-family:Arial;text-align:center;padding:60px;">'
            '<div style="color:#2E7D32;font-size:48px;">&#10003;</div>'
            '<h2 style="color:#1B5E20;">Gmail autorizado!</h2>'
            '<p style="color:#555;">Conta: <strong>' + email + '</strong></p>'
            '<p style="color:#555;">Pode fechar esta aba e voltar ao SimplificaE.</p>'
            '</body></html>'
        )
    except Exception as e:
        return '<h3>Erro:</h3><p>' + str(e) + '</p>'

@app.route('/gmail/revoke', methods=['POST'])
def gmail_revoke():
    try:
        if os.path.exists(GMAIL_TOKEN_FILE):
            os.remove(GMAIL_TOKEN_FILE)
    except Exception:
        pass
    return jsonify({'status': 'ok'})


@app.route('/send_email', methods=['POST'])

def send_email_route():

    import smtplib

    from email.mime.multipart import MIMEMultipart

    from email.mime.text import MIMEText

    from email.mime.base import MIMEBase

    from email import encoders as _enc
    from email.header import Header as _Header



    data = request.json or {}

    sid            = data.get('session_id', '')

    empresa        = data.get('empresa', '')

    safe_name      = data.get('safe_name', '')

    to_email_raw   = data.get('to_email', '')
    to_emails_list = data.get('to_emails', [])
    if to_emails_list:
        to_emails_list = [e.strip() for e in to_emails_list if e.strip()]
    elif to_email_raw:
        to_emails_list = [e.strip() for e in str(to_email_raw).split(',') if e.strip()]
    to_email = ', '.join(to_emails_list)

    custom_message = data.get('custom_message', '').strip()

    is_custom      = data.get('is_custom', False)

    taxa_pct_email  = float(data.get('taxa_pct', 0))
    perfil_cliente  = data.get('perfil_cliente', 'recorrente')  # recorrente | novo | taxa_zero
    operator_email = data.get('operator_email', '').strip()

    urs            = int(data.get('urs', 0))

    valor_total    = float(data.get('valor_total', 0))

    valor_operavel = float(data.get('valor_operavel', 0))



    if not to_email:

        return jsonify({'error': 'E-mail destinatario obrigatorio'}), 400



    cfg = load_email_config()

    if not cfg.get('smtp_user') or not cfg.get('smtp_pass'):

        return jsonify({'error': 'Configure o e-mail do operador primeiro (secao 1)'}), 400



    # Encontrar arquivo para anexar: prefer COMPLETO, fallback Elegiveis

    out_dir = os.path.join(OUTPUT_DIR, sid, safe_name)

    attach_path = None

    attach_name = None

    if os.path.exists(out_dir):

        for prefix in ['Cotacao_COMPLETO_', 'Cotacao_Elegiveis_']:

            for fname in sorted(os.listdir(out_dir)):

                if fname.startswith(prefix) and fname.endswith('.xlsx'):

                    attach_path = os.path.join(out_dir, fname)

                    attach_name = fname

                    break

            if attach_path:

                break



    def _brl(v):

        s = '{:,.2f}'.format(v)

        # pt-BR: trocar , por X, . por ,, X por .

        return 'R$ ' + s.replace(',', 'X').replace('.', ',').replace('X', '.')



    hoje = datetime.now().strftime('%d/%m/%Y')

    # pct: usar taxa do operador (inclusive 0), senao ratio
    if 'taxa_pct' in data:
        pct = round(float(data['taxa_pct']), 2)
    elif valor_total > 0:
        pct = round(valor_operavel / valor_total * 100)
    else:
        pct = 0



    subject = '[PicPay AR] Agenda de Antecipação de Recebíveis — {} — {}'.format(empresa, hoje)



    _label_agenda = 'Simulação total da operação' if is_custom else 'Agenda total disponível'



    _falar_email = operator_email or cfg.get('smtp_user', '')

    # === 3 PERFIS DE EMAIL ===
    # Calculos condicionais
    _td_em = (1 + taxa_pct_email/100)**(1/30) - 1 if taxa_pct_email > 0 else 0
    _desc  = ((1 + _td_em)**30 - 1) if _td_em > 0 else 0
    _vliq  = valor_total * (1 - _desc) if taxa_pct_email > 0 and valor_total > 0 else 0

    _tr_taxa = (
        '<tr style="border-top:1px solid #F5F5F5;background:#E8F5E9;">'
        '<td style="padding:10px;font-size:14px;color:#1B5E20;font-weight:bold;">Taxa de operação</td>'
        '<td style="padding:10px;font-size:18px;font-weight:bold;color:#1B5E20;text-align:right;">' + str(round(pct,2)) + '%</td></tr>'
    ) if pct > 0 else ''

    _tr_liq = (
        '<tr style="border-top:1px solid #F5F5F5;">'
        '<td style="padding:8px 0;font-size:13px;color:#757575;">Valor líquido estimado</td>'
        '<td style="padding:8px 0;font-size:15px;font-weight:bold;text-align:right;color:#1B5E20;">' + _brl(_vliq) + '</td></tr>'
    ) if pct > 0 and _vliq > 0 else ''

    if perfil_cliente == 'novo':
        _intro  = ('<p style="font-size:15px;">Olá,</p>'
                   '<p style="font-size:15px;">Você sabia que no PicPay você já pode antecipar valores '
                   'transacionados em outras maquininhas com uma taxa bem abaixo da média de mercado?</p>'
                   '<p style="font-size:15px;">Já estamos antecipando as principais adquirentes do mercado, '
                   'como <strong>Cielo, Rede, Getnet, Stone, Safrapay e Adyen</strong>.</p>'
                   '<p style="font-size:15px;">Segue em anexo os valores disponíveis para antecipação.</p>')
        _rodape = 'Para contratar, negociar ou simular outros valores, entre em contato com um especialista:'
        _cta    = 'Falar com um especialista'
    elif perfil_cliente == 'taxa_zero':
        _intro  = ('<p style="font-size:15px;">Olá,</p>'
                   '<p style="font-size:15px;">Segue agenda disponível para antecipação.</p>')
        _rodape = 'Para negociar ou simular uma operação sem compromisso, entre em contato com um especialista:'
        _cta    = 'Falar com um especialista'
    else:  # recorrente (default)
        _intro  = ('<p style="font-size:15px;">Olá,</p>'
                   '<p style="font-size:15px;">Segue agenda disponível para antecipação.</p>')
        _rodape = 'Para contratar, negociar ou simular outros valores, entre em contato com um especialista:'
        _cta    = 'Falar com um especialista'

    _extra = ('<p style="font-size:14px;">' + custom_message + '</p>') if custom_message else ''

    body = (
        '<html><body style="font-family:Arial,sans-serif;color:#212121;max-width:600px;margin:0 auto;">'
        '<div style="background:#1B5E20;padding:20px 24px;border-radius:8px 8px 0 0;">'
        '<h2 style="color:white;margin:0;font-size:20px;">PicPay — Antecipação de Recebíveis</h2>'
        '<p style="color:#C8E6C9;margin:4px 0 0;font-size:13px;">' + hoje + '</p></div>'
        '<div style="background:#F9FBE7;padding:20px 24px;border:1px solid #E0E0E0;border-top:none;">'
        + _intro
        + '<div style="background:white;border:1px solid #E0E0E0;border-radius:8px;padding:16px 20px;margin:16px 0;">'
        '<table style="width:100%;border-collapse:collapse;">'
        '<tr><td style="padding:8px 0;font-size:13px;color:#757575;">' + _label_agenda + '</td>'
        '<td style="padding:8px 0;font-size:15px;font-weight:bold;text-align:right;">' + _brl(valor_total) + '</td></tr>'
        + _tr_taxa + _tr_liq + '</table></div>'
        + _extra
        + '<p style="font-size:14px;">' + _rodape + '</p>'
        '<div style="text-align:center;margin:24px 0;">'
        '<a href="mailto:' + _falar_email + '" style="background:#4CAF50;color:white;'
        'padding:12px 28px;border-radius:6px;text-decoration:none;font-size:14px;font-weight:bold;">' + _cta + '</a>'
        '</div>'
        '<p style="font-size:12px;color:#9E9E9E;border-top:1px solid #E0E0E0;padding-top:12px;">'
        'Gerado automaticamente pelo SimplificaÊ — PicPay AR. Data: ' + hoje + '</p>'
        '</div></body></html>'
    )



    msg = MIMEMultipart('alternative')

    # Encodar headers para suportar caracteres nao-ASCII (ex: \xa0 em nomes de empresa)
    def _safe_header(s):
        try:
            s.encode('ascii')
            return s
        except (UnicodeEncodeError, UnicodeDecodeError):
            return str(_Header(s, 'utf-8'))

    _display = cfg.get('display_name', 'PicPay AR') or 'PicPay AR'
    msg['From']    = '{} <{}>'.format(_safe_header(_display), cfg['smtp_user'])

    msg['To']      = ', '.join(to_emails_list)

    msg['Subject'] = _safe_header(subject)

    msg.attach(MIMEText(body, 'html', 'utf-8'))



    if attach_path and os.path.exists(attach_path):

        with open(attach_path, 'rb') as af:

            part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')

            part.set_payload(af.read())

            _enc.encode_base64(part)

            part.add_header('Content-Disposition', 'attachment; filename="{}"'.format(attach_name))

            msg.attach(part)



    # Usar _smtp_send com autodetect de porta (587/STARTTLS ou 465/SSL)
    # Fix: WinError 10013 em máquinas com porta 587 bloqueada pelo firewall corporativo
    _smtp_err = _smtp_send(cfg, msg, to_emails_list)
    if _smtp_err:
        # Traduzir erro tecnico para mensagem clara para o operador
        _err_lower = _smtp_err.lower()
        if '10013' in _smtp_err or 'permission' in _err_lower or 'proibida' in _err_lower:
            _msg_erro = 'Rede bloqueada: o firewall desta m\u00e1quina n\u00e3o permite envio de e-mails. Entre em contato com o suporte TI.'
        elif 'username' in _err_lower or 'password' in _err_lower or '535' in _smtp_err:
            _msg_erro = 'Senha incorreta: verifique o App Password. Gere um novo em myaccount.google.com > Seguran\u00e7a > Senhas de app.'
        elif 'timed out' in _err_lower or 'timeout' in _err_lower:
            _msg_erro = 'Tempo esgotado ao conectar ao servidor de e-mail. Verifique a conex\u00e3o de rede.'
        elif 'less secure' in _err_lower or 'application' in _err_lower:
            _msg_erro = 'Conta Gmail n\u00e3o permite login por aplicativos. Use um App Password gerado em myaccount.google.com.'
        else:
            _msg_erro = 'Falha ao enviar e-mail. Clique em Salvar e Testar na se\u00e7\u00e3o de configura\u00e7\u00e3o para diagnosticar.'
        return jsonify({'error': _msg_erro}), 500
    status_e = 'enviado'
    error_e  = None



    save_email_history_entry({

        'timestamp':       datetime.now().strftime('%d/%m/%Y %H:%M:%S'),

        'operador':        cfg.get('smtp_user', ''),

        'empresa':         empresa,

        'session_id':      sid,

        'to_email':        to_email,

        'status':          status_e,

        'valor_total':     valor_total,

        'valor_operavel':  valor_operavel,

        'urs':             urs,

        'arquivo':         attach_name or ''

    })

    # Telemetria: evento de envio de e-mail
    if _TELEMETRIA_OK:
        _usage_log.registrar(
            operador=cfg.get('smtp_user', ''),
            evento='send_email',
            empresa=empresa,
            urs=urs,
            valor_bruto=valor_total,
            valor_operavel=valor_operavel,
            taxa=float(data.get('taxa_pct', 0) or 0),
            status=status_e,
            session_id=sid,
        )

    # Salvar e-mail para reuso

    try:

        dest = {}

        if os.path.exists(EMAIL_DEST_FILE):

            with open(EMAIL_DEST_FILE) as f:

                dest = json.load(f)

        existing = dest.get(empresa.upper(), [])
        if isinstance(existing, str): existing = [existing] if existing else []
        merged = list(existing)
        for _em in to_emails_list:
            if _em and _em not in merged: merged.append(_em)
        dest[empresa.upper()] = merged

        with open(EMAIL_DEST_FILE, 'w') as f:

            json.dump(dest, f, indent=2)

    except Exception:

        pass



    return jsonify({'status': 'ok', 'message': 'E-mail enviado para {}'.format(to_email)})



@app.route('/email_history')

def get_email_history():

    limit = request.args.get('limit', 100, type=int)

    return jsonify(load_email_history()[:limit])



@app.route('/email_destinatarios')

def get_email_destinatarios():

    if os.path.exists(EMAIL_DEST_FILE):

        try:

            with open(EMAIL_DEST_FILE) as f:

                return jsonify(json.load(f))

        except Exception:

            pass

    return jsonify({})



# ==============================================================================
# HERODASH FLOW — Cotação Rápida
# Fluxo: CNPJ/raiz -> baixa agenda HD -> upload -> gera cotação taxa=0 -> envia email
# ==============================================================================

# Token do HeroDash — lido do arquivo salvo pelo plugin herodash-connector
_HD_API_BASE   = 'https://herodash-api.picpay.com/api/v1'
_HD_API_BFF    = 'https://herodash-api.picpay.com/api/v1/herodash-bff'
_HD_SELLER_SVC = 'https://herodash-seller-service-bff.picpay.com/api/v1'

# Possíveis locais do arquivo de token salvo pelo plugin herodash-connector
_HD_TOKEN_CANDIDATES = [
    os.path.join(os.path.expanduser('~'), '.wolf', 'herodash-auth.json'),
    os.path.join('C:\\', 'tmp', '.wolf', 'herodash-auth.json'),
    os.path.join('/tmp', '.wolf', 'herodash-auth.json'),
]

def _hd_token():
    """Retorna o JWT do HeroDash salvo pelo plugin herodash-connector.
    
    Suporta múltiplos formatos:
    - {'token': 'eyJ...'}
    - {'localStorage': {'token': 'eyJ...'}}
    - storageState Playwright: {'origins': [{'localStorage': [{'name':'token','value':'eyJ...'}]}]}
    """
    for token_file in _HD_TOKEN_CANDIDATES:
        if not os.path.exists(token_file):
            continue
        try:
            with open(token_file, encoding='utf-8') as f:
                data = json.load(f)
            
            # Formato 1: {'token': 'eyJ...'}
            if isinstance(data.get('token'), str) and data['token'].startswith('eyJ'):
                return data['token']
            
            # Formato 2: {'localStorage': {'token': 'eyJ...'}}
            ls = data.get('localStorage')
            if isinstance(ls, dict):
                tok = ls.get('token')
                if tok and isinstance(tok, str) and tok.startswith('eyJ'):
                    return tok
            
            # Formato 3: storageState Playwright
            # {'origins': [{'origin': 'https://...', 'localStorage': [{'name':'token','value':'eyJ...'}]}]}
            origins = data.get('origins', [])
            for origin in origins:
                ls = origin.get('localStorage', [])
                if isinstance(ls, list):
                    for item in ls:
                        if item.get('name') == 'token':
                            val = item.get('value', '')
                            if val and isinstance(val, str) and val.startswith('eyJ'):
                                return val
                elif isinstance(ls, dict):
                    tok = ls.get('token')
                    if tok and isinstance(tok, str) and tok.startswith('eyJ'):
                        return tok
        except Exception:
            continue
    return None

def _hd_headers():
    tok = _hd_token()
    if not tok:
        return None
    return {'Authorization': f'Bearer {tok}', 'Content-Type': 'application/json'}

def _hd_search_seller(cnpj_or_raiz):
    """Busca o seller pelo CNPJ/raiz. Retorna lista [{id, name, ...}]."""
    if http_requests is None:
        return None, 'requests nao instalado'
    hdrs = _hd_headers()
    if not hdrs:
        return None, 'Token HeroDash nao encontrado. Faça login no HeroDash primeiro.'
    q = cnpj_or_raiz.replace('.', '').replace('/', '').replace('-', '')
    try:
        r = http_requests.get(
            f'{_HD_API_BASE}/sellers/search',
            params={'q': q},
            headers=hdrs,
            timeout=15
        )
        if r.status_code == 401:
            return None, 'Token HeroDash expirado. Faça login novamente.'
        r.raise_for_status()
        return r.json(), None
    except Exception as e:
        return None, str(e)

def _hd_gerar_agenda(cnpj_raiz, use_raiz=True):
    """Solicita geração de agenda no HD. Retorna file_id ou erro.
    
    Nova API (herodash-bff):
    POST /herodash-bff/advance-receivables-agenda
    Body: {"cnpjs": ["14digits"], "rootCnpjs": []} ou {"cnpjs": [], "rootCnpjs": ["8digits"]}
    """
    if http_requests is None:
        return None, 'requests nao instalado'
    hdrs = _hd_headers()
    if not hdrs:
        return None, 'Token HeroDash nao encontrado.'
    cnpj_fmt = cnpj_raiz.replace('.', '').replace('/', '').replace('-', '')
    if use_raiz:
        payload = {'cnpjs': [], 'rootCnpjs': [cnpj_fmt[:8]]}
    else:
        # CNPJ completo deve ter 14 dígitos
        if len(cnpj_fmt) < 14:
            cnpj_fmt = cnpj_fmt.zfill(14)
        payload = {'cnpjs': [cnpj_fmt], 'rootCnpjs': []}
    try:
        r = http_requests.post(
            f'{_HD_API_BFF}/advance-receivables-agenda',
            json=payload,
            headers=hdrs,
            timeout=20
        )
        if r.status_code == 401:
            return None, 'Token HeroDash expirado.'
        if r.status_code == 404:
            return None, 'Rota não encontrada. Verifique permissões Mesa AR.'
        r.raise_for_status()
        data = r.json()
        # Resposta pode ser o objeto da agenda ou lista
        file_id = data.get('fileId') or data.get('id') or data.get('file_id')
        if not file_id and isinstance(data, list) and len(data) > 0:
            file_id = data[0].get('fileId')
        return file_id, None
    except Exception as e:
        return None, str(e)

def _hd_status_agenda(file_id):
    """Verifica status da agenda consultando a lista recente.
    
    Nova API: GET /herodash-bff/advance-receivables-agenda?sort=DESC&page=1&pageSize=10
    Retorna (status_str, None) — 'processed'/'PROCESSED' quando pronto
    """
    if http_requests is None:
        return None, None
    hdrs = _hd_headers()
    if not hdrs:
        return None, None
    try:
        r = http_requests.get(
            f'{_HD_API_BFF}/advance-receivables-agenda',
            params={'sort': 'DESC', 'page': 1, 'pageSize': 20},
            headers=hdrs,
            timeout=15
        )
        r.raise_for_status()
        data = r.json()
        items = data.get('data', data) if isinstance(data, dict) else data
        for item in items:
            if item.get('fileId') == file_id:
                status = (item.get('status') or '').lower()
                return status, None
        return 'processando', None
    except Exception:
        return None, None

def _hd_download_agenda(file_id):
    """Baixa o CSV da agenda.
    
    Nova API: GET /herodash-bff/advance-receivables-agenda/file/{fileId}/download
    Retorna (bytes_content, filename, error)
    """
    if http_requests is None:
        return None, None, 'requests nao instalado'
    hdrs = _hd_headers()
    if not hdrs:
        return None, None, 'Token HeroDash nao encontrado.'
    try:
        # 1. Obter a URL presignada do S3
        r_url = http_requests.get(
            f'{_HD_API_BFF}/advance-receivables-agenda/file/{file_id}/download',
            headers=hdrs,
            timeout=30
        )
        if r_url.status_code == 401:
            return None, None, 'Token HeroDash expirado.'
        if r_url.status_code == 422:
            return None, None, f'Arquivo ainda em processamento ({r_url.json().get("message","")[:100]})'
        r_url.raise_for_status()
        resp_json = r_url.json()
        file_url = resp_json.get('fileUrl') or resp_json.get('url') or resp_json.get('download_url')
        if not file_url:
            return None, None, f'URL de download não encontrada: {r_url.text[:200]}'
        
        # 2. Baixar o CSV da URL S3 (sem Authorization — URL presignada)
        r_csv = http_requests.get(file_url, timeout=120)
        r_csv.raise_for_status()
        fname = f'{file_id}.csv'
        cd = r_csv.headers.get('Content-Disposition', '')
        if 'filename=' in cd:
            fname = cd.split('filename=')[-1].strip().strip('"')
        return r_csv.content, fname, None
    except Exception as e:
        return None, None, str(e)

@app.route('/herodash/cotacao_rapida', methods=['POST'])
def hd_cotacao_rapida():
    """
    Fluxo multi-empresa com lotes de 5 raizes (limite HD).
    Body: {empresas:[{cnpj,emails:[]}], use_raiz, operator_email, enviar_email}
    Compativel com formato antigo: {cnpj, email_cliente}
    """
    import time as _time, re as _re

    if http_requests is None:
        return jsonify({'error': 'Modulo requests nao instalado.'}), 500

    data = request.json or {}
    use_raiz       = data.get('use_raiz', True)
    operator_email = data.get('operator_email', '').strip()
    enviar_email   = data.get('enviar_email', False)

    # Compatibilidade retroativa
    if 'cnpj' in data and 'empresas' not in data:
        cnpj_raw  = data.get('cnpj', '').strip()
        email_cli = data.get('email_cliente', '').strip()
        empresas_in = [{'cnpj': cnpj_raw, 'emails': [email_cli] if email_cli else []}]
    else:
        empresas_in = data.get('empresas', [])

    if not empresas_in:
        return jsonify({'error': 'Informe pelo menos um CNPJ ou raiz.'}), 400

    hdrs = _hd_headers()
    if not hdrs:
        return jsonify({'error': 'Token HeroDash nao encontrado. Faca login no plugin HeroDash do Nitro.'}), 401

    def _norm(cnpj, raiz):
        c = cnpj.replace('.','').replace('/','').replace('-','').strip()
        if raiz:
            return c[:8].zfill(8)
        return c.zfill(14) if len(c) < 14 else c

    cnpj_email_map = {}
    cnpjs = []
    for emp in empresas_in:
        raw = (emp.get('cnpj') or '').strip()
        if not raw: continue
        cn = _norm(raw, use_raiz)
        cnpj_email_map[cn] = emp.get('emails') or []
        cnpjs.append(cn)

    if not cnpjs:
        return jsonify({'error': 'Nenhum CNPJ valido informado.'}), 400

    BATCH = 5
    all_csv = []

    for bi in range(0, len(cnpjs), BATCH):
        lote = cnpjs[bi:bi+BATCH]
        payload_hd = {'cnpjs': [], 'rootCnpjs': lote} if use_raiz else {'cnpjs': lote, 'rootCnpjs': []}
        lote_num = bi // BATCH + 1

        try:
            r1 = http_requests.post(f'{_HD_API_BFF}/advance-receivables-agenda',
                json=payload_hd, headers=hdrs, timeout=20)
            if r1.status_code == 401:
                return jsonify({'error': 'Token HeroDash expirado.'}), 401
            if r1.status_code in (403, 404):
                return jsonify({'error': 'Sem acesso a Mesa AR no HeroDash.'}), 403
            # 422 = ja em processamento, extrair fileId e continuar
            if r1.status_code != 422:
                r1.raise_for_status()
        except http_requests.exceptions.RequestException as e:
            return jsonify({'error': f'Erro ao solicitar agenda (lote {lote_num}): {e}'}), 502

        rd = r1.json()
        file_id = rd.get('fileId') or rd.get('id') or rd.get('file_id')
        if not file_id and isinstance(rd, list) and rd:
            file_id = rd[0].get('fileId')
        if not file_id:
            m = _re.search(r'[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}', r1.text)
            file_id = m.group(0) if m else None
        if not file_id:
            return jsonify({'error': f'Resposta inesperada HD (lote {lote_num}): {r1.text[:200]}'}), 502

        # Polling
        max_wait, poll_iv, elapsed, status = 180, 8, 0, 'processando'
        while elapsed < max_wait:
            _time.sleep(poll_iv); elapsed += poll_iv
            try:
                rs = http_requests.get(f'{_HD_API_BFF}/advance-receivables-agenda',
                    params={'sort':'DESC','page':1,'pageSize':20}, headers=hdrs, timeout=15)
                rs.raise_for_status()
                items = rs.json().get('data', rs.json()) if isinstance(rs.json(), dict) else rs.json()
                for item in (items if isinstance(items, list) else []):
                    if item.get('fileId') == file_id:
                        status = (item.get('status') or '').lower(); break
                if status in ('processed',) or 'finaliz' in status or 'complet' in status: break
                if 'erro' in status or 'fail' in status:
                    return jsonify({'error': f'Agenda falhou (lote {lote_num}): {status}'}), 502
            except Exception:
                pass

        if status not in ('processed',) and 'finaliz' not in status and 'complet' not in status:
            return jsonify({'error': f'Agenda nao finalizou (lote {lote_num}) apos {max_wait}s.', 'file_id': file_id}), 202

        # Download CSV (presigned URL)
        try:
            ru = http_requests.get(f'{_HD_API_BFF}/advance-receivables-agenda/file/{file_id}/download',
                headers=hdrs, timeout=30)
            ru.raise_for_status()
            url = ru.json().get('fileUrl') or ru.json().get('url')
            if not url:
                return jsonify({'error': f'URL de download nao encontrada (lote {lote_num}).'}), 502
            rc = http_requests.get(url, timeout=120)
            rc.raise_for_status()
            if rc.content: all_csv.append(rc.content)
        except Exception as e:
            return jsonify({'error': f'Erro ao baixar CSV (lote {lote_num}): {e}'}), 502

    if not all_csv:
        return jsonify({'error': 'Nenhum CSV obtido do HeroDash.'}), 502

    csv_content = b'\n'.join(all_csv)

    # Upload
    sid = datetime.now().strftime('%Y%m%d_%H%M%S')
    fpath = os.path.join(UPLOAD_DIR, f'{sid}_0.csv')
    with open(fpath, 'wb') as fout:
        fout.write(csv_content)

    records = parse_agenda(fpath)
    if not records:
        return jsonify({'error': 'CSV baixado mas nenhum registro encontrado.'}), 422

    raiz_map  = load_raizes()
    raiz_to_emp = {}
    emp_raizes = {}
    for raiz, nome in raiz_map.items():
        nu = nome.upper().strip()
        emp_raizes.setdefault(nu, {'nome': nome, 'raizes': []})['raizes'].append(raiz)
    for nu, info in emp_raizes.items():
        for raiz in info['raizes']:
            raiz_to_emp[raiz] = info['nome']
    for r in records:
        raiz = r.get('raiz', '')
        if raiz not in raiz_to_emp:
            raiz_to_emp[raiz] = f'RAIZ_{raiz}'

    empresas_an = analyze_records(records, raiz_map)
    sess_data = {'files': [fpath], 'file': fpath,
                 'empresas': [{'nome': n, 'cnpjs': len(e['cnpjs']), 'urs': e['urs'], 'valor': round(e['valor'],2)}
                              for n, e in sorted(empresas_an.items(), key=lambda x: x[1]['valor'], reverse=True)],
                 'hd_cnpjs': cnpjs}
    with open(os.path.join(UPLOAD_DIR, f'{sid}.json'), 'w') as jf:
        json.dump(sess_data, jf)

    empresa_records = {}
    for r in records:
        emp = raiz_to_emp.get(r.get('raiz', ''), 'OUTROS')
        empresa_records.setdefault(emp, []).append(r)

    taxa_pct  = 0.0
    di_periodo = 0.1465
    out_dir    = os.path.join(OUTPUT_DIR, sid)
    os.makedirs(out_dir, exist_ok=True)
    generated  = []
    cfg_email  = load_email_config()

    for emp_nome, emp_recs in empresa_records.items():
        safe    = emp_nome.replace('/','_').replace('\\','_').replace(' ','_')[:40]
        emp_out = os.path.join(out_dir, safe)
        os.makedirs(emp_out, exist_ok=True)
        cotacao_path = os.path.join(emp_out, f'Cotacao_COMPLETO_{safe}.xlsx')
        emp_result = {'empresa': emp_nome, 'safe_name': safe, 'urs': len(emp_recs)}

        try:
            generate_cotacao(emp_recs, emp_nome, taxa_pct/100.0, di_periodo, {}, cotacao_path)
        except Exception as ex:
            emp_result['error'] = str(ex)
            generated.append(emp_result)
            continue

        if enviar_email and cfg_email.get('smtp_user') and cfg_email.get('smtp_pass'):
            # Descobrir raiz da empresa
            raiz_emp = next((r for r, n in raiz_to_emp.items() if n == emp_nome), None)
            emails_dest = (cnpj_email_map.get(raiz_emp) or []) if raiz_emp else []
            if not emails_dest:
                saved = {}
                if os.path.exists(EMAIL_DEST_FILE):
                    try:
                        with open(EMAIL_DEST_FILE) as f: saved = json.load(f)
                    except Exception: pass
                v = saved.get(emp_nome) or saved.get(emp_nome.upper())
                if v: emails_dest = v if isinstance(v, list) else [v]

            if emails_dest:
                try:
                    valor_total = sum(r.get('disponivel', 0) for r in emp_recs)
                    hoje = datetime.now().strftime('%d/%m/%Y')
                    def _brl(v): return 'R$ {:,.2f}'.format(v).replace(',','X').replace('.',',').replace('X','.')
                    subject = '[PicPay AR] Cotacao Indicativa - {} - {}'.format(emp_nome, hoje)
                    body_html = (
                        '<html><body style="font-family:Arial,sans-serif;color:#212121;max-width:600px;margin:0 auto;">'
                        '<div style="background:#1B5E20;padding:20px 24px;border-radius:8px 8px 0 0;">'
                        '<h2 style="color:white;margin:0;font-size:20px;">PicPay - Antecipacao de Recebiveis</h2>'
                        '<p style="color:#C8E6C9;margin:4px 0 0;font-size:13px;">' + hoje + '</p></div>'
                        '<div style="background:#F9FBE7;padding:20px 24px;border:1px solid #E0E0E0;border-top:none;">'
                        '<p style="font-size:15px;">Ola,</p>'
                        '<p style="font-size:15px;">Segue cotacao indicativa disponivel para antecipacao.</p>'
                        '<div style="background:white;border:1px solid #E0E0E0;border-radius:8px;padding:16px 20px;margin:16px 0;">'
                        '<table style="width:100%;border-collapse:collapse;">'
                        '<tr><td style="padding:8px 0;font-size:13px;color:#757575;">Agenda disponivel</td>'
                        '<td style="padding:8px 0;font-size:15px;font-weight:bold;text-align:right;">' + _brl(valor_total) + '</td></tr>'
                        '<tr><td style="padding:8px 0;font-size:13px;color:#757575;">Taxa indicativa</td>'
                        '<td style="padding:8px 0;font-size:15px;font-weight:bold;text-align:right;">0,00% a.m. (simulacao)</td></tr>'
                        '</table></div>'
                        '<p style="font-size:14px;">Para negociar ou simular sem compromisso, entre em contato:</p>'
                        '<div style="text-align:center;margin:24px 0;">'
                        '<a href="mailto:' + (operator_email or cfg_email.get('smtp_user','')) + '" '
                        'style="background:#4CAF50;color:white;padding:12px 28px;border-radius:6px;'
                        'text-decoration:none;font-size:14px;font-weight:bold;">Falar com um especialista</a>'
                        '</div>'
                        '<p style="font-size:12px;color:#9E9E9E;border-top:1px solid #E0E0E0;padding-top:12px;">'
                        'Gerado pelo SimplificaE - PicPay AR. Data: ' + hoje + '</p>'
                        '</div></body></html>'
                    )
                    msg = MIMEMultipart('alternative')
                    _display = cfg_email.get('display_name', 'PicPay AR') or 'PicPay AR'
                    msg['From']    = '{} <{}>'.format(_display, cfg_email['smtp_user'])
                    msg['To']      = ', '.join(emails_dest)
                    msg['Subject'] = subject
                    msg.attach(MIMEText(body_html, 'html', 'utf-8'))
                    if os.path.exists(cotacao_path):
                        with open(cotacao_path, 'rb') as af:
                            part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                            part.set_payload(af.read())
                        _email_encoders.encode_base64(part)
                        part.add_header('Content-Disposition', 'attachment; filename="Cotacao_{}.xlsx"'.format(safe))
                        msg.attach(part)
                    err = _smtp_send(cfg_email, msg, emails_dest)
                    if err: emp_result['email_erro'] = err
                    else:   emp_result['email_enviado'] = True
                except Exception as ex:
                    emp_result['email_erro'] = str(ex)

        generated.append(emp_result)

    return jsonify({'session_id': sid, 'records': len(records), 'empresas': generated, 'taxa_pct': taxa_pct})


@app.route('/herodash/status/<file_id>', methods=['GET'])
def hd_status_agenda(file_id):
    """Verifica status de uma agenda em processamento no HeroDash."""
    hdrs = _hd_headers()
    if not hdrs:
        return jsonify({'error': 'Token HeroDash não encontrado.'}), 401
    if http_requests is None:
        return jsonify({'error': 'requests não instalado'}), 500
    try:
        r = http_requests.get(
            f'{_HD_API_BFF}/advance-receivables-agenda',
            params={'sort': 'DESC', 'page': 1, 'pageSize': 20},
            headers=hdrs, timeout=15
        )
        if r.status_code == 401:
            return jsonify({'error': 'Token expirado'}), 401
        r.raise_for_status()
        data = r.json()
        items = data.get('data', data) if isinstance(data, dict) else data
        for item in (items if isinstance(items, list) else []):
            if item.get('fileId') == file_id:
                return jsonify(item)
        return jsonify({'fileId': file_id, 'status': 'nao_encontrado'})
    except Exception as e:
        return jsonify({'error': str(e)}), 502


@app.route('/herodash/token_status', methods=['GET'])
def hd_token_status():
    """Diagnóstico do token HeroDash."""
    tok = _hd_token()
    if not tok:
        candidates_info = []
        for p in _HD_TOKEN_CANDIDATES:
            candidates_info.append({'path': p, 'exists': os.path.exists(p)})
        return jsonify({'ok': False, 'error': 'Token não encontrado', 'candidates': candidates_info})
    # Decodificar payload JWT sem verificar assinatura
    try:
        import base64
        parts = tok.split('.')
        if len(parts) >= 2:
            payload_b64 = parts[1] + '=='
            payload = json.loads(base64.urlsafe_b64decode(payload_b64).decode('utf-8'))
            exp = payload.get('exp', 0)
            import time as _t
            now = int(_t.time())
            valid = exp > now
            return jsonify({
                'ok': valid,
                'expired': not valid,
                'token_preview': tok[:40] + '...',
                'exp': exp,
                'expires_in_seconds': max(0, exp - now),
                'sub': payload.get('sub'),
            })
    except Exception:
        pass
    return jsonify({'ok': True, 'token_preview': tok[:40] + '...'})


if __name__ == '__main__':



    print('='*50)



    print('  SimplificaE Web App')



    print('  http://localhost:5000')



    print('='*50)



    _warmup_databricks()

    # Verificar atualizações em background (não bloqueia startup)
    if _UPDATER_OK:
        _updater.check_update_async()

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False, threaded=True)



